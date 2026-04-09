"""Enrich Assam 2026 constituencies with elector data from ECI final rolls.

Source: data/Assam_Final_Electors_10-02-2026.xlsx
Columns: Sl.No, District, AC No, AC Name, Polling Stations, Men, Women, Third Gender, Total
"""
import sys
import os
from datetime import date
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from app.database import engine
from app.models import Election, Constituency

PROJECT_ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
ELECTOR_FILE = os.path.join(PROJECT_ROOT, "data", "Assam_Final_Electors_10-02-2026.xlsx")


def run():
    wb = openpyxl.load_workbook(ELECTOR_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find header row
    data_rows = []
    for row in rows:
        if row[2] is not None and isinstance(row[2], (int, float)) and row[4] is not None:
            data_rows.append(row)

    print(f"Loaded {len(data_rows)} AC rows from elector file")

    with Session(engine) as session:
        # Set election date
        election = session.query(Election).filter(
            Election.state == "Assam", Election.year == 2026
        ).first()

        if not election:
            print("ERROR: No Assam 2026 election found")
            return

        election.election_date = date(2026, 4, 9)
        print(f"Set election_date = 2026-04-09 for {election.name}")

        # Update constituencies
        updated = 0
        for row in data_rows:
            ac_no = int(row[2])
            polling_stations = int(row[4]) if row[4] else None
            men = int(row[5]) if row[5] else None
            women = int(row[6]) if row[6] else None
            third_gender = int(row[7]) if row[7] else 0
            total = int(row[8]) if row[8] else None

            constituency = session.query(Constituency).filter(
                Constituency.election_id == election.id,
                Constituency.ac_no == ac_no,
            ).first()

            if constituency:
                constituency.total_electors = total
                constituency.male_electors = men
                constituency.female_electors = women
                constituency.third_gender_electors = third_gender
                constituency.total_polling_stations = polling_stations
                updated += 1
            else:
                print(f"  WARNING: AC {ac_no} not found in DB")

        session.commit()
        print(f"Updated {updated} constituencies with elector data")

        # Summary
        total_electors = sum(
            c.total_electors or 0
            for c in session.query(Constituency).filter(Constituency.election_id == election.id).all()
        )
        total_ps = sum(
            c.total_polling_stations or 0
            for c in session.query(Constituency).filter(Constituency.election_id == election.id).all()
        )
        print(f"\nTotal electors: {total_electors:,}")
        print(f"Total polling stations: {total_ps:,}")

        # Sample
        print("\nSample:")
        for c in session.query(Constituency).filter(Constituency.election_id == election.id).order_by(Constituency.ac_no).limit(5).all():
            print(f"  AC {c.ac_no} {c.name}: {c.total_electors:,} electors ({c.male_electors:,} M, {c.female_electors:,} F, {c.third_gender_electors} TG), {c.total_polling_stations} PS")


if __name__ == "__main__":
    run()
