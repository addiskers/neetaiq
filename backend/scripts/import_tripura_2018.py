"""Import Tripura 2018 — all candidates from myneta CSV/JSON; mark winners from ECI file.

2018 has no per-candidate detailed results from ECI, so all 296 candidates
come from myneta. The List of Successful Candidates marks position=1 winners.

Sources (Tripura/Tripura_2018/):
  - Winners:      List of Successful Candidates.xlsx
  - Electors:     Electors & Voters Data.xlsx
  - Party list:   List of Participating Poltical Parties .xlsx
  - CSV:          Tripura_candidates_2018.csv
  - JSON:         2018_candidate_profile.json
"""
import sys, os, re, csv, json
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from app.database import engine, Base
from app.models.tripura import Election, District, Constituency, Party, Candidate

PROJECT_ROOT   = os.path.join(os.path.dirname(__file__), "..", "..")
DATA_DIR       = os.path.join(PROJECT_ROOT, "Tripura", "Tripura_2018")
WINNERS_FILE   = os.path.join(DATA_DIR, "List of Successful Candidates.xlsx")
ELECTORS_FILE  = os.path.join(DATA_DIR, "Electors & Voters Data.xlsx")
PARTY_LIST     = os.path.join(DATA_DIR, "List of Participating Poltical Parties .xlsx")
PROFILES_JSON  = os.path.join(DATA_DIR, "2018_candidate_profile.json")
CANDIDATES_CSV = os.path.join(DATA_DIR, "Tripura_candidates_2018.csv")
MAPPING_FILE   = os.path.join(PROJECT_ROOT, "Tripura", "Tripura_Election_Data_Mapping.xlsx")
GEOJSON_PATH   = os.path.join(PROJECT_ROOT, "frontend", "public", "tripura_AC.geojson")

PARTY_COLORS = {
    "BJP": "#FF9933", "INC": "#00BFFF", "CPI(M)": "#FF0000", "CPM": "#FF0000",
    "CPI": "#FF6600", "AITC": "#00FF00", "TMC": "#00FF00", "IPFT": "#FF8C00",
    "TMP": "#8B4513", "INPT": "#556B2F", "IND": "#808080", "NOTA": "#000000",
    "BSP": "#0000FF", "RSP": "#FF4444",
}

ABBR_ALIASES = {"CPM": "CPI(M)", "TMC": "AITC"}

# Spelling differences between myneta CSV constituency names and mapping/ECI names
CONSTITUENCY_ALIASES = {
    "TOWN BARDOWALI": "TOWN BORDOWALI",
    "MANDAI BAZAR":   "MANDAIBAZAR",
    "MANDAI BAZAR (ST)": "MANDAIBAZAR (ST)",
    "PECHARTHAL (ST)": "PENCHARTHAL (ST)",
}


def _strip_cat(name: str) -> str:
    return re.sub(r"\s*\((SC|ST)\)\s*$", "", name.strip(), flags=re.I).strip().upper()


def _norm_name(name: str) -> str:
    n = name.upper().strip()
    n = re.sub(r"\bADV\.?\b|\bPROF\.?\b|\bDR\.?\b|\bMR\.?\b|\bSMT\.?\b", "", n)
    n = re.sub(r"\bS/O\b.*", "", n)
    n = re.sub(r"[^A-Z\s]", "", n)
    return re.sub(r"\s+", " ", n).strip()


def load_ac_name_to_no():
    """2018 AC cols (3,4) + 2023 cols (6,7) as fallback → name: ac_no."""
    wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True)
    ws = wb["AC"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    name_to_no = {}
    for row in rows[2:]:
        for no_col, name_col in [(3, 4), (6, 7), (0, 1)]:
            if row[no_col] and row[name_col]:
                raw = str(row[name_col]).strip().upper()
                clean = _strip_cat(raw)
                name_to_no.setdefault(raw, int(row[no_col]))
                name_to_no.setdefault(clean, int(row[no_col]))
    return name_to_no


def load_ac_category_map():
    """2018 AC col (3,4) → ac_no: category."""
    wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True)
    ws = wb["AC"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    cat_map = {}
    for row in rows[2:]:
        if not row[3] or not row[4]:
            continue
        try:
            ac_no = int(row[3])
        except (ValueError, TypeError):
            continue
        name = str(row[4]).strip()
        cat = "GEN"
        if "(SC)" in name:
            cat = "SC"
        elif "(ST)" in name:
            cat = "ST"
        cat_map[ac_no] = cat
    return cat_map


def load_ac_name_map():
    """2018 mapping → ac_no: clean name."""
    wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True)
    ws = wb["AC"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    name_map = {}
    for row in rows[2:]:
        if not row[3] or not row[4]:
            continue
        ac_no = int(row[3])
        name_map[ac_no] = _strip_cat(str(row[4]).strip())
    return name_map


def load_ac_district_from_winners():
    """Use ECI winners file to build ac_no → district name."""
    # Winners cols: STATE(0), DISTRICT(1), AC NO(2), AC NAME(3), WINNER(4), SEX(5), PARTY(6)
    wb = openpyxl.load_workbook(WINNERS_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    ac_dist = {}
    for row in rows[1:]:
        if not row[2]:
            continue
        try:
            ac_no = int(row[2])
        except (ValueError, TypeError):
            continue
        ac_dist[ac_no] = str(row[1]).strip().title() if row[1] else "Unknown"
    return ac_dist


def load_electors_map():
    wb = openpyxl.load_workbook(ELECTORS_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    out = {}
    for row in rows[1:]:
        if row[2] is None:
            continue
        try:
            ac_no = int(row[2])
        except (ValueError, TypeError):
            continue
        out[ac_no] = {
            "male":         int(row[4])  if row[4]  else 0,
            "female":       int(row[5])  if row[5]  else 0,
            "third_gender": int(row[6])  if row[6]  else 0,
            "total":        int(row[7])  if row[7]  else 0,
            "total_votes":  int(row[13]) if row[13] else 0,
        }
    return out


def load_party_names():
    wb = openpyxl.load_workbook(PARTY_LIST, read_only=True)
    ws = wb[wb.sheetnames[0]]
    mapping = {}
    for row in ws.iter_rows(values_only=True):
        if isinstance(row[1], (int, float)) and row[2] and row[3]:
            abbr = str(row[2]).strip()
            canon = ABBR_ALIASES.get(abbr, abbr)
            mapping[canon] = str(row[3]).strip()
    wb.close()
    return mapping


def load_winners():
    """ECI winners → ac_no: winner name upper."""
    wb = openpyxl.load_workbook(WINNERS_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    winners = {}
    for row in rows[1:]:
        if not row[2]:
            continue
        try:
            ac_no = int(row[2])
        except (ValueError, TypeError):
            continue
        winners[ac_no] = str(row[4]).strip().upper() if row[4] else ""
    return winners


def resolve_ac_no(con_raw: str, name_to_no: dict) -> int | None:
    raw = con_raw.strip().upper()
    clean = _strip_cat(raw)
    alias = CONSTITUENCY_ALIASES.get(raw, CONSTITUENCY_ALIASES.get(clean))
    alias_clean = _strip_cat(alias) if alias else None
    for key in [raw, clean, alias, alias_clean]:
        if key and key in name_to_no:
            return name_to_no[key]
    return None


def _parse_rupees(val):
    if not val or "Nil" in str(val):
        return 0
    s = str(val).strip()
    if s.startswith("http"):
        return 0
    digits = re.sub(r"[^\d]", "", s.split("~")[0])
    return int(digits) if digits else 0


def _parse_criminal(status):
    if not status or "No criminal" in str(status):
        return 0
    m = re.search(r"(\d+)", str(status))
    return int(m.group(1)) if m else 0


def load_myneta_candidates():
    csv_by_id = {}
    with open(CANDIDATES_CSV, encoding="latin-1") as f:
        for row in csv.DictReader(f):
            cid = row.get("candidate_id", "").strip()
            if cid:
                csv_by_id[cid] = row

    with open(PROFILES_JSON, encoding="utf-8") as f:
        profiles = json.load(f)

    candidates = []
    for p in profiles:
        cp = p.get("candidate_profile", {})
        cid = str(cp.get("candidate_id", ""))
        csv_r = csv_by_id.get(cid, {})

        con_raw = (csv_r.get("constituency") or cp.get("ac_name") or "").strip()
        name = (csv_r.get("candidate_name") or cp.get("name") or "").strip()
        party_abbr = (csv_r.get("party") or cp.get("party") or "").strip()
        party_abbr = ABBR_ALIASES.get(party_abbr, party_abbr)

        edu = (csv_r.get("education") or cp.get("education") or "").strip()
        if edu.startswith("Category:"):
            edu = edu[len("Category:"):].strip()

        assets_s = p.get("assets_summary", {})
        assets_raw = csv_r.get("assets", "")
        assets_val = _parse_rupees(assets_raw) if assets_raw and not str(assets_raw).startswith("http") else _parse_rupees(assets_s.get("total_assets"))

        liab_raw = csv_r.get("liabilities", "")
        liab_val = _parse_rupees(liab_raw) if liab_raw and not str(liab_raw).startswith("http") else _parse_rupees(assets_s.get("total_liabilities"))

        age_raw = cp.get("age", "")
        try:
            age = int(str(age_raw).strip())
        except (ValueError, TypeError):
            age = None

        candidates.append({
            "constituency": con_raw,
            "name": name,
            "party_abbr": party_abbr,
            "education": edu if edu and edu not in ("Not mentioned", "Not Given", "") else None,
            "occupation": (p.get("profession", {}).get("self") or "").strip() or None,
            "declared_assets": assets_val if assets_val > 0 else None,
            "liabilities": liab_val if liab_val > 0 else None,
            "criminal_cases": int(csv_r.get("criminal_cases") or 0) or _parse_criminal(cp.get("crime_status")),
            "image_url": (cp.get("image_url") or "").strip() or None,
            "age": age,
            "gender": cp.get("gender", "").strip().upper() or None,
        })
    return candidates


def import_data():
    Base.metadata.create_all(bind=engine)

    name_to_no  = load_ac_name_to_no()
    cat_map     = load_ac_category_map()
    ac_name_map = load_ac_name_map()
    ac_dist_map = load_ac_district_from_winners()
    electors    = load_electors_map()
    party_names = load_party_names()
    winners     = load_winners()
    myneta_cands = load_myneta_candidates()

    with Session(engine) as session:
        election = Election(
            state="Tripura", year=2018, type="Assembly",
            name="Tripura Legislative Assembly Election 2018",
        )
        session.add(election)
        session.flush()
        print(f"Created election: {election.name} (id={election.id})")

        party_cache = {p.abbr.upper(): p for p in session.query(Party).all()}
        for abbr, full_name in party_names.items():
            key = abbr.upper()
            if key not in party_cache:
                p = Party(name=full_name, abbr=abbr[:50], color=PARTY_COLORS.get(abbr))
                session.add(p)
                session.flush()
                party_cache[key] = p

        district_cache = {}
        constituency_cache = {}

        for ac_no in sorted(ac_name_map.keys()):
            dist_name = ac_dist_map.get(ac_no, "Unknown")
            if dist_name not in district_cache:
                d = District(election_id=election.id, name=dist_name)
                session.add(d)
                session.flush()
                district_cache[dist_name] = d

            el = electors.get(ac_no, {})
            total_votes    = el.get("total_votes")
            total_electors = el.get("total")
            turnout_pct    = round(total_votes / total_electors * 100, 2) if total_electors and total_votes else None

            con = Constituency(
                election_id=election.id,
                district_id=district_cache[dist_name].id,
                ac_no=ac_no, name=ac_name_map[ac_no],
                category=cat_map.get(ac_no, "GEN"),
                total_electors=total_electors,
                male_electors=el.get("male"), female_electors=el.get("female"),
                third_gender_electors=el.get("third_gender"),
                total_votes_polled=total_votes, turnout_pct=turnout_pct,
            )
            session.add(con)
            session.flush()
            constituency_cache[ac_no] = con

        print(f"Created {len(constituency_cache)} constituencies, {len(district_cache)} districts")

        ac_cands: dict[int, list] = {}
        skipped = 0
        for cand in myneta_cands:
            ac_no = resolve_ac_no(cand["constituency"], name_to_no)
            if ac_no is None:
                skipped += 1
                continue
            ac_cands.setdefault(ac_no, []).append(cand)

        if skipped:
            print(f"Warning: {skipped} candidates could not be matched to an AC")

        total_inserted = 0
        for ac_no, cands in ac_cands.items():
            con = constituency_cache.get(ac_no)
            if not con:
                continue

            winner_name = winners.get(ac_no, "")
            winner_norm = _norm_name(winner_name)

            def is_winner(c, _wn=winner_name, _wno=winner_norm):
                n = c["name"].upper()
                return n == _wn or (_wno and _norm_name(n) == _wno)

            sorted_cands = sorted(cands, key=lambda c: (0 if is_winner(c) else 1))

            for cand in sorted_cands:
                position = 1 if is_winner(cand) else 2
                party_abbr = cand["party_abbr"][:50] if cand["party_abbr"] else None
                party = None
                if party_abbr:
                    key = party_abbr.upper()
                    if key not in party_cache:
                        p = Party(name=party_abbr, abbr=party_abbr, color=PARTY_COLORS.get(party_abbr))
                        session.add(p)
                        session.flush()
                        party_cache[key] = p
                    party = party_cache[key]

                gender_raw = (cand.get("gender") or "").upper()
                gender = "MALE" if gender_raw in ("MALE", "M") else ("FEMALE" if gender_raw in ("FEMALE", "F") else (gender_raw or None))

                occ = cand.get("occupation") or None
                if occ and occ == "Not mentioned":
                    occ = None

                session.add(Candidate(
                    election_id=election.id, constituency_id=con.id,
                    party_id=party.id if party else None,
                    name=cand["name"], gender=gender, age=cand.get("age"),
                    position=position,
                    votes_general=None, votes_postal=None, votes_total=None, vote_pct=None,
                    is_nota=False,
                    education=cand.get("education"),
                    occupation=occ,
                    declared_assets=cand.get("declared_assets"),
                    liabilities=cand.get("liabilities"),
                    criminal_cases=cand.get("criminal_cases", 0),
                    image_url=cand.get("image_url"),
                ))
                total_inserted += 1

        session.commit()
        winners_in_db = session.query(Candidate).filter_by(election_id=election.id, position=1).count()
        print(f"\n=== Import Summary ===")
        print(f"Candidates inserted: {total_inserted}, Winners (position=1): {winners_in_db}")


if __name__ == "__main__":
    import_data()
