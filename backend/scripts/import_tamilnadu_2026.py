"""Import Tamil Nadu 2026 nomination data (no results yet).

Data sources (all under tamil_nadu/Tamil Nadu_2026/):
  - Electors: District wise AC wise electors_ceo_2026.xlsx
  - Candidate profiles: my_neta_2026_candidates_profile.json
  - Candidate bridge: my_neta_tamilnadu_candidates_2026.csv
  - Affidavit list: affidavit_candidate_list_TamilNadu_2026.xlsx

Usage: cd backend && python -m scripts.import_tamilnadu_2026
"""
import sys, os, re, json, csv
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from app.database import engine, Base
from app.models.tamilnadu import Election, District, Constituency, Party, Candidate

PROJECT_ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
DATA_DIR = os.path.join(PROJECT_ROOT, "tamil_nadu", "Tamil Nadu_2026")

ELECTORS_FILE  = os.path.join(DATA_DIR, "District wise AC wise electors_ceo_2026.xlsx")
PROFILES_JSON  = os.path.join(DATA_DIR, "my_neta_2026_candidates_profile.json")
CANDIDATES_CSV = os.path.join(DATA_DIR, "my_neta_tamilnadu_candidates_2026.csv")
AFFIDAVIT_FILE = os.path.join(DATA_DIR, "affidavit_candidate_list_TamilNadu_2026.xlsx")

PARTY_COLORS = {
    "DMK": "#FF0000", "AIADMK": "#006400", "INC": "#00BFFF",
    "BJP": "#FF9933", "CPI(M)": "#FF0000", "CPI": "#FF4444",
    "PMK": "#FFFF00", "AITC": "#00FF00", "MDMK": "#FF6600",
    "VCK": "#0000FF", "DMDK": "#800080", "IND": "#808080",
}

PARTY_ABBR = {
    "Bharatiya Janata Party": "BJP",
    "Indian National Congress": "INC",
    "Dravida Munnetra Kazhagam": "DMK",
    "All India Anna Dravida Munnetra Kazhagam": "AIADMK",
    "Pattali Makkal Katchi": "PMK",
    "Communist Party of India (Marxist)": "CPI(M)",
    "Communist Party of India": "CPI",
    "Bahujan Samaj Party": "BSP",
    "Aam Aadmi Party": "AAP",
    "Nationalist Congress Party": "NCP",
    "All India Trinamool Congress": "AITC",
    "Marumalarchi Dravida Munnetra Kazhagam": "MDMK",
    "Viduthalai Chiruthaigal Katchi": "VCK",
    "Desiya Murpokku Dravida Kazhagam": "DMDK",
    "Independent": "IND",
}

ABBR_ALIASES = {"CPM": "CPI(M)", "ADMK": "AIADMK"}


def parse_rupees(val):
    if not val or "Nil" in str(val):
        return 0
    cleaned = re.sub(r"[Rs\s\xa0,~]", "", str(val))
    digits = re.match(r"^([\d.]+)", cleaned)
    return int(float(digits.group(1))) if digits else 0


def parse_criminal(status):
    if not status or "No criminal" in status:
        return 0
    m = re.search(r"(\d+)", status)
    return int(m.group(1)) if m else 0


def safe_int_comma(val):
    """Parse integers that may be formatted as comma strings like '1,23,359'."""
    if val is None:
        return None
    try:
        return int(str(val).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def load_electors():
    # Cols: SL.No(0), District No(1), District Name(2), AC NO(3), Name of Assembly(4),
    #       Male(5), Female(6), Third Gender(7), Total(8)
    # Numbers are comma strings like '1,23,359'
    wb = openpyxl.load_workbook(ELECTORS_FILE, read_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()
    em = {}
    for row in rows[1:]:
        if not any(row):
            continue
        ac_no_raw = row[3]
        if ac_no_raw is None:
            continue
        try:
            ac_no = int(str(ac_no_raw).replace(",", "").strip())
        except (TypeError, ValueError):
            continue

        district_name = re.sub(r"\s+", " ", str(row[2]).strip()).upper() if row[2] else "UNKNOWN"
        ac_name = re.sub(r"\s+", " ", str(row[4]).strip()).upper() if row[4] else f"AC_{ac_no}"

        em[ac_no] = {
            "district": district_name,
            "name": ac_name,
            "category": "GEN",  # No category column in 2026 electors file
            "male_electors": safe_int_comma(row[5]),
            "female_electors": safe_int_comma(row[6]),
            "third_gender_electors": safe_int_comma(row[7]),
            "total_electors": safe_int_comma(row[8]),
        }
    return em


def run():
    Base.metadata.create_all(bind=engine)

    em = load_electors()
    print(f"Loaded {len(em)} ACs from electors file")

    # Load AC category from 2021 constituencies (categories don't change between elections)
    from app.models.tamilnadu import Constituency as TNCon
    ac_categories = {}
    with Session(engine) as session:
        prev = session.query(Election).filter_by(state="Tamil Nadu", year=2021).first()
        if prev:
            for c in session.query(TNCon).filter_by(election_id=prev.id).all():
                ac_categories[c.ac_no] = c.category

    if not os.path.exists(PROFILES_JSON):
        print(f"WARNING: Profiles JSON not found at {PROFILES_JSON}")
        profiles = []
    else:
        with open(PROFILES_JSON, "r", encoding="utf-8") as f:
            profiles = json.load(f)
    print(f"Loaded {len(profiles)} candidate profiles")

    with Session(engine) as session:
        if session.query(Election).filter_by(state="Tamil Nadu", year=2026).first():
            print("Tamil Nadu 2026 already exists — skipping.")
            return

        election = Election(state="Tamil Nadu", year=2026, type="Assembly",
                            name="Tamil Nadu Legislative Assembly Election 2026")
        session.add(election)
        session.flush()

        # Collect unique districts from electors file
        district_names = sorted(set(
            info["district"] for info in em.values()
            if info.get("district") and info["district"] not in ("UNKNOWN", "")
        ))
        if not district_names:
            district_names = ["Tamil Nadu"]
        district_cache = {dname: District(election_id=election.id, name=dname)
                          for dname in district_names}
        for d in district_cache.values():
            session.add(d)
        session.flush()
        default_district = list(district_cache.values())[0]

        # Create constituencies from electors file
        constituency_cache = {}
        for ac_no, info in em.items():
            dist_name = info.get("district", "")
            dist = district_cache.get(dist_name, default_district)
            con = Constituency(
                election_id=election.id,
                district_id=dist.id if dist else default_district.id,
                ac_no=ac_no,
                name=info["name"],
                category=info.get("category") or ac_categories.get(ac_no, "GEN"),
                total_electors=info["total_electors"],
                male_electors=info["male_electors"],
                female_electors=info["female_electors"],
                third_gender_electors=info["third_gender_electors"],
            )
            session.add(con)
            constituency_cache[ac_no] = con
        session.flush()
        print(f"Created {len(constituency_cache)} constituencies")

        party_cache = {p.abbr.upper(): p for p in session.query(Party).all()}
        imported = skipped = 0

        # Build ac_name -> ac_no lookup
        name_to_ac = {v["name"]: k for k, v in em.items()}

        for cand_data in profiles:
            cp = cand_data.get("candidate_profile", {})
            ac_name_raw = (cp.get("ac_name") or "").strip().upper()
            ac_name = re.sub(r"\s+", " ", ac_name_raw)

            # Find constituency by name
            con = None
            ac_no_match = name_to_ac.get(ac_name)
            if ac_no_match and ac_no_match in constituency_cache:
                con = constituency_cache[ac_no_match]
            else:
                # Fuzzy match: look for partial match
                for ac_no, c in constituency_cache.items():
                    if c.name == ac_name:
                        con = c
                        break
            if not con:
                print(f"  WARNING: No constituency match for '{ac_name}'")
                skipped += 1
                continue

            # Party
            raw_party = (cp.get("party") or "").strip().rstrip()
            party = None
            if raw_party and raw_party.upper() not in ("NOTA", ""):
                abbr = ABBR_ALIASES.get(raw_party, PARTY_ABBR.get(raw_party, raw_party[:20].upper()))
                if abbr.upper() not in party_cache:
                    p = Party(name=raw_party, abbr=abbr, color=PARTY_COLORS.get(abbr))
                    session.add(p)
                    session.flush()
                    party_cache[abbr.upper()] = p
                party = party_cache[abbr.upper()]

            # Name
            name = (cp.get("name") or "").strip().rstrip()

            # Age
            age_str = str(cp.get("age") or "").strip()
            age = int(age_str) if age_str.isdigit() else None

            # Education
            edu = (cp.get("education") or "").strip()
            if edu.startswith("Category:"):
                edu = edu[len("Category:"):].strip()
            education = edu if edu and edu not in ("Not mentioned", "Not Given", "Not Given ") else None

            occupation = (cand_data.get("profession", {}).get("self") or "").strip() or None

            assets_summary = cand_data.get("assets_summary", {})
            declared_assets = parse_rupees(assets_summary.get("total_assets"))
            liabilities_val = parse_rupees(assets_summary.get("total_liabilities"))

            criminal_cases = parse_criminal(cp.get("crime_status"))
            image_url = (cp.get("image_url") or "").strip() or None

            gender_raw = str(cp.get("gender") or "").strip().upper()
            gender = "M" if gender_raw in ("M", "MALE") else ("F" if gender_raw in ("F", "FEMALE") else None)

            session.add(Candidate(
                election_id=election.id, constituency_id=con.id,
                party_id=party.id if party else None,
                name=name or f"CANDIDATE_{cp.get('candidate_id', '')}",
                gender=gender,
                age=age, education=education, occupation=occupation,
                declared_assets=declared_assets if declared_assets > 0 else None,
                liabilities=liabilities_val if liabilities_val > 0 else None,
                criminal_cases=criminal_cases, image_url=image_url,
                is_nota=False,
            ))
            imported += 1

        session.commit()

    print(f"\n=== Tamil Nadu 2026 ===")
    print(f"Constituencies: {len(constituency_cache)}")
    print(f"Candidates: {imported} imported, {skipped} skipped")
    total_electors = sum(v["total_electors"] or 0 for v in em.values())
    print(f"Total electors: {total_electors:,}")


if __name__ == "__main__":
    run()
