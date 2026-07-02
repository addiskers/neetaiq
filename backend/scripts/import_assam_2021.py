"""Import Assam 2021 election data from Excel files into netaiq database.

Data sources:
  - District mapping: Assam District and AC name (2016,2021,2026).xlsx (2021 cols 0-3)
  - Results: data/10. Detailed Results.xlsx
  - Party names: assam2021/3. List Of Political Parties Participated.xlsx
"""
import sys
import os
import re
import csv
import json
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from app.database import engine, Base
from app.models.assam import Election, District, Constituency, Party, Candidate

PROJECT_ROOT   = os.path.join(os.path.dirname(__file__), "..", "..")
MAPPING_FILE   = os.path.join(PROJECT_ROOT, "Assam District and AC name (2016,2021,2026).xlsx")
DETAILED_RESULTS = os.path.join(PROJECT_ROOT, "data", "10. Detailed Results.xlsx")
PARTY_LIST     = os.path.join(PROJECT_ROOT, "assam2021", "3. List Of Political Parties Participated.xlsx")
PROFILES_JSON  = os.path.join(PROJECT_ROOT, "assam2021", "structured_output_2021.json")
CANDIDATES_CSV = os.path.join(PROJECT_ROOT, "assam2021", "assam_candidates_2021.csv")

PARTY_COLORS = {
    "BJP": "#FF9933", "INC": "#00BFFF", "AIUDF": "#006400", "AGP": "#FFFFFF",
    "BOPF": "#228B22", "UPPL": "#FFD700", "CPI(M)": "#FF0000", "CPI": "#FF0000",
    "AITC": "#00FF00", "NCP": "#004080", "JD(U)": "#003366", "SP": "#FF0000",
    "RJD": "#00A300", "LJP": "#0000FF", "IND": "#808080", "NOTA": "#000000",
}


def load_ac_mapping():
    """Parse 2021 columns from the district-AC mapping file.

    Returns dict: ac_no → {district, name, category}

    File layout (2021 is cols 0-3):
      Col 0: Sl.No (district number, only on first row of each district group)
      Col 1: District name (only on first row of each district group)
      Col 2: AC number
      Col 3: AC name (may have "(SC)" or "(ST)" suffix for category)
    """
    wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    ac_map = {}
    current_district = None

    for row in rows[2:]:  # skip header rows
        # Update district when a new one appears
        if row[1]:
            current_district = str(row[1]).strip().title()  # normalize: "CACHAR" → "Cachar"

        # Parse AC if present
        if row[2] is not None and row[3]:
            ac_no = int(row[2])
            raw_name = str(row[3]).strip()

            # Extract category from suffix
            category = "GEN"
            if "(SC)" in raw_name:
                category = "SC"
                raw_name = raw_name.replace("(SC)", "").strip()
            elif "(ST)" in raw_name:
                category = "ST"
                raw_name = raw_name.replace("(ST)", "").strip()

            ac_map[ac_no] = {
                "district": current_district,
                "name": raw_name,
                "category": category,
            }

    return ac_map


def load_party_names():
    """Load abbreviation → full name mapping from party list Excel."""
    if not os.path.exists(PARTY_LIST):
        print(f"  Party list not found at {PARTY_LIST}, using abbreviations as names")
        return {}
    wb = openpyxl.load_workbook(PARTY_LIST, read_only=True)
    ws = wb[wb.sheetnames[0]]
    mapping = {}
    for row in ws.iter_rows(values_only=True):
        if isinstance(row[0], (int, float)) and row[1] and row[2]:
            mapping[row[1].strip()] = row[2].strip()
    wb.close()
    return mapping


def parse_candidate_name(raw_name):
    """Strip position prefix: '1 BIJOY MALAKAR' → 'BIJOY MALAKAR'."""
    match = re.match(r"^\d+\s+(.+)$", raw_name.strip())
    return match.group(1).strip() if match else raw_name.strip()


def import_data():
    Base.metadata.create_all(bind=engine)

    ac_map = load_ac_mapping()
    print(f"Loaded mapping: {len(ac_map)} ACs across {len(set(a['district'] for a in ac_map.values()))} districts")

    party_names = load_party_names()

    # Load detailed results
    wb = openpyxl.load_workbook(DETAILED_RESULTS, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    data_rows = rows[4:]  # skip header rows

    with Session(engine) as session:
        # 1. Create election
        election = Election(
            state="Assam",
            year=2021,
            type="Assembly",
            name="Assam Legislative Assembly Election 2021",
        )
        session.add(election)
        session.flush()
        print(f"Created election: {election.name} (id={election.id})")

        # 2. Create parties from party list (reuse existing ones)
        party_cache = {p.abbr.upper(): p for p in session.query(Party).all()}
        for abbr, full_name in party_names.items():
            if abbr.upper() not in party_cache:
                party = Party(name=full_name, abbr=abbr, color=PARTY_COLORS.get(abbr))
                session.add(party)
                session.flush()
                party_cache[abbr.upper()] = party
        print(f"Created {len(party_cache)} parties from party list")

        # 3. Create districts from mapping
        district_cache = {}
        for ac_info in ac_map.values():
            dname = ac_info["district"]
            if dname not in district_cache:
                district = District(election_id=election.id, name=dname)
                session.add(district)
                district_cache[dname] = district
        session.flush()
        print(f"Created {len(district_cache)} districts")

        # 4. Parse detailed results — group by AC
        current_ac_no = None
        current_ac_name = None
        current_electors = None
        ac_candidates = []
        ac_turnout = None

        def flush_ac():
            nonlocal current_ac_no, current_ac_name, current_electors, ac_candidates, ac_turnout
            if current_ac_no is None:
                return

            mapping_info = ac_map.get(current_ac_no, {})
            district_name = mapping_info.get("district")
            district = district_cache.get(district_name)

            if not district:
                print(f"  WARNING: No district mapping for AC {current_ac_no} ({current_ac_name})")

            # Turnout
            total_votes = ac_turnout[2] if ac_turnout else None
            turnout_pct = ac_turnout[3] if ac_turnout else None

            # Winning margin = 1st place votes - 2nd place votes
            winning_margin = None
            if len(ac_candidates) >= 2:
                v1 = ac_candidates[0][8]  # votes_total of position 1
                v2 = ac_candidates[1][8]  # votes_total of position 2
                if v1 and v2:
                    winning_margin = v1 - v2

            constituency = Constituency(
                election_id=election.id,
                district_id=district.id if district else None,
                ac_no=current_ac_no,
                name=mapping_info.get("name", current_ac_name),
                category=mapping_info.get("category"),
                total_electors=current_electors,
                total_votes_polled=total_votes,
                turnout_pct=turnout_pct,
                winning_margin=winning_margin,
            )
            session.add(constituency)
            session.flush()

            for pos, cand_row in enumerate(ac_candidates, start=1):
                raw_name, sex, age, _, party_abbr, _, gen_votes, postal_votes, total_votes_c, vote_pct = cand_row
                name = parse_candidate_name(raw_name)
                is_nota = (party_abbr == "NOTA")

                party = None
                if not is_nota and party_abbr:
                    key = party_abbr.upper()
                    if key not in party_cache:
                        new_party = Party(
                            name=party_abbr,
                            abbr=party_abbr,
                            color=PARTY_COLORS.get(party_abbr),
                        )
                        session.add(new_party)
                        session.flush()
                        party_cache[key] = new_party
                    party = party_cache[key]

                session.add(Candidate(
                    election_id=election.id,
                    constituency_id=constituency.id,
                    party_id=party.id if party else None,
                    name=name,
                    gender=sex,
                    age=age,
                    position=pos,
                    votes_general=gen_votes,
                    votes_postal=postal_votes,
                    votes_total=total_votes_c,
                    vote_pct=vote_pct,
                    is_nota=is_nota,
                ))

            current_ac_no = None
            ac_candidates = []
            ac_turnout = None

        for row in data_rows:
            state_col = row[0]
            ac_no = row[1]
            ac_name = row[2]
            candidate_name = row[3]

            if state_col is None or state_col in ("GRAND TOTAL:", "Disclaimer"):
                continue

            if state_col == "TURN OUT":
                ac_turnout = (
                    row[9] if row[9] else 0,
                    row[10] if row[10] else 0,
                    row[11] if row[11] else 0,
                    row[12] if row[12] else 0,
                )
                flush_ac()
                continue

            if ac_no is not None and ac_no != current_ac_no:
                flush_ac()
                current_ac_no = int(ac_no)
                current_ac_name = ac_name.strip() if ac_name else None
                current_electors = int(row[13]) if row[13] else None
                ac_candidates = []
                ac_turnout = None

            if candidate_name:
                ac_candidates.append((
                    str(candidate_name),
                    row[4],                                    # sex
                    int(row[5]) if row[5] else None,           # age
                    row[6],                                    # category (unused, from mapping instead)
                    str(row[7]) if row[7] else None,           # party abbr
                    row[8],                                    # symbol
                    int(row[9]) if row[9] else 0,              # general votes
                    int(row[10]) if row[10] else 0,            # postal votes
                    int(row[11]) if row[11] else 0,            # total votes
                    float(row[12]) if row[12] else 0,          # vote %
                ))

        flush_ac()
        session.commit()

        # Summary
        n_elections = session.query(Election).count()
        n_districts = session.query(District).count()
        n_constituencies = session.query(Constituency).count()
        n_parties = session.query(Party).count()
        n_candidates = session.query(Candidate).count()
        n_nota = session.query(Candidate).filter(Candidate.is_nota == True).count()

        print(f"\n=== Import Summary ===")
        print(f"Elections:      {n_elections}")
        print(f"Districts:      {n_districts}")
        print(f"Constituencies: {n_constituencies}")
        print(f"Parties:        {n_parties}")
        print(f"Candidates:     {n_candidates} ({n_nota} NOTA)")
        print(f"Real candidates: {n_candidates - n_nota}")

        # District breakdown
        print(f"\n=== District Breakdown ===")
        for d in session.query(District).order_by(District.name).all():
            count = session.query(Constituency).filter(Constituency.district_id == d.id).count()
            print(f"  {d.name}: {count} ACs")


def _parse_rupees(val):
    if not val or "Nil" in str(val): return 0
    s = str(val).strip()
    if s.startswith("http"): return 0
    digits = re.sub(r"[^\d]", "", s.split("~")[0])
    return int(digits) if digits else 0


def _parse_criminal(status):
    if not status or "No criminal" in str(status): return 0
    m = re.search(r"(\d+)", str(status))
    return int(m.group(1)) if m else 0


def _expand_ac(name: str) -> str:
    """Expand truncated AC name suffixes: ' EA' → ' EAST', ' WE' → ' WEST'."""
    if name.endswith(" EA"):  return name[:-3] + " EAST"
    if name.endswith(" WE"):  return name[:-3] + " WEST"
    return name


def load_profiles():
    """Load 2021 candidate profiles indexed by (constituency_upper, name_upper)."""
    csv_by_id = {}
    with open(CANDIDATES_CSV, encoding="latin-1") as f:
        for row in csv.DictReader(f):
            cid = row.get("candidate_id", "").strip()
            con = re.sub(r"\s*\(?(SC|ST)\)?\s*$", "", row.get("constituency", "").strip(), flags=re.I).upper()
            con = _expand_ac(con)  # fix truncated names: BILASIPARA EA → BILASIPARA EAST
            if cid:
                csv_by_id[cid] = {
                    "name":           row.get("candidate_name", "").strip(),
                    "constituency":   con,
                    "criminal_cases": int(row.get("criminal_cases") or 0),
                    "education":      row.get("education", "").strip(),
                    "assets":         row.get("assets", ""),
                    "liabilities":    row.get("liabilities", ""),
                }

    with open(PROFILES_JSON, encoding="utf-8") as f:
        profiles_json = json.load(f)

    by_key = {}
    for p in profiles_json:
        cp  = p.get("candidate_profile", {})
        cid = str(cp.get("candidate_id", ""))
        info = csv_by_id.get(cid)
        if info:
            key = (info["constituency"], info["name"].upper())
            by_key[key] = {**p, "_csv": info}
    return by_key


def enrich_2021_profiles():
    """Update already-imported 2021 candidates with affidavit/profile data."""
    profile_map = load_profiles()
    print(f"Loaded {len(profile_map)} profiles for 2021")

    with Session(engine) as session:
        el = session.query(Election).filter_by(state="Assam", year=2021).first()
        if not el:
            print("Assam 2021 not found in DB — run import_data() first")
            return

        updated = missing = 0
        for con in session.query(Constituency).filter_by(election_id=el.id).all():
            ac_upper = con.name.upper()
            for cand in session.query(Candidate).filter_by(constituency_id=con.id, is_nota=False).all():
                key = (ac_upper, cand.name.upper())
                prof = profile_map.get(key)
                if not prof:
                    missing += 1
                    continue

                cp       = prof.get("candidate_profile", {})
                csv_info = prof.get("_csv", {})
                assets_s = prof.get("assets_summary", {})

                edu = (csv_info.get("education") or cp.get("education") or "").strip()
                if edu.startswith("Category:"): edu = edu[len("Category:"):].strip()
                cand.education = edu if edu and edu not in ("Not mentioned", "Not Given", "") else None

                cand.criminal_cases = int(csv_info.get("criminal_cases") or 0) or _parse_criminal(cp.get("crime_status"))

                assets_raw = csv_info.get("assets", "")
                # CSV may contain a URL (image endpoint) instead of text — fall back to JSON
                if not assets_raw or str(assets_raw).startswith("http"):
                    assets_val = _parse_rupees(assets_s.get("total_assets"))
                else:
                    assets_val = _parse_rupees(assets_raw)
                cand.declared_assets = assets_val if assets_val > 0 else None

                liab_raw = csv_info.get("liabilities", "")
                if not liab_raw or str(liab_raw).startswith("http"):
                    liab_val = _parse_rupees(assets_s.get("total_liabilities"))
                else:
                    liab_val = _parse_rupees(liab_raw)
                cand.liabilities = liab_val if liab_val > 0 else None

                occ = (prof.get("profession", {}).get("self") or "").strip()
                cand.occupation = occ if occ and occ != "Not mentioned" else None

                img = (cp.get("image_url") or "").strip()
                cand.image_url = img if img and img != "None" else None

                updated += 1

        session.commit()
        print(f"Enriched {updated} candidates, {missing} not matched")


if __name__ == "__main__":
    import_data()
    enrich_2021_profiles()
