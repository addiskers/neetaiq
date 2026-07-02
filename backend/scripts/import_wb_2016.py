"""Import West Bengal 2016 election data — detailed results + affidavit data.

Data sources (all under west_bengal_election_data 1/west_bengal_election_data/wb_2016/):
  - Results: Detailed Results_2016.xlsx
  - Electors: AC_wise_elector_2016.xlsx (district-AC mapping + elector breakdown)
  - Parties: List Of Political Parties Participated_2016.xlsx
  - Affidavit bridge: wb2016_all_candidates.xlsx
  - Affidavit data: 2016_candidate_profiles.json
"""
import sys
import os
import re
import json

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.database import engine, Base
from app.models.westbengal import Election, District, Constituency, Party, Candidate

PROJECT_ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
DATA_DIR = os.path.join(PROJECT_ROOT, "west_bengal_election_data 1", "west_bengal_election_data", "wb_2016")

RESULTS_FILE = os.path.join(DATA_DIR, "Detailed Results_2016.xlsx")
ELECTORS_FILE = os.path.join(DATA_DIR, "AC_wise_elector_2016.xlsx")
PARTIES_FILE = os.path.join(DATA_DIR, "List Of Political Parties Participated_2016.xlsx")
CANDIDATES_XLSX = os.path.join(DATA_DIR, "wb2016_all_candidates.xlsx")
PROFILES_JSON = os.path.join(DATA_DIR, "2016_candidate_profiles.json")

PARTY_COLORS = {
    "AITC": "#00FF00", "CPM": "#FF0000", "CPI(M)": "#FF0000", "INC": "#00BFFF",
    "BJP": "#FF9933", "AIFB": "#CC0000", "RSP": "#FF6600", "CPI": "#FF4444",
    "JD(U)": "#003366", "BSP": "#0000FF", "NCP": "#004080", "IND": "#808080",
    "NOTA": "#000000", "GJM": "#FFD700", "JMM": "#2E8B57", "SUCI": "#8B0000",
    "SDF": "#FF69B4", "RJD": "#00A300", "SP": "#FF0000",
}

ABBR_ALIASES = {
    "CPM": "CPI(M)",
}


def load_electors():
    """Parse AC_wise_elector_2016.xlsx -> {ac_no: {district, name, category, ...electors}}."""
    wb = openpyxl.load_workbook(ELECTORS_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    ac_map = {}
    for row in rows[1:]:
        district = str(row[0]).strip() if row[0] else None
        if not district or district == "State Total":
            continue
        ac_no = int(row[1]) if row[1] else None
        raw_name = str(row[2]).strip() if row[2] else None
        if not ac_no or not raw_name:
            continue

        category = "GEN"
        if "(SC)" in raw_name:
            category = "SC"
        elif "(ST)" in raw_name:
            category = "ST"
        clean_name = re.sub(r"\s*\((SC|ST)\)\s*$", "", raw_name).strip()

        ac_map[ac_no] = {
            "district": district,
            "name": clean_name,
            "category": category,
            "total_polling_stations": int(row[3]) if row[3] else None,
            "male_electors": int(row[4]) if row[4] else None,
            "female_electors": int(row[5]) if row[5] else None,
            "third_gender_electors": int(row[6]) if row[6] else None,
            "total_electors": int(row[7]) if row[7] else None,
        }
    return ac_map


def load_party_names():
    """Parse parties file -> {abbr: full_name}."""
    wb = openpyxl.load_workbook(PARTIES_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    party_map = {}
    for row in rows[1:]:
        full_name = str(row[1]).strip() if row[1] else None
        abbr = str(row[2]).strip() if row[2] else None
        if abbr and full_name:
            party_map[abbr] = full_name
    return party_map


def load_affidavit_bridge():
    """Build {(constituency_upper, candidate_name_upper): affidavit_data}."""
    # Load candidate xlsx for id -> (name, constituency) mapping
    wb = openpyxl.load_workbook(CANDIDATES_XLSX, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    csv_by_id = {}
    for row in rows[1:]:
        cid = str(int(row[1])) if row[1] else None
        name = str(row[2]).strip() if row[2] else ""
        constituency = str(row[4]).strip().upper() if row[4] else ""
        constituency = re.sub(r"\s*\((SC|ST)\)\s*$", "", constituency).strip()
        if cid:
            csv_by_id[cid] = {"name": name, "constituency": constituency}

    # Load JSON profiles
    with open(PROFILES_JSON, "r", encoding="utf-8") as f:
        profiles = json.load(f)

    json_by_key = {}
    for p in profiles:
        cp = p.get("candidate_profile", {})
        cid = str(cp.get("candidate_id", ""))
        csv_info = csv_by_id.get(cid)
        if csv_info:
            key = (csv_info["constituency"], csv_info["name"].upper())
            json_by_key[key] = p

    return json_by_key


def parse_rupees(val):
    if not val or "Nil" in str(val):
        return 0
    cleaned = re.sub(r"[Rs\s\xa0,~]", "", str(val))
    digits = re.match(r"^([\d.]+)", cleaned)
    return int(float(digits.group(1))) if digits else 0


def parse_criminal(status):
    if not status or "No criminal" in status:
        return 0
    m = re.search(r"(\d+)", status)
    return int(m.group(1)) if m else 0


def run():
    Base.metadata.create_all(bind=engine)

    ac_map = load_electors()
    print(f"Loaded electors: {len(ac_map)} ACs, {len(set(a['district'] for a in ac_map.values()))} districts")

    party_names = load_party_names()
    print(f"Loaded {len(party_names)} party name mappings")

    affidavit_data = load_affidavit_bridge()
    print(f"Loaded {len(affidavit_data)} affidavit profiles")

    # Load detailed results
    wb = openpyxl.load_workbook(RESULTS_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    result_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    print(f"Loaded {len(result_rows) - 1} result rows")

    with Session(engine) as session:
        existing = session.query(Election).filter(
            Election.state == "West Bengal", Election.year == 2016
        ).first()
        if existing:
            print(f"ERROR: West Bengal 2016 already exists (id={existing.id}). Skipping.")
            return

        # 1. Create election
        election = Election(
            state="West Bengal", year=2016, type="Assembly",
            name="West Bengal Legislative Assembly Election 2016",
        )
        session.add(election)
        session.flush()
        print(f"Created election: {election.name} (id={election.id})")

        # 2. Create districts
        district_cache = {}
        for ac_info in ac_map.values():
            dname = ac_info["district"]
            if dname not in district_cache:
                district = District(election_id=election.id, name=dname.title())
                session.add(district)
                district_cache[dname] = district
        session.flush()
        print(f"Created {len(district_cache)} districts")

        # 3. Parse results -> group by AC
        ac_data = {}
        for row in result_rows[1:]:
            ac_no = row[0]
            if ac_no is None or not isinstance(ac_no, (int, float)):
                continue
            ac_no = int(ac_no)
            cand_name = str(row[2]).strip() if row[2] else None
            sex = str(row[3]).strip() if row[3] else None
            age = int(row[4]) if row[4] else None
            party_abbr = str(row[6]).strip() if row[6] else None
            gen_votes = int(row[7]) if row[7] else 0
            postal_votes = int(row[8]) if row[8] else 0
            total_valid = int(row[9]) if row[9] else 0
            total_electors_col = int(row[10]) if row[10] else None
            total_votes_col = int(row[11]) if row[11] else None

            if ac_no not in ac_data:
                ac_data[ac_no] = {
                    "total_electors_from_results": total_electors_col,
                    "total_votes_from_results": total_votes_col,
                    "candidates": [],
                }

            ac_data[ac_no]["candidates"].append({
                "name": cand_name,
                "sex": sex,
                "age": age,
                "party": party_abbr,
                "gen_votes": gen_votes,
                "postal_votes": postal_votes,
                "total_votes": total_valid,
            })

        # 4. Create constituencies
        constituency_cache = {}
        for ac_no, info in ac_data.items():
            mapping = ac_map.get(ac_no, {})
            district_name = mapping.get("district")
            district = district_cache.get(district_name)

            total_electors = mapping.get("total_electors") or info.get("total_electors_from_results")
            total_votes = info.get("total_votes_from_results") or sum(c["total_votes"] for c in info["candidates"])
            turnout = round(total_votes / total_electors * 100, 2) if total_electors and total_votes else None

            sorted_cands = sorted(info["candidates"], key=lambda x: x["total_votes"], reverse=True)
            real_cands = [c for c in sorted_cands if c["party"] != "NOTA" and c["name"] != "None of the Above"]
            margin = None
            if len(real_cands) >= 2:
                margin = real_cands[0]["total_votes"] - real_cands[1]["total_votes"]

            c = Constituency(
                election_id=election.id,
                district_id=district.id if district else None,
                ac_no=ac_no,
                name=mapping.get("name", str(ac_no)),
                category=mapping.get("category"),
                total_electors=total_electors,
                male_electors=mapping.get("male_electors"),
                female_electors=mapping.get("female_electors"),
                third_gender_electors=mapping.get("third_gender_electors"),
                total_polling_stations=mapping.get("total_polling_stations"),
                total_votes_polled=total_votes,
                turnout_pct=turnout,
                winning_margin=margin,
            )
            session.add(c)
            constituency_cache[ac_no] = c

        session.flush()
        print(f"Created {len(constituency_cache)} constituencies")

        # 5. Build party cache (reuse existing parties)
        party_cache = {}
        for p in session.query(Party).all():
            party_cache[p.abbr.upper()] = p
            party_cache[p.name.upper()] = p

        # 6. Import candidates
        imported = 0
        for ac_no, info in ac_data.items():
            constituency = constituency_cache[ac_no]
            sorted_cands = sorted(info["candidates"], key=lambda x: x["total_votes"], reverse=True)
            total_valid_in_ac = sum(c["total_votes"] for c in sorted_cands)

            for pos, cand in enumerate(sorted_cands, start=1):
                name = cand["name"] or "UNKNOWN"
                is_nota = (cand["party"] == "NOTA" or name == "None of the Above")
                if name == "None of the Above":
                    name = "NOTA"

                party = None
                if not is_nota and cand["party"]:
                    raw_abbr = cand["party"]
                    canonical_abbr = ABBR_ALIASES.get(raw_abbr, raw_abbr)

                    p_key = canonical_abbr.upper()
                    if p_key not in party_cache:
                        p_key = raw_abbr.upper()

                    if p_key not in party_cache:
                        full_name = party_names.get(raw_abbr, raw_abbr)
                        new_party = Party(
                            name=full_name,
                            abbr=canonical_abbr,
                            color=PARTY_COLORS.get(raw_abbr) or PARTY_COLORS.get(canonical_abbr),
                        )
                        session.add(new_party)
                        session.flush()
                        party_cache[canonical_abbr.upper()] = new_party
                        party_cache[full_name.upper()] = new_party
                        p_key = canonical_abbr.upper()

                    party = party_cache[p_key]

                vote_pct = round(cand["total_votes"] / total_valid_in_ac * 100, 2) if total_valid_in_ac > 0 else 0

                ac_name_upper = constituency.name.upper()
                affidavit = affidavit_data.get((ac_name_upper, name.upper()))

                education = None
                occupation = None
                declared_assets = None
                liabilities_val = None
                criminal_cases = 0
                image_url = None

                if affidavit:
                    cp = affidavit.get("candidate_profile", {})
                    edu = (cp.get("education") or "").strip()
                    if edu.startswith("Category:"):
                        edu = edu[len("Category:"):].strip()
                    if edu and edu not in ("Not mentioned", "Not Given"):
                        education = edu

                    occ = (affidavit.get("profession", {}).get("self") or "").strip()
                    if occ and occ not in ("Not mentioned", "Not Given"):
                        occupation = occ

                    assets = parse_rupees(affidavit.get("assets_summary", {}).get("total_assets"))
                    if assets > 0:
                        declared_assets = assets
                    liab = parse_rupees(affidavit.get("assets_summary", {}).get("total_liabilities"))
                    if liab > 0:
                        liabilities_val = liab

                    criminal_cases = parse_criminal(cp.get("crime_status"))

                    img = (cp.get("image_url") or "").strip()
                    if img and img != "None":
                        image_url = img

                session.add(Candidate(
                    election_id=election.id,
                    constituency_id=constituency.id,
                    party_id=party.id if party else None,
                    name=name,
                    gender=cand["sex"],
                    age=cand["age"],
                    position=pos,
                    votes_general=cand["gen_votes"],
                    votes_postal=cand["postal_votes"],
                    votes_total=cand["total_votes"],
                    vote_pct=vote_pct,
                    is_nota=is_nota,
                    education=education,
                    occupation=occupation,
                    declared_assets=declared_assets,
                    liabilities=liabilities_val,
                    criminal_cases=criminal_cases,
                    image_url=image_url,
                ))
                imported += 1

        session.flush()
        print(f"Imported {imported} candidates")

        session.commit()

        # Summary
        print(f"\n=== Summary ===")
        n_real = session.query(Candidate).filter(
            Candidate.election_id == election.id, Candidate.is_nota == False
        ).count()
        n_nota = session.query(Candidate).filter(
            Candidate.election_id == election.id, Candidate.is_nota == True
        ).count()
        has_edu = session.query(Candidate).filter(
            Candidate.election_id == election.id, Candidate.education != None
        ).count()
        has_assets = session.query(Candidate).filter(
            Candidate.election_id == election.id, Candidate.declared_assets != None
        ).count()
        has_img = session.query(Candidate).filter(
            Candidate.election_id == election.id, Candidate.image_url != None
        ).count()
        print(f"  Candidates: {n_real} real + {n_nota} NOTA")
        print(f"  With education: {has_edu}")
        print(f"  With assets: {has_assets}")
        print(f"  With image: {has_img}")

        print(f"\n=== Top Parties by Seats Won ===")
        winners = session.query(Candidate).filter(
            Candidate.election_id == election.id,
            Candidate.position == 1,
            Candidate.is_nota == False,
        ).all()
        party_seats = {}
        for w in winners:
            abbr = w.party.abbr if w.party else "IND"
            party_seats[abbr] = party_seats.get(abbr, 0) + 1
        for p, seats in sorted(party_seats.items(), key=lambda x: x[1], reverse=True)[:10]:
            print(f"  {p}: {seats} seats")


if __name__ == "__main__":
    run()
