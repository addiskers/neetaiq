"""Import Assam 2026 candidate nomination data from JSON + mapping file.

Data sources:
  - Candidates: assam 2026/candidate_profile_2026 1.json (722 candidates)
  - District/AC mapping: Assam District and AC name (2016,2021,2026).xlsx (2026 cols 4-7)
  - District mapping: same file, same-row pairs 2021<->2026 districts
"""
import sys
import os
import re
import json
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.database import engine, Base
from app.models import Election, District, Constituency, Party, Candidate, DistrictMapping

PROJECT_ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
MAPPING_FILE = os.path.join(PROJECT_ROOT, "Assam District and AC name (2016,2021,2026).xlsx")
JSON_FILE = os.path.join(PROJECT_ROOT, "assam 2026", "candidate_profile_2026 1.json")

PARTY_COLORS = {
    "Bharatiya Janata Party": "#FF9933",
    "Indian National Congress": "#00BFFF",
    "All India United Democratic Front": "#006400",
    "Asom Gana Parishad": "#FFFFFF",
    "Bodoland Peoples Front": "#228B22",
    "United Peoples Party Liberal": "#FFD700",
    "Communist Party of India (Marxist)": "#FF0000",
    "All India Trinamool Congress": "#00FF00",
    "Aam Aadmi Party": "#0066CC",
}

# Known abbreviations for common parties (JSON has full names)
PARTY_ABBR = {
    "Bharatiya Janata Party": "BJP",
    "Indian National Congress": "INC",
    "All India United Democratic Front": "AIUDF",
    "Asom Gana Parishad": "AGP",
    "United Peoples Party Liberal": "UPPL",
    "Bodoland Peoples Front": "BOPF",
    "Communist Party of India (Marxist)": "CPI(M)",
    "Communist Party of India": "CPI",
    "All India Trinamool Congress": "AITC",
    "Nationalist Congress Party": "NCP",
    "Janata Dal (United)": "JD(U)",
    "Samajwadi Party": "SP",
    "Rashtriya Janata Dal": "RJD",
    "Aam Aadmi Party": "AAP",
    "Bahujan Samaj Party": "BSP",
    "Independent": "IND",
}


def load_2026_ac_mapping():
    """Parse 2026 columns (4-7) from the district-AC mapping file."""
    wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    ac_map = {}
    current_district = None

    for row in rows[2:]:
        if row[5]:
            current_district = str(row[5]).strip().title()
        if row[6] is not None and row[7]:
            ac_no = int(row[6])
            raw_name = str(row[7]).strip()
            category = "GEN"
            if "(SC)" in raw_name:
                category = "SC"
                raw_name = raw_name.replace("(SC)", "").strip()
            elif "(ST)" in raw_name:
                category = "ST"
                raw_name = raw_name.replace("(ST)", "").strip()
            ac_map[ac_no] = {
                "district": current_district,
                "name": raw_name,
                "category": category,
            }
    return ac_map


def load_district_pairs():
    """Parse same-row district pairs: 2021 (col 1) ↔ 2026 (col 5).

    Returns list of (district_2021_name, district_2026_name) tuples.
    For new 2026 districts (Bajali, Tamulpur), district_2021_name = None.
    """
    wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    pairs = []
    seen_2021 = set()
    seen_2026 = set()

    for row in rows[2:]:
        d2021 = str(row[1]).strip().title() if row[1] else None
        d2026 = str(row[5]).strip().title() if row[5] else None

        if d2026 and d2026 not in seen_2026:
            seen_2026.add(d2026)
            if d2021 and d2021 not in seen_2021:
                seen_2021.add(d2021)
                pairs.append((d2021, d2026))
            elif not d2021:
                # New 2026 district with no 2021 equivalent
                pairs.append((None, d2026))

    return pairs


def parse_rupees(val):
    """Parse 'Rs 64,09,147' or 'Rs\xa064,09,147' to integer. Returns 0 for 'Nil'/None."""
    if not val or val == "Nil":
        return 0
    # Remove "Rs", non-breaking spaces, commas
    cleaned = re.sub(r"[Rs\s\xa0,~]", "", str(val))
    # Take only digits
    digits = re.match(r"^(\d+)", cleaned)
    return int(digits.group(1)) if digits else 0


def parse_criminal_cases(status):
    """Extract count from 'Number of Criminal Cases: 3' or return 0."""
    if not status or "No criminal" in status:
        return 0
    match = re.search(r"(\d+)", status)
    return int(match.group(1)) if match else 0


def import_data():
    # Add new tables (don't drop — keep 2021 data!)
    print("Creating new tables (if any)...")
    Base.metadata.create_all(bind=engine)

    ac_map = load_2026_ac_mapping()
    print(f"Loaded 2026 mapping: {len(ac_map)} ACs, {len(set(a['district'] for a in ac_map.values()))} districts")

    district_pairs = load_district_pairs()
    print(f"Loaded {len(district_pairs)} district pairs (2021<->2026)")

    # Load JSON
    with open(JSON_FILE, "r", encoding="utf-8") as f:
        candidates_json = json.load(f)
    print(f"Loaded {len(candidates_json)} candidates from JSON")

    with Session(engine) as session:
        # Check if 2026 election already exists
        existing = session.query(Election).filter(Election.state == "Assam", Election.year == 2026).first()
        if existing:
            print(f"ERROR: Assam 2026 election already exists (id={existing.id}). Delete it first or skip.")
            return

        # 1. Create election
        election = Election(
            state="Assam",
            year=2026,
            type="Assembly",
            name="Assam Legislative Assembly Election 2026",
        )
        session.add(election)
        session.flush()
        print(f"Created election: {election.name} (id={election.id})")

        # 2. Create districts from mapping
        district_cache = {}  # name (title case) -> District
        for ac_info in ac_map.values():
            dname = ac_info["district"]
            if dname not in district_cache:
                district = District(election_id=election.id, name=dname)
                session.add(district)
                district_cache[dname] = district
        session.flush()
        print(f"Created {len(district_cache)} districts")

        # 3. Create constituencies from mapping
        constituency_cache = {}  # uppercase name -> Constituency
        for ac_no, ac_info in ac_map.items():
            district = district_cache[ac_info["district"]]
            c = Constituency(
                election_id=election.id,
                district_id=district.id,
                ac_no=ac_no,
                name=ac_info["name"],
                category=ac_info["category"],
            )
            session.add(c)
            constituency_cache[ac_info["name"].upper()] = c
        session.flush()
        print(f"Created {len(constituency_cache)} constituencies")

        # 4. Build party cache from existing parties
        party_cache = {}  # name (lowercase) -> Party
        for p in session.query(Party).all():
            party_cache[p.name.lower()] = p
            party_cache[p.abbr.lower()] = p

        # 5. Import candidates from JSON
        imported = 0
        skipped = 0
        for cand in candidates_json:
            cp = cand.get("candidate_profile", {})
            ac_name = (cp.get("ac_name") or "").strip().upper()
            constituency = constituency_cache.get(ac_name)

            if not constituency:
                print(f"  WARNING: No constituency match for '{ac_name}', skipping candidate {cp.get('candidate_id')}")
                skipped += 1
                continue

            # Party
            party_name = (cp.get("party") or "").strip()
            party = None
            if party_name:
                party_key = party_name.lower()
                if party_key not in party_cache:
                    abbr = PARTY_ABBR.get(party_name, party_name[:10].upper())
                    # Check if abbr already exists
                    existing_party = session.query(Party).filter(func.lower(Party.abbr) == abbr.lower()).first()
                    if existing_party:
                        party_cache[party_key] = existing_party
                    else:
                        new_party = Party(
                            name=party_name,
                            abbr=abbr,
                            color=PARTY_COLORS.get(party_name),
                        )
                        session.add(new_party)
                        session.flush()
                        party_cache[party_key] = new_party
                party = party_cache[party_key]

            # Parse fields
            age_str = (cp.get("age") or "").strip()
            age = int(age_str) if age_str.isdigit() else None

            education = (cp.get("education") or "").strip()
            if education.startswith("Category:"):
                education = education[len("Category:"):].strip()

            occupation = (cand.get("profession", {}).get("self") or "").strip() or None

            assets_summary = cand.get("assets_summary", {})
            declared_assets = parse_rupees(assets_summary.get("total_assets"))
            liabilities_val = parse_rupees(assets_summary.get("total_liabilities"))

            criminal_cases = parse_criminal_cases(cp.get("crime_status"))
            image_url = (cp.get("image_url") or "").strip() or None

            candidate = Candidate(
                election_id=election.id,
                constituency_id=constituency.id,
                party_id=party.id if party else None,
                name=party_name,  # placeholder — will fix below
                gender=None,  # not in JSON
                age=age,
                education=education or None,
                occupation=occupation,
                declared_assets=declared_assets if declared_assets > 0 else None,
                liabilities=liabilities_val if liabilities_val > 0 else None,
                criminal_cases=criminal_cases,
                image_url=image_url,
                is_nota=False,
            )
            # Get actual name from CSV or use candidate_id
            # JSON has 'name' = "Assam 2026" (not the candidate name!)
            # We need the real name from CSV
            candidate.name = f"CANDIDATE_{cp.get('candidate_id', '')}"
            session.add(candidate)
            imported += 1

        session.flush()
        print(f"Imported {imported} candidates ({skipped} skipped)")

        # 6. Now fix candidate names from CSV
        print("Fixing candidate names from CSV...")
        import csv
        csv_file = os.path.join(PROJECT_ROOT, "assam 2026", "assam_candidates_2026.csv")
        name_map = {}  # candidate_id -> name
        with open(csv_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                cid = row.get("candidate_id", "").strip()
                name = row.get("candidate_name", "").strip()
                if cid and name:
                    name_map[cid] = name

        updated_names = 0
        for cand in session.query(Candidate).filter(Candidate.election_id == election.id).all():
            if cand.name.startswith("CANDIDATE_"):
                cid = cand.name.replace("CANDIDATE_", "")
                if cid in name_map:
                    cand.name = name_map[cid]
                    updated_names += 1
        print(f"Updated {updated_names} candidate names from CSV")

        # 7. Populate district_mappings
        print("Populating district mappings (2021<->2026)...")
        election_2021 = session.query(Election).filter(Election.state == "Assam", Election.year == 2021).first()

        if election_2021:
            districts_2021 = {d.name.lower(): d for d in session.query(District).filter(District.election_id == election_2021.id).all()}
            districts_2026 = {d.name.lower(): d for d in session.query(District).filter(District.election_id == election.id).all()}

            mapped = 0
            for d2021_name, d2026_name in district_pairs:
                d2021 = districts_2021.get(d2021_name.lower()) if d2021_name else None
                d2026 = districts_2026.get(d2026_name.lower()) if d2026_name else None

                if d2026:
                    dm = DistrictMapping(
                        district_2021_id=d2021.id if d2021 else None,
                        district_2026_id=d2026.id,
                    )
                    session.add(dm)
                    mapped += 1
                    label = f"{d2021_name or '(new)'} -> {d2026_name}"
                    if not d2021:
                        label += " [NEW]"
                    print(f"  {label}")

            print(f"Created {mapped} district mappings")
        else:
            print("  WARNING: No 2021 election found, skipping district mappings")

        session.commit()

        # Summary
        n_districts = session.query(District).filter(District.election_id == election.id).count()
        n_constituencies = session.query(Constituency).filter(Constituency.election_id == election.id).count()
        n_candidates = session.query(Candidate).filter(Candidate.election_id == election.id).count()
        n_mappings = session.query(DistrictMapping).count()

        print(f"\n=== Import Summary ===")
        print(f"Election:       {election.name}")
        print(f"Districts:      {n_districts}")
        print(f"Constituencies: {n_constituencies}")
        print(f"Candidates:     {n_candidates}")
        print(f"District maps:  {n_mappings}")

        # Verify both elections
        print(f"\n=== All Elections ===")
        for e in session.query(Election).all():
            count = session.query(Candidate).filter(Candidate.election_id == e.id).count()
            print(f"  {e.state} {e.year}: {count} candidates")


if __name__ == "__main__":
    import_data()
