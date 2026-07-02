"""Import Puducherry 2026 nomination data (no results yet).

Data sources (all under Puducherry/Puducherry_2026/):
  - Candidate profiles: my_neta_2026_candidate_profile.json
  - Candidate bridge: my_neta_puducherry_candidates_2026.csv
  - Electors: District_Constituency wise Electors & Voters Data_2026.xlsx

Usage: cd backend && python -m scripts.import_puducherry_2026
"""
import sys, os, re, json, csv
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from app.database import engine, Base
from app.models.puducherry import Election, District, Constituency, Party, Candidate

PROJECT_ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
DATA_DIR = os.path.join(PROJECT_ROOT, "Puducherry", "Puducherry_2026")

PROFILES_JSON  = os.path.join(DATA_DIR, "my_neta_2026_candidate_profile.json")
CANDIDATES_CSV = os.path.join(DATA_DIR, "my_neta_puducherry_candidates_2026.csv")
ELECTORS_FILE  = os.path.join(DATA_DIR, "District_Constituency wise Electors & Voters Data_2026.xlsx")
MAPPING_FILE   = os.path.join(PROJECT_ROOT, "puducherry", "puducherry_checking.xlsx")

PARTY_COLORS = {
    "AINRC": "#800080", "DMK": "#FF0000", "AIADMK": "#228B22", "INC": "#00BFFF",
    "BJP": "#FF9933", "PMK": "#FFFF00", "CPI": "#FF4444", "CPI(M)": "#FF0000",
    "BSP": "#0000FF", "IND": "#808080", "NCP": "#004080", "AAP": "#0066CC",
    "AITC": "#00FF00", "NR_CONGRESS": "#800080",
}

PARTY_ABBR = {
    "Bharatiya Janata Party": "BJP",
    "Indian National Congress": "INC",
    "Dravida Munnetra Kazhagam": "DMK",
    "All India Anna Dravida Munnetra Kazhagam": "AIADMK",
    "N.R. Congress": "AINRC",
    "NR Congress": "AINRC",
    "Pattali Makkal Katchi": "PMK",
    "Communist Party of India (Marxist)": "CPI(M)",
    "Communist Party of India": "CPI",
    "Bahujan Samaj Party": "BSP",
    "Aam Aadmi Party": "AAP",
    "Nationalist Congress Party": "NCP",
    "All India Trinamool Congress": "AITC",
    "Independent": "IND",
}

# ACs 1-23: Puducherry, 24-28: Karaikal, 29: Mahe, 30: Yanam
AC_DISTRICT = {**{n: "Puducherry" for n in range(1, 24)},
               **{n: "Karaikal" for n in range(24, 29)},
               29: "Mahe", 30: "Yanam"}


def _normalize(name):
    return re.sub(r"[.\s@]+", "", name).upper()


def load_ac_mapping():
    wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True)
    ws = wb["Ac Name"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    name_to_ac = {}
    for row in rows[1:]:
        if not row[0]:
            continue
        ac_no = int(row[0])
        for col in [1, 4, 7, 10]:
            if col < len(row) and row[col]:
                v = str(row[col]).strip()
                stripped = re.sub(r"\s*\(?(SC|ST)\)?\s*$", "", v, flags=re.I).strip()
                name_to_ac[_normalize(v)] = ac_no
                name_to_ac[_normalize(stripped)] = ac_no
    return name_to_ac


def load_electors():
    # Cols: District(0), AC No(1), AC Name(2), Male(3), Female(4), Third(5), Total(6)
    # Name may include "(SC)" or "(ST)" suffix — strip it and derive category
    wb = openpyxl.load_workbook(ELECTORS_FILE, read_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()
    em = {}
    for row in rows[1:]:
        if row[1] is None:
            continue
        ac_no = int(row[1])
        raw_name = re.sub(r"\s+", " ", str(row[2]).strip()).upper() if row[2] else ""
        # Extract category from name suffix, then strip it
        cat_match = re.search(r"\s*\((SC|ST)\)\s*$", raw_name)
        category = cat_match.group(1) if cat_match else "GEN"
        name = re.sub(r"\s*\((SC|ST)\)\s*$", "", raw_name).strip()
        em[ac_no] = {
            "name": name,
            "category": category,
            "male_electors": int(row[3]) if row[3] else None,
            "female_electors": int(row[4]) if row[4] else None,
            "third_gender_electors": int(row[5]) if row[5] else None,
            "total_electors": int(row[6]) if row[6] else None,
        }
    return em


def parse_rupees(val):
    if not val or "Nil" in str(val):
        return 0
    cleaned = re.sub(r"[Rs\s\xa0,~]", "", str(val))
    digits = re.match(r"^([\d.]+)", cleaned)
    return int(float(digits.group(1))) if digits else 0


def parse_criminal(status):
    if not status or "No criminal" in status:
        return 0
    m = re.search(r"(\d+)", status)
    return int(m.group(1)) if m else 0


def run():
    Base.metadata.create_all(bind=engine)

    em = load_electors()
    name_to_ac = load_ac_mapping()
    print(f"Loaded {len(em)} ACs from electors file")

    # Load AC category from 2021 constituencies (reuse same categories as they don't change)
    ac_categories = {}
    with Session(engine) as session:
        prev = session.query(Election).filter_by(state="Puducherry", year=2021).first()
        if prev:
            for c in session.query(Constituency).filter_by(election_id=prev.id).all():
                ac_categories[c.ac_no] = c.category

    with open(PROFILES_JSON, "r", encoding="utf-8") as f:
        profiles = json.load(f)
    print(f"Loaded {len(profiles)} candidate profiles")

    with Session(engine) as session:
        if session.query(Election).filter_by(state="Puducherry", year=2026).first():
            print("Puducherry 2026 already exists — skipping.")
            return

        election = Election(state="Puducherry", year=2026, type="Assembly",
                            name="Puducherry Legislative Assembly Election 2026")
        session.add(election)
        session.flush()

        district_cache = {dname: District(election_id=election.id, name=dname)
                          for dname in sorted(set(AC_DISTRICT.values()))}
        for d in district_cache.values():
            session.add(d)
        session.flush()

        # Create constituencies from electors file
        constituency_cache = {}
        for ac_no, info in em.items():
            dist = district_cache.get(AC_DISTRICT.get(ac_no, "Puducherry"))
            con = Constituency(
                election_id=election.id, district_id=dist.id if dist else None,
                ac_no=ac_no, name=info["name"],
                category=info.get("category") or ac_categories.get(ac_no, "GEN"),
                total_electors=info["total_electors"],
                male_electors=info["male_electors"],
                female_electors=info["female_electors"],
                third_gender_electors=info["third_gender_electors"],
            )
            session.add(con)
            constituency_cache[ac_no] = con
        session.flush()
        print(f"Created {len(constituency_cache)} constituencies")

        party_cache = {p.abbr.upper(): p for p in session.query(Party).all()}
        imported = skipped = 0

        for cand_data in profiles:
            cp = cand_data.get("candidate_profile", {})
            ac_name_raw = re.sub(r"\s*\(?(SC|ST)\)?\s*$", "", (cp.get("ac_name") or "").strip(), flags=re.I).strip()
            ac_no_found = name_to_ac.get(_normalize(ac_name_raw))
            con = constituency_cache.get(ac_no_found) if ac_no_found else None
            if not con:
                print(f"  WARNING: No constituency match for '{ac_name_raw}'")
                skipped += 1
                continue
            con_id = con.id
            # Use ac_no from mapping for clean lookup
            ac_no = ac_no_found

            # Party
            raw_party = (cp.get("party") or "").strip().rstrip()
            party = None
            if raw_party and raw_party.upper() not in ("NOTA", ""):
                abbr = PARTY_ABBR.get(raw_party, raw_party[:15].upper())
                if abbr.upper() not in party_cache:
                    p = Party(name=raw_party, abbr=abbr, color=PARTY_COLORS.get(abbr))
                    session.add(p)
                    session.flush()
                    party_cache[abbr.upper()] = p
                party = party_cache[abbr.upper()]

            # Name
            name = (cp.get("name") or "").strip().rstrip()

            # Age
            age_str = (cp.get("age") or "").strip()
            age = int(age_str) if age_str.isdigit() else None

            # Education
            edu = (cp.get("education") or "").strip()
            if edu.startswith("Category:"):
                edu = edu[len("Category:"):].strip()
            education = edu if edu and edu not in ("Not mentioned", "Not Given", "Not Given ") else None

            occupation = (cand_data.get("profession", {}).get("self") or "").strip() or None

            assets_summary = cand_data.get("assets_summary", {})
            declared_assets = parse_rupees(assets_summary.get("total_assets"))
            liabilities_val = parse_rupees(assets_summary.get("total_liabilities"))

            criminal_cases = parse_criminal(cp.get("crime_status"))
            image_url = (cp.get("image_url") or "").strip() or None

            session.add(Candidate(
                election_id=election.id, constituency_id=con.id,
                party_id=party.id if party else None,
                name=name or f"CANDIDATE_{cp.get('candidate_id', '')}",
                age=age, education=education, occupation=occupation,
                declared_assets=declared_assets if declared_assets > 0 else None,
                liabilities=liabilities_val if liabilities_val > 0 else None,
                criminal_cases=criminal_cases, image_url=image_url,
                is_nota=False,
            ))
            imported += 1

        session.commit()

    print(f"\n=== Puducherry 2026 ===")
    print(f"Constituencies: {len(constituency_cache)}")
    print(f"Candidates: {imported} imported, {skipped} skipped")
    total_electors = sum(v["total_electors"] or 0 for v in em.values())
    print(f"Total electors: {total_electors:,}")


if __name__ == "__main__":
    run()
