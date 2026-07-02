"""Import Manipur 2012 election data.

Data sources (all under Manipur/Manipur_2012/):
  - Results:  Detailed_Results_2012.xlsx
  - Electors: Electors & Voters Data_2012.xlsx
  - CSV:      Manipur_candidates_2012.csv
  - JSON:     2012_candidate_profile.json
  - Mapping:  Manipur/Manipur_Data_Mapping.xlsx

Results columns (0-indexed):
  STATE(0), District(1), AC_NO(2), AC_Name(3), Candidate(4),
  Sex(5), Age(6), Category(7), Party(8),
  General(9), Postal(10), Total(11), %(12), Electors(13)

Electors columns:
  District(0), AC_NO(1), AC_Name(2), Male(3), Female(4), Others(5), Total(6)

Usage: cd backend && python -m scripts.import_manipur_2012
"""
import sys, os, re, json, csv
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from app.database import engine, Base
from app.models.manipur import Election, District, Constituency, Party, Candidate

PROJECT_ROOT   = os.path.join(os.path.dirname(__file__), "..", "..")
DATA_DIR       = os.path.join(PROJECT_ROOT, "Manipur", "Manipur_2012")
MAPPING_FILE   = os.path.join(PROJECT_ROOT, "Manipur", "Manipur_Data_Mapping.xlsx")
RESULTS_FILE   = os.path.join(DATA_DIR, "Detailed_Results_2012.xlsx")
ELECTORS_FILE  = os.path.join(DATA_DIR, "Electors & Voters Data_2012.xlsx")
PROFILES_JSON  = os.path.join(DATA_DIR, "2012_candidate_profile.json")
CANDIDATES_CSV = os.path.join(DATA_DIR, "Manipur_candidates_2012.csv")

PARTY_COLORS = {
    "BJP": "#FF9933", "INC": "#00BFFF", "IND": "#808080",
    "NPF": "#8B4513", "NPP": "#228B22", "AITC": "#00FF00",
    "CPI": "#FF4444", "CPI(M)": "#CC0000", "BSP": "#0000FF",
    "NCP": "#004080", "JDU": "#00CC66", "LJP": "#FFCC00",
    "IVD": "#9932CC", "KSD": "#DAA520", "MNDF": "#708090",
    "MSCP": "#CD853F", "NPC": "#4169E1",
}


def _normalize(name: str) -> str:
    return re.sub(r"[.\s@]+", "", str(name)).upper()


def load_ac_mapping():
    """Build {normalized_name: ac_no} from Manipur_Data_Mapping.xlsx AC sheet.

    AC sheet layout (0-indexed):
      row0: year labels  (2012, None, None, 2017, None, None, 2022, ...)
      row1: field labels (AC_NO, AC_Name, None, AC_NO, AC_Name, ...)
      row2+: data rows
    Name columns: 1 (2012), 4 (2017), 7 (2022); AC_NO in col 0.
    """
    wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True)
    ws = wb["AC"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    name_to_ac = {}
    for row in rows[2:]:  # skip 2 header rows
        if not row[0]:
            continue
        ac_no = int(row[0])
        for col in [1, 4, 7]:  # 2012, 2017, 2022 name columns
            if col < len(row) and row[col]:
                v = str(row[col]).strip()
                stripped = re.sub(r"\s*\(?(SC|ST)\)?\s*$", "", v, flags=re.I).strip()
                name_to_ac[_normalize(v)] = ac_no
                name_to_ac[_normalize(stripped)] = ac_no
        # Parse myneta alternate spellings from col 5 comments
        if len(row) > 5 and row[5]:
            comment = str(row[5])
            alt_m = re.search(r"showing as\s+(\w+)", comment, re.I)
            if alt_m:
                name_to_ac[_normalize(alt_m.group(1))] = ac_no
    # Hardcoded aliases for CSV/JSON spelling variants not in the mapping file
    name_to_ac.update({
        "NOURIYAPAKHANGLAKPA": 21,  # 2012 CSV; mapping: Naoriya Pakhanglakpa
        "NAORIAPAKHANGLAKPA":  21,  # 2022 CSV; mapping: Naoriya Pakhanglakpa
        "SAGOIBAND":           11,  # 2012 CSV; mapping: Sagolband
        "THONJU":               5,  # 2012 CSV; mapping: Thongju
        "TIPAIMUK":            55,  # 2012 CSV; mapping: Tipaimukh
    })
    return name_to_ac


def load_electors():
    """Electors & Voters Data_2012.xlsx:
    District(0), AC_NO(1), AC_Name(2), Male(3), Female(4), Others(5), Total(6)
    """
    wb = openpyxl.load_workbook(ELECTORS_FILE, read_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()
    em = {}
    for row in rows[1:]:
        if not isinstance(row[1], (int, float)):
            continue
        ac_no = int(row[1])
        raw_name = re.sub(r"\s+", " ", str(row[2]).strip()) if row[2] else ""
        cat_m = re.search(r"\s*\((SC|ST)\)\s*$", raw_name, re.I)
        category = cat_m.group(1).upper() if cat_m else None
        name = re.sub(r"\s*\((SC|ST)\)\s*$", "", raw_name, flags=re.I).strip()
        em[ac_no] = {
            "name": name, "category": category,
            "male_electors":          int(row[3]) if row[3] else None,
            "female_electors":        int(row[4]) if row[4] else None,
            "third_gender_electors":  int(row[5]) if row[5] else None,
            "total_electors":         int(row[6]) if row[6] else None,
        }
    return em


def load_profiles(name_to_ac):
    """Merge JSON profiles with CSV summary data, keyed by (ac_no, norm_name)."""
    csv_by_id = {}
    with open(CANDIDATES_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            cid = row.get("candidate_id", "").strip()
            con_raw = re.sub(r"\s*\(?(SC|ST)\)?\s*$", "", row.get("constituency", "").strip(), flags=re.I).strip()
            if cid:
                csv_by_id[cid] = {
                    "name":           row.get("candidate_name", "").strip(),
                    "constituency":   con_raw,
                    "criminal_cases": int(row.get("criminal_cases") or 0),
                    "education":      row.get("education", "").strip(),
                    "assets":         row.get("assets", ""),
                    "liabilities":    row.get("liabilities", ""),
                }

    with open(PROFILES_JSON, encoding="utf-8") as f:
        profiles = json.load(f)

    by_acno = {}
    for p in profiles:
        cp   = p.get("candidate_profile", {})
        cid  = str(cp.get("candidate_id", ""))
        info = csv_by_id.pop(cid, None)

        ac_no = None
        if info:
            ac_no = name_to_ac.get(_normalize(info["constituency"]))
        if not ac_no:
            ac_raw = re.sub(r"\s*\(?(SC|ST)\)?\s*$", "", (cp.get("ac_name") or "").strip(), flags=re.I).strip()
            ac_no  = name_to_ac.get(_normalize(ac_raw))

        if not ac_no:
            continue

        # Store under both JSON name and CSV name so either can be looked up exactly
        json_name = (cp.get("name") or "").strip()
        csv_name  = (info.get("name") or "").strip() if info else ""
        for kname in dict.fromkeys(n for n in [json_name, csv_name] if n):
            by_acno.setdefault((ac_no, _normalize(kname)), {**p, "_csv": info})

    for cid, info in csv_by_id.items():
        ac_no = name_to_ac.get(_normalize(info["constituency"]))
        if ac_no and info["name"]:
            by_acno.setdefault((ac_no, _normalize(info["name"])),
                               {"candidate_profile": {}, "assets_summary": {}, "_csv": info})
    return by_acno


def parse_rupees(val):
    """Parse Indian currency string to integer rupees.
    Strips all non-digit characters to handle double-encoded Unicode and URL artifacts.
    """
    if not val or "Nil" in str(val):
        return 0
    s = str(val).strip()
    if s.startswith("http"):  # URL placeholder, no actual value
        return 0
    digits = re.sub(r"[^\d]", "", s.split("~")[0])
    return int(digits) if digits else 0


def parse_criminal(status):
    if not status or "No criminal" in str(status):
        return 0
    m = re.search(r"(\d+)", str(status))
    return int(m.group(1)) if m else 0


def _name_words(name: str) -> frozenset:
    base = re.split(r"\s*@", name)[0].strip()
    return frozenset(w for w in re.findall(r"[A-Za-z]{3,}", base) if w.upper() not in ("MRS", "MR", "DR", "SH", "SHRI"))


def build_word_map(profile_map: dict) -> dict:
    wm = {}
    for (ac_no, _), profile in profile_map.items():
        csv_info = profile.get("_csv")
        cp = profile.get("candidate_profile", {})
        # Prefer JSON name for word extraction (matches results file romanization better)
        orig = (cp.get("name") or "").strip() or (csv_info.get("name") if csv_info else None) or ""
        words = frozenset(w.upper() for w in re.findall(r"[A-Za-z]{3,}", orig)
                         if w.upper() not in ("MRS", "MR", "DR", "SH", "SHRI"))
        if words:
            wm.setdefault((ac_no, words), profile)
    return wm


def _lev(a: str, b: str) -> int:
    if a == b: return 0
    if not a: return len(b)
    if not b: return len(a)
    dp = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        prev, dp[0] = dp[0], i
        for j, cb in enumerate(b, 1):
            prev, dp[j] = dp[j], prev if ca == cb else 1 + min(prev, dp[j], dp[j - 1])
    return dp[len(b)]


def _fuzzy_dist(word: str, w: str) -> int:
    """Levenshtein treating adjacent transpositions as 1 edit."""
    d = _lev(word, w)
    if d == 2 and len(word) == len(w):
        diffs = [i for i in range(len(word)) if word[i] != w[i]]
        if len(diffs) == 2:
            i, j = diffs
            if j == i + 1 and word[i] == w[j] and word[j] == w[i]:
                return 1
    return d


def _word_in_set(word: str, word_set: frozenset) -> bool:
    """True if word matches any element: exact, prefix (both directions), or fuzzy."""
    if word in word_set:
        return True
    # Forward prefix: short abbreviation matches long word (e.g. "KSH" → "KSHETRIMAYUM")
    if len(word) >= 3:
        for w in word_set:
            if len(w) > len(word) and w.startswith(word):
                return True
    # Reverse prefix: long word starts with a shorter word (e.g. "JADUMANI" matches "JADU", "KSHETRIMAYUM" matches "KSH")
    if len(word) >= 6:
        for w in word_set:
            if len(w) >= 3 and word.startswith(w):
                return True
    # Fuzzy: 1 edit for 5+ chars, 2 edits for 7+ chars; transpositions treated as 1 edit
    if len(word) >= 5:
        tol = 2 if len(word) >= 7 else 1
        for w in word_set:
            if len(w) >= 5 and _fuzzy_dist(word, w) <= tol:
                return True
        # Also try with L→I substitution (Excel font sometimes renders I as l)
        word_li = word.replace("L", "I")
        if word_li != word:
            for w in word_set:
                if len(w) >= 5 and _fuzzy_dist(word_li, w) <= tol:
                    return True
    return False


def lookup_profile(ac_no, name, profile_map, word_map):
    """Exact → @/(alias)-stripped → bidirectional word-subset → 1-unmatched-word fallback."""
    prof = profile_map.get((ac_no, _normalize(name or "")))
    if prof:
        return prof
    # Strip @alias or (alias) suffix
    base = re.split(r"\s*[@(]", name)[0].strip() if name else ""
    if base != name:
        prof = profile_map.get((ac_no, _normalize(base)))
        if prof:
            return prof
    rwords = frozenset(w.upper() for w in _name_words(base))
    if not rwords:
        return None
    # Pass 1: all words in shorter name must match longer name
    for (k_ac, k_words), profile in word_map.items():
        if k_ac != ac_no:
            continue
        shorter = rwords if len(rwords) <= len(k_words) else k_words
        longer  = k_words if len(rwords) <= len(k_words) else rwords
        if (all(_word_in_set(w, longer) for w in shorter)
                and any(len(w) >= 4 for w in shorter)):
            return profile
    # Pass 2: allow 1 unmatched word — anchor word ≥6 chars, or ≥4 chars when unmatched word is a long name (8+ chars, likely abbreviated)
    for (k_ac, k_words), profile in word_map.items():
        if k_ac != ac_no:
            continue
        shorter = rwords if len(rwords) <= len(k_words) else k_words
        longer  = k_words if len(rwords) <= len(k_words) else rwords
        if len(shorter) < 2:
            continue
        matched   = [w for w in shorter if _word_in_set(w, longer)]
        unmatched = [w for w in shorter if not _word_in_set(w, longer)]
        if (len(matched) >= len(shorter) - 1
                and (any(len(w) >= 6 for w in matched)
                     or (any(len(w) >= 4 for w in matched)
                         and all(len(w) >= 8 for w in unmatched)))):
            return profile
    return None


def run():
    Base.metadata.create_all(bind=engine)

    name_to_ac  = load_ac_mapping()
    em          = load_electors()
    profile_map = load_profiles(name_to_ac)
    word_map    = build_word_map(profile_map)
    print(f"Loaded {len(em)} ACs, {len(profile_map)} profiles")

    # Results: STATE(0), District(1), AC_NO(2), AC_Name(3), Candidate(4),
    #          Sex(5), Age(6), Category(7), Party(8),
    #          General(9), Postal(10), Total(11), %(12), Electors(13)
    wb = openpyxl.load_workbook(RESULTS_FILE, read_only=True)
    result_rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()

    ac_data = {}
    for row in result_rows[1:]:
        if not isinstance(row[2], (int, float)):
            continue
        ac_no    = int(row[2])
        district = str(row[1]).strip() if row[1] else "Unknown"
        cat_raw  = str(row[7]).strip().upper() if row[7] else "GEN"
        category = "GEN" if cat_raw in ("GEN", "GENERAL") else cat_raw

        if ac_no not in ac_data:
            ac_data[ac_no] = {
                "district":             district,
                "name":                 str(row[3]).strip() if row[3] else "",
                "category":             category,
                "electors_from_results": int(row[13]) if row[13] else None,
                "candidates":           [],
            }
        sex_raw = str(row[5]).strip().upper() if row[5] else ""
        ac_data[ac_no]["candidates"].append({
            "name":         str(row[4]).strip().upper() if row[4] else None,
            "sex":          "M" if sex_raw in ("M", "MALE") else ("F" if sex_raw in ("F", "FEMALE") else None),
            "age":          int(row[6]) if row[6] else None,
            "party":        str(row[8]).strip() if row[8] else None,
            "gen_votes":    int(row[9])  if row[9]  else 0,
            "postal_votes": int(row[10]) if row[10] else 0,
            "total_votes":  int(row[11]) if row[11] else 0,
        })

    with Session(engine) as session:
        if session.query(Election).filter_by(state="Manipur", year=2012).first():
            print("Manipur 2012 already exists — skipping.")
            return

        election = Election(state="Manipur", year=2012, type="Assembly",
                            name="Manipur Legislative Assembly Election 2012")
        session.add(election)
        session.flush()

        # Build districts from results data
        all_districts = sorted(set(info["district"] for info in ac_data.values()))
        district_cache = {dname: District(election_id=election.id, name=dname)
                          for dname in all_districts}
        for d in district_cache.values():
            session.add(d)
        session.flush()

        constituency_cache = {}
        for ac_no, info in ac_data.items():
            elec = em.get(ac_no, {})
            dist = district_cache.get(info["district"])
            total_electors = elec.get("total_electors") or info["electors_from_results"]
            real = [c for c in info["candidates"] if c["party"] not in ("NOTA", None)
                    and c["name"] != "None of the Above"]
            real_sorted = sorted(real, key=lambda x: x["total_votes"], reverse=True)
            total_votes  = sum(c["total_votes"] for c in info["candidates"])
            turnout = round(total_votes / total_electors * 100, 2) if total_electors and total_votes else None
            margin  = (real_sorted[0]["total_votes"] - real_sorted[1]["total_votes"]
                       if len(real_sorted) >= 2 else None)
            category = elec.get("category") or info.get("category") or "GEN"
            con = Constituency(
                election_id=election.id, district_id=dist.id if dist else None,
                ac_no=ac_no, name=elec.get("name") or info["name"],
                category=category,
                total_electors=total_electors,
                male_electors=elec.get("male_electors"),
                female_electors=elec.get("female_electors"),
                third_gender_electors=elec.get("third_gender_electors"),
                total_votes_polled=total_votes, turnout_pct=turnout, winning_margin=margin,
            )
            session.add(con)
            constituency_cache[ac_no] = con
        session.flush()
        print(f"Created {len(constituency_cache)} constituencies")

        party_cache = {p.abbr.upper(): p for p in session.query(Party).all()}
        imported = 0

        for ac_no, info in ac_data.items():
            con  = constituency_cache[ac_no]
            real = [c for c in info["candidates"] if c["party"] not in ("NOTA", None)
                    and c["name"] != "None of the Above"]
            nota = [c for c in info["candidates"] if c["party"] == "NOTA"
                    or c["name"] == "None of the Above"]
            sorted_cands = sorted(real, key=lambda x: x["total_votes"], reverse=True) + nota
            total_valid  = sum(c["total_votes"] for c in sorted_cands)

            for pos, cand in enumerate(sorted_cands, start=1):
                is_nota = cand["party"] == "NOTA" or cand["name"] == "None of the Above"
                name    = "NOTA" if is_nota else cand["name"]

                party = None
                if not is_nota and cand["party"]:
                    abbr = cand["party"].strip()
                    if abbr.upper() not in party_cache:
                        p = Party(name=abbr, abbr=abbr, color=PARTY_COLORS.get(abbr))
                        session.add(p); session.flush()
                        party_cache[abbr.upper()] = p
                    party = party_cache[abbr.upper()]

                vote_pct  = round(cand["total_votes"] / total_valid * 100, 2) if total_valid else 0
                prof      = lookup_profile(ac_no, name or "", profile_map, word_map)
                csv_info  = prof.get("_csv") if prof else None
                cp        = (prof or {}).get("candidate_profile", {})
                assets_s  = (prof or {}).get("assets_summary", {})

                if not is_nota and prof:
                    if cp and cp.get("name", "").strip():
                        name = cp["name"].strip().title()
                    elif csv_info and csv_info.get("name", "").strip():
                        name = csv_info["name"].strip()

                education = declared_assets = liabilities_val = image_url = None
                criminal_cases = 0

                if csv_info:
                    edu = csv_info.get("education", "").strip()
                    if edu and edu not in ("Not mentioned", "Not Given", ""):
                        education = edu
                    criminal_cases = int(csv_info.get("criminal_cases") or 0)
                elif cp:
                    edu = (cp.get("education") or "").strip()
                    if edu.startswith("Category:"):
                        edu = edu[len("Category:"):].strip()
                    if edu and edu not in ("Not mentioned", "Not Given"):
                        education = edu
                    criminal_cases = parse_criminal(cp.get("crime_status"))

                # Assets/liabilities: CSV first, fall through to JSON when CSV has URL or returns 0
                a = parse_rupees(csv_info.get("assets")) if csv_info else 0
                if a > 0:
                    declared_assets = a
                elif assets_s:
                    a = parse_rupees(assets_s.get("total_assets"))
                    if a > 0: declared_assets = a

                l = parse_rupees(csv_info.get("liabilities")) if csv_info else 0
                if l > 0:
                    liabilities_val = l
                elif assets_s:
                    l = parse_rupees(assets_s.get("total_liabilities"))
                    if l > 0: liabilities_val = l

                if cp:
                    img = (cp.get("image_url") or "").strip()
                    if img and img != "None" and not img.endswith("/"):
                        image_url = img

                session.add(Candidate(
                    election_id=election.id, constituency_id=con.id,
                    party_id=party.id if party else None, name=name,
                    gender=cand["sex"], age=cand["age"], position=pos,
                    votes_general=cand["gen_votes"], votes_postal=cand["postal_votes"],
                    votes_total=cand["total_votes"], vote_pct=vote_pct, is_nota=is_nota,
                    education=education, declared_assets=declared_assets,
                    liabilities=liabilities_val, criminal_cases=criminal_cases,
                    image_url=image_url,
                ))
                imported += 1

        session.commit()
        election_id = election.id

    print(f"\n=== Manipur 2012 ===")
    print(f"Constituencies: {len(constituency_cache)}, Candidates: {imported}")
    with Session(engine) as session:
        winners = session.query(Candidate).filter_by(
            election_id=election_id, position=1, is_nota=False
        ).all()
        seats = {}
        for w in winners:
            p = w.party.abbr if w.party else "IND"
            seats[p] = seats.get(p, 0) + 1
        for p, s in sorted(seats.items(), key=lambda x: -x[1]):
            print(f"  {p}: {s} seats")


if __name__ == "__main__":
    run()
