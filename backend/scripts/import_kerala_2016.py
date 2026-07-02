"""Import Kerala 2016 election data into the database.

Data sources (Kerala/Kerala_2016/):
  - Detailed Results: Detailed Results.xlsx
  - Electors data:    Electors & Voters Data.xlsx
  - Party list:       List Of Political Parties Participated.xlsx
  - Candidate CSV:    Kerala_candidates_2016.csv
  - Profile JSON:     2016_candidate_profile.json
"""
import sys
import os
import re
import csv
import json

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from app.database import engine, Base
from app.models.kerala import Election, District, Constituency, Party, Candidate

PROJECT_ROOT    = os.path.join(os.path.dirname(__file__), "..", "..")
DATA_DIR        = os.path.join(PROJECT_ROOT, "Kerala", "Kerala_2016")
DETAILED_RESULTS = os.path.join(DATA_DIR, "Detailed Results.xlsx")
ELECTORS_FILE   = os.path.join(DATA_DIR, "Electors & Voters Data.xlsx")
PARTY_LIST      = os.path.join(DATA_DIR, "List Of Political Parties Participated.xlsx")
PROFILES_JSON   = os.path.join(DATA_DIR, "2016_candidate_profile.json")
CANDIDATES_CSV  = os.path.join(DATA_DIR, "Kerala_candidates_2016.csv")
MAPPING_FILE    = os.path.join(PROJECT_ROOT, "Kerala", "Kerala_Election_Data_Mapping.xlsx")

PARTY_COLORS = {
    "BJP": "#FF9933", "INC": "#00BFFF", "CPI(M)": "#FF0000", "CPM": "#FF0000",
    "CPI": "#FF6600", "IUML": "#006400", "JD(S)": "#FFD700", "NCP": "#004080",
    "BSP": "#0000FF", "IND": "#808080", "NOTA": "#000000", "AAP": "#0066CC",
    "KC(M)": "#8B008B", "KEC(M)": "#8B008B", "RSP": "#FF4444", "AIFB": "#CC0000",
}

# 2016 uses "CPM" as abbreviation — normalise to canonical CPI(M)
ABBR_ALIASES = {"CPM": "CPI(M)"}


def load_ac_category_map():
    """Parse 2016 AC column from mapping file → ac_no: category."""
    wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True)
    ws = wb["AC"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    cat_map = {}
    for row in rows[2:]:
        ac_no_val = row[3]   # 2016 AC NO column
        ac_name_val = row[4]  # 2016 AC NAME column
        if ac_no_val is None or ac_name_val is None:
            continue
        try:
            ac_no = int(ac_no_val)
        except (ValueError, TypeError):
            continue
        name = str(ac_name_val).strip()
        category = "GEN"
        if "(SC)" in name:
            category = "SC"
        elif "(ST)" in name:
            category = "ST"
        cat_map[ac_no] = category
    return cat_map


def load_electors_map():
    """Parse electors file → ac_no: {male, female, third_gender, total, total_votes}."""
    wb = openpyxl.load_workbook(ELECTORS_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    electors = {}
    for row in rows[1:]:
        ac_no_val = row[2]
        if ac_no_val is None:
            continue
        try:
            ac_no = int(ac_no_val)
        except (ValueError, TypeError):
            continue
        electors[ac_no] = {
            "male": int(row[4]) if row[4] else 0,
            "female": int(row[5]) if row[5] else 0,
            "third_gender": int(row[6]) if row[6] else 0,
            "total": int(row[7]) if row[7] else 0,
            "total_votes": int(row[13]) if row[13] else 0,
        }
    return electors


def load_party_names():
    """2016 format: Party Type, Party Name, Party Abbreviation (cols 0,1,2)."""
    wb = openpyxl.load_workbook(PARTY_LIST, read_only=True)
    ws = wb[wb.sheetnames[0]]
    mapping = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] and row[1] and row[2] and str(row[0]).strip() != "Party Type":
            abbr = str(row[2]).strip()
            name = str(row[1]).strip()
            canonical = ABBR_ALIASES.get(abbr, abbr)
            mapping[canonical] = name
    wb.close()
    return mapping


def _parse_rupees(val):
    if not val or "Nil" in str(val):
        return 0
    s = str(val).strip()
    if s.startswith("http"):
        return 0
    digits = re.sub(r"[^\d]", "", s.split("~")[0])
    return int(digits) if digits else 0


def _parse_criminal(status):
    if not status or "No criminal" in str(status):
        return 0
    m = re.search(r"(\d+)", str(status))
    return int(m.group(1)) if m else 0


def load_profiles():
    """Build profile map keyed by (constituency_upper, name_upper)."""
    csv_by_id = {}
    with open(CANDIDATES_CSV, encoding="latin-1") as f:
        for row in csv.DictReader(f):
            cid = row.get("candidate_id", "").strip()
            con = row.get("constituency", "").strip().upper()
            if cid:
                csv_by_id[cid] = {
                    "name":           row.get("candidate_name", "").strip(),
                    "constituency":   con,
                    "criminal_cases": int(row.get("criminal_cases") or 0),
                    "education":      row.get("education", "").strip(),
                    "assets":         row.get("assets", ""),
                    "liabilities":    row.get("liabilities", ""),
                }

    with open(PROFILES_JSON, encoding="utf-8") as f:
        profiles_json = json.load(f)

    by_key = {}
    for p in profiles_json:
        cp = p.get("candidate_profile", {})
        cid = str(cp.get("candidate_id", ""))
        info = csv_by_id.get(cid)
        if info:
            key = (info["constituency"], info["name"].upper())
            by_key[key] = {**p, "_csv": info}
    return by_key


def import_data():
    Base.metadata.create_all(bind=engine)

    cat_map = load_ac_category_map()
    electors_map = load_electors_map()
    party_names = load_party_names()

    # 2016 columns: District(0), AC No(1), AC Name(2), Candidate No(3), Candidate Name(4),
    # Candidate Sex(5), Candidate Age(6), Candidate Category(7), Party Name(8) [abbr],
    # VALID VOTES in General(9), VALID VOTES in Postal(10), Total Valid Votes(11),
    # Total Electors(12), Total Votes(13)
    wb = openpyxl.load_workbook(DETAILED_RESULTS, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    data_rows = rows[1:]

    with Session(engine) as session:
        election = Election(
            state="Kerala",
            year=2016,
            type="Assembly",
            name="Kerala Legislative Assembly Election 2016",
        )
        session.add(election)
        session.flush()
        print(f"Created election: {election.name} (id={election.id})")

        party_cache = {p.abbr.upper(): p for p in session.query(Party).all()}
        for abbr, full_name in party_names.items():
            key = abbr.upper()
            if key not in party_cache:
                p = Party(name=full_name, abbr=abbr, color=PARTY_COLORS.get(abbr))
                session.add(p)
                session.flush()
                party_cache[key] = p
        print(f"Loaded {len(party_cache)} parties")

        district_cache = {}
        ac_groups: dict[int, list] = {}
        ac_meta: dict[int, dict] = {}

        for row in data_rows:
            if row[1] is None:
                continue
            try:
                ac_no = int(row[1])
            except (ValueError, TypeError):
                continue
            district_raw = str(row[0]).strip().title() if row[0] else "Unknown"
            ac_name_raw = str(row[2]).strip().upper() if row[2] else ""
            if ac_no not in ac_meta:
                ac_meta[ac_no] = {"district": district_raw, "ac_name": ac_name_raw}
            if ac_no not in ac_groups:
                ac_groups[ac_no] = []
            ac_groups[ac_no].append(row)

        for ac_no in sorted(ac_groups.keys()):
            meta = ac_meta[ac_no]
            cand_rows = ac_groups[ac_no]
            district_name = meta["district"]
            ac_name = meta["ac_name"]

            if district_name not in district_cache:
                d = District(election_id=election.id, name=district_name)
                session.add(d)
                session.flush()
                district_cache[district_name] = d

            district = district_cache[district_name]
            el_info = electors_map.get(ac_no, {})

            valid_cands = [(r, int(r[11]) if r[11] else 0) for r in cand_rows if r[4]]
            valid_cands.sort(key=lambda x: -x[1])

            winning_margin = None
            if len(valid_cands) >= 2:
                winning_margin = valid_cands[0][1] - valid_cands[1][1]

            total_votes = el_info.get("total_votes") or (int(cand_rows[0][13]) if cand_rows and cand_rows[0][13] else None)
            total_electors = el_info.get("total") or (int(cand_rows[0][12]) if cand_rows and cand_rows[0][12] else None)
            turnout_pct = round(total_votes / total_electors * 100, 2) if total_electors and total_votes else None

            con = Constituency(
                election_id=election.id,
                district_id=district.id,
                ac_no=ac_no,
                name=ac_name,
                category=cat_map.get(ac_no, "GEN"),
                total_electors=total_electors,
                male_electors=el_info.get("male"),
                female_electors=el_info.get("female"),
                third_gender_electors=el_info.get("third_gender"),
                total_votes_polled=total_votes,
                turnout_pct=turnout_pct,
                winning_margin=winning_margin,
            )
            session.add(con)
            session.flush()

            for pos, (row, _) in enumerate(valid_cands, start=1):
                party_abbr = str(row[8]).strip() if row[8] else None
                party_abbr = ABBR_ALIASES.get(party_abbr, party_abbr)
                cand_name = str(row[4]).strip()
                is_nota = bool(cand_name.lower() == "none of the above" or (party_abbr and party_abbr.upper() == "NOTA"))
                sex_raw = str(row[5]).strip().upper() if row[5] else None
                gender = "MALE" if sex_raw in ("M", "MALE") else ("FEMALE" if sex_raw in ("F", "FEMALE") else sex_raw)
                total_c = int(row[11]) if row[11] else 0
                vote_pct = round(total_c / total_votes * 100, 2) if total_votes and total_votes > 0 else None

                party = None
                if not is_nota and party_abbr:
                    key = party_abbr.upper()
                    if key not in party_cache:
                        p = Party(name=party_abbr, abbr=party_abbr, color=PARTY_COLORS.get(party_abbr))
                        session.add(p)
                        session.flush()
                        party_cache[key] = p
                    party = party_cache[key]

                session.add(Candidate(
                    election_id=election.id,
                    constituency_id=con.id,
                    party_id=party.id if party else None,
                    name=cand_name,
                    gender=gender,
                    age=int(row[6]) if row[6] else None,
                    position=pos,
                    votes_general=int(row[9]) if row[9] else 0,
                    votes_postal=int(row[10]) if row[10] else 0,
                    votes_total=total_c,
                    vote_pct=vote_pct,
                    is_nota=is_nota,
                ))

        session.commit()

        n_c = session.query(Constituency).filter_by(election_id=election.id).count()
        n_cand = session.query(Candidate).filter_by(election_id=election.id).count()
        print(f"\n=== Import Summary ===")
        print(f"Districts: {len(district_cache)}, Constituencies: {n_c}, Candidates: {n_cand}")


def enrich_profiles():
    """Enrich 2016 candidates with education, assets, liabilities, occupation, image."""
    profile_map = load_profiles()
    print(f"Loaded {len(profile_map)} profiles")

    with Session(engine) as session:
        el = session.query(Election).filter_by(state="Kerala", year=2016).first()
        if not el:
            print("Kerala 2016 not found — run import_data() first")
            return

        updated = missing = 0
        for con in session.query(Constituency).filter_by(election_id=el.id).all():
            for cand in session.query(Candidate).filter_by(constituency_id=con.id, is_nota=False).all():
                key = (con.name.upper(), cand.name.upper())
                prof = profile_map.get(key)
                if not prof:
                    missing += 1
                    continue

                cp = prof.get("candidate_profile", {})
                csv_info = prof.get("_csv", {})
                assets_s = prof.get("assets_summary", {})

                edu = (csv_info.get("education") or cp.get("education") or "").strip()
                if edu.startswith("Category:"):
                    edu = edu[len("Category:"):].strip()
                cand.education = edu if edu and edu not in ("Not mentioned", "Not Given", "") else None

                cand.criminal_cases = int(csv_info.get("criminal_cases") or 0) or _parse_criminal(cp.get("crime_status"))

                assets_raw = csv_info.get("assets", "")
                if not assets_raw or str(assets_raw).startswith("http"):
                    assets_val = _parse_rupees(assets_s.get("total_assets"))
                else:
                    assets_val = _parse_rupees(assets_raw)
                cand.declared_assets = assets_val if assets_val > 0 else None

                liab_raw = csv_info.get("liabilities", "")
                if not liab_raw or str(liab_raw).startswith("http"):
                    liab_val = _parse_rupees(assets_s.get("total_liabilities"))
                else:
                    liab_val = _parse_rupees(liab_raw)
                cand.liabilities = liab_val if liab_val > 0 else None

                occ = (prof.get("profession", {}).get("self") or "").strip()
                cand.occupation = occ if occ and occ != "Not mentioned" else None

                img = (cp.get("image_url") or "").strip()
                cand.image_url = img if img and img != "None" else None

                updated += 1

        session.commit()
        print(f"Enriched {updated} candidates, {missing} not matched")


if __name__ == "__main__":
    import_data()
    enrich_profiles()
