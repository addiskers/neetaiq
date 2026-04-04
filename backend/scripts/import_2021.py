"""
Import Assam 2021 election data.
Sources: assam_ac_summary.xlsx, assam_candidate_results.xlsx, 10. Detailed Results.xlsx
"""
import sys
import os
import re
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from app.database import engine, Base, SessionLocal
from app.models import Election, District, Constituency, Party, Candidate

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "assam2021")

# 2021 district name -> normalized name (use 2021 names as-is, they are what they are)
# We store them under their original 2021 names since this is the 2021 election
DISTRICT_2021_TO_2026 = {
    "Kamrup": "Kamrup",
    "Kamrup Metropolitan": "Kamrup Metropolitan",
    "Karimganj": "Karimganj",
    "Marigaon": "Marigaon",
    "Sibsagar": "Sibsagar",
}


def import_2021():
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    try:
        # Check if 2021 already exists
        existing = db.query(Election).filter_by(year=2021, state="Assam").first()
        if existing:
            print("Assam 2021 already imported. Skipping.")
            return

        # Step 1: Create election
        print("\n[1/5] Creating 2021 election record...")
        election = Election(
            year=2021,
            election_name="Assam Legislative Assembly Election 2021",
            election_type="AC - GENERAL",
            state="Assam",
        )
        db.add(election)
        db.flush()
        print(f"  Election ID: {election.id}")

        # Step 2: Load AC summary (has district names + constituency data)
        print("\n[2/5] Importing districts & constituencies from ac_summary...")
        ac_summary = _load_ac_summary()
        district_map = _import_districts_2021(db, election.id, ac_summary)
        constituency_map = _import_constituencies_2021(db, election.id, ac_summary, district_map)
        print(f"  Districts: {len(district_map)}, Constituencies: {len(constituency_map)}")

        # Step 3: Ensure parties exist
        print("\n[3/5] Ensuring parties exist...")
        party_map = _ensure_parties(db)
        print(f"  Total parties: {len(party_map)}")

        # Step 4: Import candidates from Detailed Results (has age, sex, category, votes)
        print("\n[4/5] Importing candidates from Detailed Results...")
        cand_count = _import_candidates_2021(db, election.id, constituency_map, party_map)
        print(f"  Imported {cand_count} candidates")

        # Step 5: Mark winners from ac_summary
        print("\n[5/5] Marking winners...")
        _mark_winners(db, election.id, ac_summary, constituency_map)

        db.commit()
        print("\n=== 2021 Import Complete! ===")
        _print_summary(db, election.id)

    except Exception as e:
        db.rollback()
        print(f"\nERROR: {e}")
        raise
    finally:
        db.close()


def _load_ac_summary() -> list:
    """Load assam_ac_summary.xlsx into list of dicts."""
    path = os.path.join(DATA_DIR, "assam_ac_summary.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    headers = rows[0]
    data = []
    for r in rows[1:]:
        d = dict(zip(headers, r))
        data.append(d)
    wb.close()
    return data


def _import_districts_2021(db: Session, election_id: int, ac_summary: list) -> dict:
    """Create districts for 2021 election. Returns {district_name_lower: district_id}."""
    district_names = set()
    for row in ac_summary:
        dn = str(row.get("district", "")).strip()
        if dn:
            district_names.add(dn)

    district_map = {}
    for name in sorted(district_names):
        district = District(
            election_id=election_id,
            name=name,
            name_previous=None,
            rename_status=None,
        )
        db.add(district)
        db.flush()
        district_map[name.lower()] = district.id

    return district_map


def _parse_int(val) -> int | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).replace(",", "").replace(" ", "").strip()
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _parse_float(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").replace("%", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _import_constituencies_2021(db: Session, election_id: int, ac_summary: list, district_map: dict) -> dict:
    """Import constituencies from ac_summary. Returns {ac_number: constituency_id}."""
    constituency_map = {}
    for row in ac_summary:
        ac_no = int(row["ac_no"])
        ac_name = str(row["ac_name"]).strip()
        district_name = str(row["district"]).strip()
        total_electors = _parse_int(row.get("total_electors"))
        total_votes = _parse_int(row.get("total_votes"))
        poll_pct = _parse_float(row.get("poll_percent"))
        margin = _parse_int(row.get("margin"))
        ac_type = str(row.get("type", "")).strip()
        winner_party = str(row.get("party", "")).strip()

        district_id = district_map.get(district_name.lower())
        if not district_id:
            print(f"  WARNING: District '{district_name}' not found for AC {ac_no} {ac_name}")
            continue

        # Compute male/female from detailed data later, for now use total
        constituency = Constituency(
            election_id=election_id,
            district_id=district_id,
            ac_number=ac_no,
            name=ac_name,
            total_electors=total_electors,
            turnout_percentage=poll_pct,
            ruling_party=winner_party,
            swing_status="Safe" if (margin and margin > 20000) else "Swing" if margin else "Stable",
        )
        db.add(constituency)
        db.flush()
        constituency_map[ac_no] = constituency.id

    return constituency_map


def _ensure_parties(db: Session) -> dict:
    """Get existing parties + add any missing from 2021 data. Returns {name: party_id}."""
    # Load existing
    existing = db.query(Party).all()
    party_map = {p.name: p.id for p in existing}

    # Load 2021 parties from detailed results
    path = os.path.join(DATA_DIR, "10. Detailed Results.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))

    # Also load party list file for abbreviation mapping
    path2 = os.path.join(DATA_DIR, "3. List Of Political Parties Participated.xlsx")
    wb2 = openpyxl.load_workbook(path2, read_only=True)
    rows2 = list(wb2.active.iter_rows(values_only=True))
    abbr_to_full = {}
    for r in rows2[1:]:
        if r[1] and r[2]:
            abbr_to_full[str(r[1]).strip()] = str(r[2]).strip()
    wb2.close()

    # Get unique parties from detailed results (uses abbreviations)
    party_abbrs = set()
    for r in rows[4:]:  # data starts at row 4
        if r[7]:
            party_abbrs.add(str(r[7]).strip())

    for abbr in sorted(party_abbrs):
        if abbr in ("NOTA",):
            continue
        full_name = abbr_to_full.get(abbr, abbr)
        # Check if already exists by full name or abbreviation
        if full_name not in party_map:
            # Check by short_name
            existing_p = db.query(Party).filter(Party.short_name == abbr).first()
            if existing_p:
                party_map[full_name] = existing_p.id
                party_map[abbr] = existing_p.id
            else:
                p = Party(name=full_name, short_name=abbr, color="#A9A9A9")
                db.add(p)
                db.flush()
                party_map[full_name] = p.id
                party_map[abbr] = p.id
                print(f"  Added party: {abbr} = {full_name}")
        else:
            party_map[abbr] = party_map[full_name]

    wb.close()
    return party_map


def _import_candidates_2021(db: Session, election_id: int, constituency_map: dict, party_map: dict) -> int:
    """Import candidates from '10. Detailed Results.xlsx'."""
    path = os.path.join(DATA_DIR, "10. Detailed Results.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))

    count = 0
    for r in rows[4:]:  # data starts at row index 4
        ac_no = r[1]
        if not ac_no or not isinstance(ac_no, int):
            continue

        raw_name = str(r[3]).strip() if r[3] else None
        if not raw_name:
            continue

        # Name format: "1 BIJOY MALAKAR" - strip leading number
        name = re.sub(r'^\d+\s+', '', raw_name).strip()
        if name == "NOTA":
            continue

        sex = str(r[4]).strip() if r[4] else None
        age = int(r[5]) if r[5] and isinstance(r[5], (int, float)) else None
        category = str(r[6]).strip() if r[6] else None
        party_abbr = str(r[7]).strip() if r[7] else None
        total_votes = int(r[11]) if r[11] and isinstance(r[11], (int, float)) else None
        vote_pct = float(r[12]) if r[12] and isinstance(r[12], (int, float)) else None

        constituency_id = constituency_map.get(ac_no)
        if not constituency_id:
            continue

        party_id = party_map.get(party_abbr)
        if not party_id:
            # Try full name lookup
            for k, v in party_map.items():
                if party_abbr and party_abbr.lower() in k.lower():
                    party_id = v
                    break
            if not party_id:
                print(f"  WARNING: Party '{party_abbr}' not found for {name}")
                continue

        gender = None
        if sex:
            if sex.upper() in ("MALE", "M"):
                gender = "Male"
            elif sex.upper() in ("FEMALE", "F"):
                gender = "Female"
            else:
                gender = "Other"

        candidate = Candidate(
            election_id=election_id,
            constituency_id=constituency_id,
            party_id=party_id,
            name=name,
            status="Accepted",
            is_contesting=True,
            age=age,
            gender=gender,
            declared_assets=total_votes,  # Store votes in declared_assets temporarily? No, use a better approach
        )
        # Actually, we don't have assets for 2021 from this file. Just store candidate info.
        candidate.declared_assets = None
        db.add(candidate)
        count += 1

    db.flush()
    wb.close()
    return count


def _mark_winners(db: Session, election_id: int, ac_summary: list, constituency_map: dict):
    """Mark winning candidates as incumbent."""
    for row in ac_summary:
        ac_no = int(row["ac_no"])
        winner_name = str(row.get("winning_candidate", "")).strip().upper()
        if not winner_name:
            continue

        constituency_id = constituency_map.get(ac_no)
        if not constituency_id:
            continue

        # Find the candidate
        candidates = db.query(Candidate).filter(
            Candidate.constituency_id == constituency_id,
            Candidate.election_id == election_id,
        ).all()

        for c in candidates:
            if c.name.upper() == winner_name or winner_name in c.name.upper() or c.name.upper() in winner_name:
                c.is_incumbent = True
                break


def _print_summary(db: Session, election_id: int):
    print(f"\n--- Summary for Election ID {election_id} (2021) ---")
    print(f"  Districts:       {db.query(District).filter_by(election_id=election_id).count()}")
    print(f"  Constituencies:  {db.query(Constituency).filter_by(election_id=election_id).count()}")
    print(f"  Parties:         {db.query(Party).count()}")
    print(f"  Candidates:      {db.query(Candidate).filter_by(election_id=election_id).count()}")
    winners = db.query(Candidate).filter_by(election_id=election_id, is_incumbent=True).count()
    print(f"  Winners marked:  {winners}")


if __name__ == "__main__":
    import_2021()
