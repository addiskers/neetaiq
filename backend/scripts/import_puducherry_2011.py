"""Import Puducherry 2011 election data.

Data sources (all under Puducherry/Puducherry_2011/):
  - Results: Detailed_Results_2011.xlsx
  - Electors: District_Constituency wise Electors & Voters Data_2011.xlsx
  - Party names: List Of Political Parties Participated_2011.xlsx
  - Candidate profiles: my_neta_2011_candidate_profile.json
  - Candidate bridge: my_neta_puducherry_candidates_2011.csv

Usage: cd backend && python -m scripts.import_puducherry_2011
"""
import sys, os, re, json, csv
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from app.database import engine, Base
from app.models.puducherry import Election, District, Constituency, Party, Candidate

PROJECT_ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
DATA_DIR = os.path.join(PROJECT_ROOT, "Puducherry", "Puducherry_2011")

RESULTS_FILE   = os.path.join(DATA_DIR, "Detailed_Results_2011.xlsx")
ELECTORS_FILE  = os.path.join(DATA_DIR, "District_Constituency wise Electors & Voters Data_2011.xlsx")
PARTIES_FILE   = os.path.join(DATA_DIR, "List Of Political Parties Participated_2011.xlsx")
PROFILES_JSON  = os.path.join(DATA_DIR, "my_neta_2011_candidate_profile.json")
CANDIDATES_CSV = os.path.join(DATA_DIR, "my_neta_puducherry_candidates_2011.csv")
MAPPING_FILE   = os.path.join(PROJECT_ROOT, "puducherry", "puducherry_checking.xlsx")

PARTY_COLORS = {
    "AINRC": "#800080", "DMK": "#FF0000", "AIADMK": "#228B22", "INC": "#00BFFF",
    "BJP": "#FF9933", "PMK": "#FFFF00", "CPI": "#FF4444", "CPI(M)": "#FF0000",
    "BSP": "#0000FF", "IND": "#808080", "JD(U)": "#003366",
}

ABBR_ALIASES = {"CPM": "CPI(M)"}

# ACs 1-23: Puducherry, 24-28: Karaikal, 29: Mahe, 30: Yanam
AC_DISTRICT = {**{n: "Puducherry" for n in range(1, 24)},
               **{n: "Karaikal" for n in range(24, 29)},
               29: "Mahe", 30: "Yanam"}

CANDIDATE_NAME_FIXES = {
    ("MANNADIPET", "D. GUNASHAKARAN"): "D. GUNASEKARAN",
    ("OUSSUDU", "M. PITCHEKARANE @ PITCHAIAPPAN"): "M.PITCHAIAPPANE @ PITCHAIKARANE",
    ("VILLIANUR", "NADARAJAN K."): "K. NADARAJAN",
    ("VILLIANUR", "K. ATHIRIYEN"): "K. ATHIRAYEN",
    ("OZHUKARAI", "N.G. PANNIR SELVAM"): "N.G.PANNEERSELVAM",
    ("THATTANCHAVADY", "MASTHANJI MOHAMED KALIMULLA"): "MASTHANJI MOHAMMED KALIMULLA",
    ("MUTHIALPET", "KRISHNAKANTHAN @ BASKAR"): "KRISHNAKANTH @ BHASKAR",
    ("MUTHIALPET", "NANDHA T. SARAVANAN"): "NANDA T. SARAVANAN",
    ("MUTHIALPET", "R. KUMARAN @ PAVADAISAMI"): "R. KUMARAN @ PAVADAISAMY",
    ("OUPALAM", "ANNIBAL KENNEDY"): "ANIBAL KENNEDY",
    ("OUPALAM", "ANGAPPIN @ ROUSSEAU ANGAPPIN"): "ANGAPPIN @ ROUSSEAU ANGAPPIN. S",
    ("OUPALAM", "S. ANTHONI"): "ANTHONI. S",
    ("ARIANKUPPAM", "V. SABAPATHY"): "V. SABAPATHY @ KOTHANDARAMAN",
    ("LAWSPET", "G. SATHIARAJ"): "G. SATHIRAJ",
    ("MUTHIALPET", "S. MOTHILAL"): "S. MOTILAL",
}

def _normalize(name):
    return re.sub(r"[.\s@]+", "", name).upper()


def load_ac_mapping():
    """Load puducherry_checking.xlsx and return:
    - name_to_ac: {normalized_name: ac_no} covering all 4 year name variants
    """
    wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True)
    ws = wb["Ac Name"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    name_to_ac = {}
    for row in rows[1:]:
        if not row[0]:
            continue
        ac_no = int(row[0])
        for col in [1, 4, 7, 10]:  # 2011, 2016, 2021, 2026 name columns
            if col < len(row) and row[col]:
                v = str(row[col]).strip()
                # Add both with and without SC/ST suffix
                stripped = re.sub(r"\s*\(?(SC|ST)\)?\s*$", "", v, flags=re.I).strip()
                name_to_ac[_normalize(v)] = ac_no
                name_to_ac[_normalize(stripped)] = ac_no
    return name_to_ac


def load_electors():
    wb = openpyxl.load_workbook(ELECTORS_FILE, read_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()
    em = {}
    for row in rows[1:]:
        if row[1] is None:
            continue
        ac_no = int(row[1])
        cat = str(row[4]).strip().upper() if row[4] else "GEN"
        em[ac_no] = {
            "name": re.sub(r"\s+", " ", str(row[2]).strip()).upper() if row[2] else "",
            "category": "GEN" if cat == "GENERAL" else cat,
            "male_electors": int(row[5]) if row[5] else None,
            "female_electors": int(row[6]) if row[6] else None,
            "third_gender_electors": int(row[7]) if row[7] else None,
            "total_electors": int(row[8]) if row[8] else None,
        }
    return em


def load_party_names():
    wb = openpyxl.load_workbook(PARTIES_FILE, read_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()
    return {str(r[2]).strip(): str(r[1]).strip() for r in rows[1:] if r[1] and r[2]}


def load_affidavit_bridge(name_to_ac):
    """Load profiles keyed by (ac_no, normalized_candidate_name) for reliable matching."""
    csv_by_id = {}
    with open(CANDIDATES_CSV, "r", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            cid = row.get("candidate_id", "").strip()
            con_raw = re.sub(r"\s*\(?(SC|ST)\)?\s*$", "", row.get("constituency", "").strip(), flags=re.I).strip()
            if cid:
                csv_by_id[cid] = {"name": row.get("candidate_name", "").strip(), "constituency": con_raw}

    with open(PROFILES_JSON, "r", encoding="utf-8") as f:
        profiles = json.load(f)

    by_acno = {}  # (ac_no, norm_name) → profile
    for p in profiles:
        cp = p.get("candidate_profile", {})
        cid = str(cp.get("candidate_id", ""))
        info = csv_by_id.get(cid)

        ac_no = None
        cand_name = None
        if info:
            ac_no = name_to_ac.get(_normalize(info["constituency"]))
            cand_name = info["name"]
        if not ac_no:
            ac_raw = re.sub(r"\s*\(?(SC|ST)\)?\s*$", "", (cp.get("ac_name") or "").strip(), flags=re.I).strip()
            ac_no = name_to_ac.get(_normalize(ac_raw))
            if not cand_name:
                cand_name = (cp.get("name") or "").strip()
        if ac_no and cand_name:
            by_acno.setdefault((ac_no, _normalize(cand_name)), p)

    return by_acno


def parse_rupees(val):
    if not val or "Nil" in str(val):
        return 0
    cleaned = re.sub(r"[Rs\s\xa0,~]", "", str(val))
    digits = re.match(r"^([\d.]+)", cleaned)
    return int(float(digits.group(1))) if digits else 0


def parse_criminal(status):
    if not status or "No criminal" in status:
        return 0
    m = re.search(r"(\d+)", status)
    return int(m.group(1)) if m else 0


def run():
    Base.metadata.create_all(bind=engine)

    em = load_electors()
    party_names = load_party_names()
    name_to_ac = load_ac_mapping()
    profile_map = load_affidavit_bridge(name_to_ac)
    print(f"Loaded {len(em)} ACs, {len(party_names)} party names, {len(profile_map)} profiles")

    wb = openpyxl.load_workbook(RESULTS_FILE, read_only=True)
    result_rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()
    # Cols: AC NO(0), Constituency(1), Total Electors(2), Candidate(3),
    #       Sex(4), Age(5), Category(6), Party(7), General(8), Postal(9), Total(10), %(11)

    with Session(engine) as session:
        if session.query(Election).filter_by(state="Puducherry", year=2011).first():
            print("Puducherry 2011 already exists — skipping.")
            return

        election = Election(state="Puducherry", year=2011, type="Assembly",
                            name="Puducherry Legislative Assembly Election 2011")
        session.add(election)
        session.flush()

        district_cache = {dname: District(election_id=election.id, name=dname)
                          for dname in sorted(set(AC_DISTRICT.values()))}
        for d in district_cache.values():
            session.add(d)
        session.flush()

        # Group by AC
        ac_data = {}
        for row in result_rows[1:]:
            if not isinstance(row[0], (int, float)):
                continue
            ac_no = int(row[0])
            if ac_no not in ac_data:
                cat = str(row[6]).strip().upper() if row[6] else "GEN"
                ac_data[ac_no] = {"name": str(row[1]).strip(), "category": "GEN" if cat == "GENERAL" else cat,
                                  "electors_from_results": int(row[2]) if row[2] else None, "candidates": []}
            sex = str(row[4]).strip().upper() if row[4] else ""
            ac_data[ac_no]["candidates"].append({
                "name": str(row[3]).strip() if row[3] else None,
                "sex": "M" if "MALE" in sex else ("F" if "FEMALE" in sex else None),
                "age": int(row[5]) if row[5] else None,
                "party": str(row[7]).strip() if row[7] else None,
                "gen_votes": int(row[8]) if row[8] else 0,
                "postal_votes": int(row[9]) if row[9] else 0,
                "total_votes": int(row[10]) if row[10] else 0,
            })

        constituency_cache = {}
        for ac_no, info in ac_data.items():
            elec_data = em.get(ac_no, {})
            dist = district_cache.get(AC_DISTRICT.get(ac_no, "Puducherry"))
            total_electors = elec_data.get("total_electors") or info["electors_from_results"]
            total_votes = sum(c["total_votes"] for c in info["candidates"])
            turnout = round(total_votes / total_electors * 100, 2) if total_electors and total_votes else None
            real = sorted([c for c in info["candidates"]
                           if c["party"] != "NOTA" and c["name"] != "None of the Above"],
                          key=lambda x: x["total_votes"], reverse=True)
            margin = real[0]["total_votes"] - real[1]["total_votes"] if len(real) >= 2 else None
            con = Constituency(
                election_id=election.id, district_id=dist.id if dist else None, ac_no=ac_no,
                name=re.sub(r"\s+", " ", (elec_data.get("name") or info["name"]).upper()),
                category=elec_data.get("category") or info["category"],
                total_electors=total_electors, male_electors=elec_data.get("male_electors"),
                female_electors=elec_data.get("female_electors"),
                third_gender_electors=elec_data.get("third_gender_electors"),
                total_votes_polled=total_votes, turnout_pct=turnout, winning_margin=margin,
            )
            session.add(con)
            constituency_cache[ac_no] = con
        session.flush()
        print(f"Created {len(constituency_cache)} constituencies")

        party_cache = {p.abbr.upper(): p for p in session.query(Party).all()}
        imported = 0

        for ac_no, info in ac_data.items():
            con = constituency_cache[ac_no]
            sorted_cands = sorted(info["candidates"], key=lambda x: x["total_votes"], reverse=True)
            total_valid = sum(c["total_votes"] for c in sorted_cands)

            for pos, cand in enumerate(sorted_cands, start=1):
                is_nota = cand["party"] == "NOTA" or cand["name"] == "None of the Above"
                name = "NOTA" if cand["name"] == "None of the Above" else cand["name"]

                party = None
                if not is_nota and cand["party"]:
                    raw = cand["party"]
                    abbr = ABBR_ALIASES.get(raw, raw)
                    if abbr.upper() not in party_cache:
                        full_name = party_names.get(raw, raw)
                        p = Party(name=full_name, abbr=abbr,
                                  color=PARTY_COLORS.get(raw) or PARTY_COLORS.get(abbr))
                        session.add(p)
                        session.flush()
                        party_cache[abbr.upper()] = p
                    party = party_cache[abbr.upper()]

                vote_pct = round(cand["total_votes"] / total_valid * 100, 2) if total_valid else 0

                aff = profile_map.get((ac_no, _normalize(name or "")))
                if not aff:
                    fixed = CANDIDATE_NAME_FIXES.get((con.name.upper(), (name or "").upper()))
                    if fixed:
                        aff = profile_map.get((ac_no, _normalize(fixed)))

                education = occupation = declared_assets = liabilities_val = image_url = None
                criminal_cases = 0
                if aff:
                    cp = aff.get("candidate_profile", {})
                    edu = (cp.get("education") or "").strip()
                    if edu.startswith("Category:"):
                        edu = edu[len("Category:"):].strip()
                    if edu and edu not in ("Not mentioned", "Not Given"):
                        education = edu
                    occ = (aff.get("profession", {}).get("self") or "").strip()
                    if occ and occ not in ("Not mentioned", "Not Given"):
                        occupation = occ
                    a = parse_rupees(aff.get("assets_summary", {}).get("total_assets"))
                    if a > 0:
                        declared_assets = a
                    l = parse_rupees(aff.get("assets_summary", {}).get("total_liabilities"))
                    if l > 0:
                        liabilities_val = l
                    criminal_cases = parse_criminal(cp.get("crime_status"))
                    img = (cp.get("image_url") or "").strip()
                    if img and img != "None":
                        image_url = img

                session.add(Candidate(
                    election_id=election.id, constituency_id=con.id,
                    party_id=party.id if party else None, name=name,
                    gender=cand["sex"], age=cand["age"], position=pos,
                    votes_general=cand["gen_votes"], votes_postal=cand["postal_votes"],
                    votes_total=cand["total_votes"], vote_pct=vote_pct, is_nota=is_nota,
                    education=education, occupation=occupation,
                    declared_assets=declared_assets, liabilities=liabilities_val,
                    criminal_cases=criminal_cases, image_url=image_url,
                ))
                imported += 1

        session.commit()
        election_id = election.id

    print(f"\n=== Puducherry 2011 ===")
    print(f"Constituencies: {len(constituency_cache)}, Candidates: {imported}")
    with Session(engine) as session:
        winners = session.query(Candidate).filter_by(election_id=election_id, position=1, is_nota=False).all()
        seats = {}
        for w in winners:
            p = w.party.abbr if w.party else "IND"
            seats[p] = seats.get(p, 0) + 1
        for p, s in sorted(seats.items(), key=lambda x: -x[1]):
            print(f"  {p}: {s} seats")


if __name__ == "__main__":
    run()
