"""Import Assam 2016 election data — detailed results + affidavit data (JSON/CSV).

Data sources:
  - Results: assam 2016/DETAILED RESULTS 2016.xlsx
  - Affidavit: assam 2016/2016_candidate_profile.json + assam_candidates_2016.csv
  - District mapping: Assam District and AC name (2016,2021,2026) 1.xlsx (2016 cols 0-3)
"""
import sys
import os
import re
import json
import csv
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.database import engine, Base
from app.models import Election, District, Constituency, Party, Candidate, DistrictMapping

PROJECT_ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
MAPPING_FILE = os.path.join(PROJECT_ROOT, "Assam District and AC name (2016,2021,2026) 1.xlsx")
RESULTS_FILE = os.path.join(PROJECT_ROOT, "assam 2016", "DETAILED RESULTS 2016.xlsx")
JSON_FILE = os.path.join(PROJECT_ROOT, "assam 2016", "2016_candidate_profile.json")
CSV_FILE = os.path.join(PROJECT_ROOT, "assam 2016", "assam_candidates_2016.csv")

PARTY_COLORS = {
    "BJP": "#FF9933", "INC": "#00BFFF", "AIUDF": "#006400", "AGP": "#FFFFFF",
    "BOPF": "#228B22", "CPI(M)": "#FF0000", "CPI": "#FF0000",
    "AITC": "#00FF00", "NCP": "#004080", "JD(U)": "#003366",
    "IND": "#808080", "NOTA": "#000000",
}


def load_2016_ac_mapping():
    """Parse 2016 columns (0-3) from the mapping file."""
    wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    ac_map = {}
    current_district = None
    for row in rows[2:]:
        if row[1]:
            current_district = str(row[1]).strip()
        if row[2] is not None and row[3]:
            ac_no = int(row[2])
            raw_name = str(row[3]).strip()
            category = "GEN"
            if "(SC)" in raw_name:
                category = "SC"
                raw_name = raw_name.replace("(SC)", "").strip()
            elif "(ST)" in raw_name:
                category = "ST"
                raw_name = raw_name.replace("(ST)", "").strip()
            ac_map[ac_no] = {"district": current_district, "name": raw_name, "category": category}
    return ac_map


def load_district_triples():
    """Parse same-row district triples: 2016 (col 1), 2021 (col 5), 2026 (col 9)."""
    wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    triples = []
    seen_16 = set()
    seen_21 = set()
    seen_26 = set()

    for row in rows[2:]:
        d16 = str(row[1]).strip() if row[1] else None
        d21 = str(row[5]).strip() if row[5] else None
        d26 = str(row[9]).strip() if row[9] else None

        new_entry = False
        if d16 and d16 not in seen_16:
            seen_16.add(d16)
            new_entry = True
        if d21 and d21 not in seen_21:
            seen_21.add(d21)
            new_entry = True
        if d26 and d26 not in seen_26:
            seen_26.add(d26)
            new_entry = True

        if new_entry and any([d16, d21, d26]):
            triples.append((d16 if d16 and d16 not in [t[0] for t in triples] else None,
                            d21 if d21 and d21 not in [t[1] for t in triples] else None,
                            d26 if d26 and d26 not in [t[2] for t in triples] else None))

    # Filter out fully-None triples
    return [(a, b, c) for a, b, c in triples if a or b or c]


def parse_rupees(val):
    if not val or "Nil" in str(val):
        return 0
    cleaned = re.sub(r"[Rs\s\xa0,~]", "", str(val))
    digits = re.match(r"^([\d.]+)", cleaned)
    return int(float(digits.group(1))) if digits else 0


def parse_criminal(status):
    if not status or "No criminal" in status:
        return 0
    m = re.search(r"(\d+)", status)
    return int(m.group(1)) if m else 0


def run():
    Base.metadata.create_all(bind=engine)

    ac_map = load_2016_ac_mapping()
    print(f"Loaded 2016 mapping: {len(ac_map)} ACs, {len(set(a['district'] for a in ac_map.values()))} districts")

    # Load detailed results
    wb = openpyxl.load_workbook(RESULTS_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    result_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    print(f"Loaded {len(result_rows)} rows from detailed results")

    # Load CSV for candidate names
    csv_by_id = {}
    with open(CSV_FILE, "r", encoding="latin-1") as f:
        reader = csv.DictReader(f)
        for row in reader:
            cid = row.get("candidate_id", "").strip()
            con = row.get("constituency", "").strip().upper()
            con = re.sub(r"\s*\(SC\)\s*$|\s*\(ST\)\s*$", "", con)
            csv_by_id[cid] = {
                "name": row.get("candidate_name", "").strip(),
                "constituency": con,
                "party_abbr": row.get("party", "").strip(),
            }

    # Load JSON for affidavit data, indexed by (constituency, name)
    with open(JSON_FILE, "r", encoding="utf-8") as f:
        candidates_json = json.load(f)
    json_by_key = {}
    for c in candidates_json:
        cp = c.get("candidate_profile", {})
        cid = str(cp.get("candidate_id", ""))
        csv_info = csv_by_id.get(cid)
        if csv_info:
            key = (csv_info["constituency"], csv_info["name"].upper())
            json_by_key[key] = c
    print(f"Loaded {len(json_by_key)} affidavit profiles (matched to CSV)")

    with Session(engine) as session:
        existing = session.query(Election).filter(Election.state == "Assam", Election.year == 2016).first()
        if existing:
            print(f"ERROR: Assam 2016 already exists (id={existing.id}). Skipping.")
            return

        # 1. Create election
        election = Election(
            state="Assam", year=2016, type="Assembly",
            name="Assam Legislative Assembly Election 2016",
        )
        session.add(election)
        session.flush()
        print(f"Created election: {election.name} (id={election.id})")

        # 2. Create districts
        district_cache = {}
        for ac_info in ac_map.values():
            dname = ac_info["district"]
            if dname not in district_cache:
                district = District(election_id=election.id, name=dname)
                session.add(district)
                district_cache[dname] = district
        session.flush()
        print(f"Created {len(district_cache)} districts")

        # 3. Create constituencies with vote data from results
        constituency_cache = {}  # ac_no -> Constituency
        constituency_by_name = {}  # uppercase name -> Constituency

        # Parse results to get constituency-level totals
        ac_data = {}  # ac_no -> {total_electors, total_votes, candidates: [...]}
        for row in result_rows[3:]:  # skip header rows
            ac_no = row[0]
            if ac_no is None or not isinstance(ac_no, (int, float)):
                continue
            ac_no = int(ac_no)
            ac_name = str(row[1]).strip() if row[1] else None
            cand_name = str(row[2]).strip() if row[2] else None
            sex = row[3]
            age = int(row[4]) if row[4] else None
            category = row[5]
            party = str(row[6]).strip() if row[6] else None
            gen_votes = int(row[7]) if row[7] else 0
            postal_votes = int(row[8]) if row[8] else 0
            total_valid = int(row[9]) if row[9] else 0
            total_electors = int(row[10]) if row[10] else None
            total_votes = int(row[11]) if row[11] else None

            if ac_no not in ac_data:
                ac_data[ac_no] = {
                    "name": ac_name,
                    "total_electors": total_electors,
                    "total_votes": total_votes,
                    "candidates": [],
                }

            ac_data[ac_no]["candidates"].append({
                "name": cand_name,
                "sex": sex,
                "age": age,
                "party": party,
                "gen_votes": gen_votes,
                "postal_votes": postal_votes,
                "total_votes": total_valid,
            })

        # Create constituencies
        for ac_no, info in ac_data.items():
            mapping = ac_map.get(ac_no, {})
            district_name = mapping.get("district")
            district = district_cache.get(district_name)

            # Compute turnout and margin
            total_electors = info["total_electors"]
            total_votes = info["total_votes"]
            turnout = round(total_votes / total_electors * 100, 2) if total_electors and total_votes else None

            # Sort candidates by votes to find margin
            sorted_cands = sorted(info["candidates"], key=lambda x: x["total_votes"], reverse=True)
            # Exclude NOTA from margin calc
            real_cands = [c for c in sorted_cands if c["party"] != "NOTA"]
            margin = None
            if len(real_cands) >= 2:
                margin = real_cands[0]["total_votes"] - real_cands[1]["total_votes"]

            c = Constituency(
                election_id=election.id,
                district_id=district.id if district else None,
                ac_no=ac_no,
                name=mapping.get("name", info["name"]),
                category=mapping.get("category"),
                total_electors=total_electors,
                total_votes_polled=total_votes,
                turnout_pct=turnout,
                winning_margin=margin,
            )
            session.add(c)
            constituency_cache[ac_no] = c
            constituency_by_name[mapping.get("name", info["name"]).upper()] = c

        session.flush()
        print(f"Created {len(constituency_cache)} constituencies")

        # 4. Build party cache
        party_cache = {}
        for p in session.query(Party).all():
            party_cache[p.abbr.lower()] = p
            party_cache[p.name.lower()] = p

        # 5. Import candidates from results + affidavit
        imported = 0
        for ac_no, info in ac_data.items():
            constituency = constituency_cache[ac_no]
            sorted_cands = sorted(info["candidates"], key=lambda x: x["total_votes"], reverse=True)

            for pos, cand in enumerate(sorted_cands, start=1):
                is_nota = (cand["party"] == "NOTA")
                name = cand["name"]
                if name == "None of the Above":
                    name = "NOTA"

                # Get or create party
                party = None
                if not is_nota and cand["party"]:
                    p_key = cand["party"].lower()
                    if p_key not in party_cache:
                        new_party = Party(
                            name=cand["party"], abbr=cand["party"],
                            color=PARTY_COLORS.get(cand["party"]),
                        )
                        session.add(new_party)
                        session.flush()
                        party_cache[p_key] = new_party
                    party = party_cache[p_key]

                # Total votes polled in constituency for vote_pct
                total_valid_in_ac = sum(c["total_votes"] for c in sorted_cands)
                vote_pct = round(cand["total_votes"] / total_valid_in_ac * 100, 2) if total_valid_in_ac > 0 else 0

                # Look up affidavit data
                ac_name_upper = constituency.name.upper()
                affidavit = json_by_key.get((ac_name_upper, name.upper()))

                education = None
                occupation = None
                declared_assets = None
                liabilities_val = None
                criminal_cases = 0
                image_url = None

                if affidavit:
                    cp = affidavit["candidate_profile"]
                    edu = (cp.get("education") or "").strip()
                    if edu.startswith("Category:"):
                        edu = edu[len("Category:"):].strip()
                    if edu and edu != "Not mentioned":
                        education = edu

                    occ = (affidavit.get("profession", {}).get("self") or "").strip()
                    if occ and occ != "Not mentioned":
                        occupation = occ

                    assets = parse_rupees(affidavit.get("assets_summary", {}).get("total_assets"))
                    if assets > 0:
                        declared_assets = assets
                    liab = parse_rupees(affidavit.get("assets_summary", {}).get("total_liabilities"))
                    if liab > 0:
                        liabilities_val = liab

                    criminal_cases = parse_criminal(cp.get("crime_status"))

                    img = (cp.get("image_url") or "").strip()
                    if img and img != "None":
                        image_url = img

                session.add(Candidate(
                    election_id=election.id,
                    constituency_id=constituency.id,
                    party_id=party.id if party else None,
                    name=name,
                    gender=cand["sex"],
                    age=cand["age"],
                    position=pos,
                    votes_general=cand["gen_votes"],
                    votes_postal=cand["postal_votes"],
                    votes_total=cand["total_votes"],
                    vote_pct=vote_pct,
                    is_nota=is_nota,
                    education=education,
                    occupation=occupation,
                    declared_assets=declared_assets,
                    liabilities=liabilities_val,
                    criminal_cases=criminal_cases,
                    image_url=image_url,
                ))
                imported += 1

        session.flush()
        print(f"Imported {imported} candidates")

        # 6. Rebuild district_mappings for all 3 years
        print("Rebuilding district mappings...")
        session.query(DistrictMapping).delete()

        e16 = session.query(Election).filter(Election.state == "Assam", Election.year == 2016).first()
        e21 = session.query(Election).filter(Election.state == "Assam", Election.year == 2021).first()
        e26 = session.query(Election).filter(Election.state == "Assam", Election.year == 2026).first()

        d16 = {d.name.lower(): d for d in session.query(District).filter(District.election_id == e16.id).all()} if e16 else {}
        d21 = {d.name.lower(): d for d in session.query(District).filter(District.election_id == e21.id).all()} if e21 else {}
        d26 = {d.name.lower(): d for d in session.query(District).filter(District.election_id == e26.id).all()} if e26 else {}

        triples = load_district_triples()
        mapped = 0
        for d16_name, d21_name, d26_name in triples:
            dd16 = d16.get(d16_name.lower()) if d16_name else None
            dd21 = d21.get(d21_name.lower()) if d21_name else None
            dd26 = d26.get(d26_name.lower()) if d26_name else None
            if dd16 or dd21 or dd26:
                session.add(DistrictMapping(
                    district_2016_id=dd16.id if dd16 else None,
                    district_2021_id=dd21.id if dd21 else None,
                    district_2026_id=dd26.id if dd26 else None,
                ))
                mapped += 1
        print(f"Created {mapped} district mappings")

        session.commit()

        # Summary
        print(f"\n=== All Elections ===")
        for e in session.query(Election).order_by(Election.year).all():
            n_real = session.query(Candidate).filter(Candidate.election_id == e.id, Candidate.is_nota == False).count()
            n_nota = session.query(Candidate).filter(Candidate.election_id == e.id, Candidate.is_nota == True).count()
            has_votes = session.query(Candidate).filter(Candidate.election_id == e.id, Candidate.votes_total != None).count()
            has_edu = session.query(Candidate).filter(Candidate.election_id == e.id, Candidate.education != None).count()
            has_assets = session.query(Candidate).filter(Candidate.election_id == e.id, Candidate.declared_assets != None).count()
            print(f"  {e.state} {e.year}: {n_real} candidates + {n_nota} NOTA | votes={has_votes} edu={has_edu} assets={has_assets}")


if __name__ == "__main__":
    run()
