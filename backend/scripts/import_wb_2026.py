"""Import West Bengal 2026 election data — candidates only (no results yet).

Data sources (all under west_bengal_election_data 1/west_bengal_election_data/wb_2026/):
  - Electors: AC_wise_Draft_Elector_SIR_2026.xlsx (district, AC, polling stations, electors)
  - Candidates CSV: west_bengal_candidates_2026.csv (candidate_id bridge)
  - Candidate profiles: 2026_West_Bengal_candidate_profile.json (affidavit data)
"""
import sys
import os
import re
import json
import csv

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from app.database import engine, Base
from app.models import Election, District, Constituency, Party, Candidate

PROJECT_ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
DATA_DIR = os.path.join(PROJECT_ROOT, "west_bengal_election_data 1", "west_bengal_election_data", "wb_2026")

ELECTORS_FILE = os.path.join(DATA_DIR, "AC_wise_Draft_Elector_SIR_2026.xlsx")
CANDIDATES_CSV = os.path.join(DATA_DIR, "west_bengal_candidates_2026.csv")
PROFILES_JSON = os.path.join(DATA_DIR, "2026_West_Bengal_candidate_profile.json")

PARTY_COLORS = {
    "AITC": "#00FF00", "CPM": "#FF0000", "CPI(M)": "#FF0000", "INC": "#00BFFF",
    "BJP": "#FF9933", "AIFB": "#CC0000", "RSP": "#FF6600", "CPI": "#FF4444",
    "JD(U)": "#003366", "BSP": "#0000FF", "NCP": "#004080", "IND": "#808080",
}

# Full party name -> abbreviation mapping for common WB parties
PARTY_ABBR_MAP = {
    "All India Trinamool Congress": "AITC",
    "Bharatiya Janata Party": "BJP",
    "Indian National Congress": "INC",
    "Communist Party of India (Marxist)": "CPI(M)",
    "Communist Party of India": "CPI",
    "All India Forward Bloc": "AIFB",
    "Revolutionary Socialist Party": "RSP",
    "Bahujan Samaj Party": "BSP",
    "Nationalist Congress Party - Sharadchandra Pawar": "NCP-SP",
    "Independent": "IND",
    "Janata Dal (United)": "JD(U)",
}


def load_electors():
    """Parse electors file -> {ac_no: {district, name, polling_stations, electors}}."""
    wb = openpyxl.load_workbook(ELECTORS_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    ac_map = {}
    for row in rows[1:]:
        district = str(row[0]).strip() if row[0] else None
        if not district or district in ("Total", "Grand Total", "State Total"):
            continue
        ac_no = int(row[1]) if row[1] else None
        raw_name = str(row[2]).strip() if row[2] else None
        if not ac_no or not raw_name:
            continue

        category = "GEN"
        if "(SC)" in raw_name.upper():
            category = "SC"
        elif "(ST)" in raw_name.upper():
            category = "ST"
        clean_name = re.sub(r"\s*\((SC|ST)\)\s*$", "", raw_name, flags=re.IGNORECASE).strip()

        ac_map[ac_no] = {
            "district": district.upper(),
            "name": clean_name.upper(),
            "category": category,
            "total_polling_stations": int(row[3]) if row[3] else None,
            "male_electors": int(row[4]) if row[4] else None,
            "female_electors": int(row[5]) if row[5] else None,
            "third_gender_electors": int(row[6]) if row[6] else None,
            "total_electors": int(row[7]) if row[7] else None,
        }
    return ac_map


def load_candidates():
    """Load candidate data from JSON + CSV."""
    # Load CSV for candidate_id -> name mapping
    csv_by_id = {}
    with open(CANDIDATES_CSV, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            cid = row.get("candidate_id", "").strip()
            csv_by_id[cid] = {
                "name": row.get("candidate_name", "").strip(),
                "constituency": re.sub(r"\s*\((SC|ST)\)\s*$", "", row.get("constituency", "").strip().upper()),
                "party_name": row.get("party", "").strip(),
            }

    # Load JSON profiles
    with open(PROFILES_JSON, "r", encoding="utf-8") as f:
        profiles = json.load(f)

    # Group by constituency
    candidates_by_ac = {}
    for p in profiles:
        cp = p.get("candidate_profile", {})
        cid = str(cp.get("candidate_id", ""))
        ac_name = re.sub(r"\s+", " ", (cp.get("ac_name") or "").strip().upper())

        csv_info = csv_by_id.get(cid)
        name = csv_info["name"] if csv_info else (cp.get("name") or "").strip()
        party_name = csv_info["party_name"] if csv_info else (cp.get("party") or "").strip()

        if not ac_name:
            continue

        if ac_name not in candidates_by_ac:
            candidates_by_ac[ac_name] = []

        candidates_by_ac[ac_name].append({
            "name": name,
            "party_name": party_name,
            "age": int(cp.get("age", "0").strip()) if cp.get("age", "").strip().isdigit() else None,
            "education": (cp.get("education") or "").strip(),
            "crime_status": (cp.get("crime_status") or "").strip(),
            "image_url": (cp.get("image_url") or "").strip(),
            "profession": (p.get("profession", {}).get("self") or "").strip(),
            "total_assets": p.get("assets_summary", {}).get("total_assets"),
            "total_liabilities": p.get("assets_summary", {}).get("total_liabilities"),
        })

    return candidates_by_ac


def parse_rupees(val):
    if not val or "Nil" in str(val):
        return 0
    cleaned = re.sub(r"[Rs\s\xa0,~]", "", str(val))
    digits = re.match(r"^([\d.]+)", cleaned)
    return int(float(digits.group(1))) if digits else 0


def parse_criminal(status):
    if not status or "No criminal" in status:
        return 0
    m = re.search(r"(\d+)", status)
    return int(m.group(1)) if m else 0


def run():
    Base.metadata.create_all(bind=engine)

    ac_map = load_electors()
    print(f"Loaded electors: {len(ac_map)} ACs")

    candidates_by_ac = load_candidates()
    print(f"Loaded candidates for {len(candidates_by_ac)} ACs")

    with Session(engine) as session:
        existing = session.query(Election).filter(
            Election.state == "West Bengal", Election.year == 2026
        ).first()
        if existing:
            print(f"ERROR: West Bengal 2026 already exists (id={existing.id}). Skipping.")
            return

        election = Election(
            state="West Bengal", year=2026, type="Assembly",
            name="West Bengal Legislative Assembly Election 2026",
        )
        session.add(election)
        session.flush()
        print(f"Created election: {election.name} (id={election.id})")

        # Create districts
        district_cache = {}
        for ac_info in ac_map.values():
            dname = ac_info["district"]
            if dname not in district_cache:
                district = District(election_id=election.id, name=dname.title())
                session.add(district)
                district_cache[dname] = district
        session.flush()
        print(f"Created {len(district_cache)} districts")

        # Create constituencies (no vote data)
        constituency_cache = {}
        for ac_no, info in ac_map.items():
            district = district_cache.get(info["district"])
            c = Constituency(
                election_id=election.id,
                district_id=district.id if district else None,
                ac_no=ac_no,
                name=info["name"],
                category=info["category"],
                total_electors=info["total_electors"],
                male_electors=info["male_electors"],
                female_electors=info["female_electors"],
                third_gender_electors=info["third_gender_electors"],
                total_polling_stations=info["total_polling_stations"],
            )
            session.add(c)
            constituency_cache[ac_no] = c
            constituency_cache[info["name"]] = c
        session.flush()
        print(f"Created {len(ac_map)} constituencies")

        # Build party cache
        party_cache = {}
        for p in session.query(Party).all():
            party_cache[p.abbr.upper()] = p
            party_cache[p.name.upper()] = p

        # Import candidates
        imported = 0
        for ac_name, cands in candidates_by_ac.items():
            constituency = constituency_cache.get(ac_name)
            if not constituency:
                # Try fuzzy match
                for key in constituency_cache:
                    if isinstance(key, str) and (ac_name in key or key in ac_name):
                        constituency = constituency_cache[key]
                        break
            if not constituency:
                continue

            for cand in cands:
                # Resolve party
                party_name = cand["party_name"]
                abbr = PARTY_ABBR_MAP.get(party_name)
                if not abbr:
                    abbr = party_name[:10] if len(party_name) > 10 else party_name

                p_key = abbr.upper()
                if p_key not in party_cache:
                    p_key = party_name.upper()
                if p_key not in party_cache:
                    new_party = Party(
                        name=party_name, abbr=abbr,
                        color=PARTY_COLORS.get(abbr),
                    )
                    session.add(new_party)
                    session.flush()
                    party_cache[abbr.upper()] = new_party
                    party_cache[party_name.upper()] = new_party
                    p_key = abbr.upper()
                party = party_cache[p_key]

                edu = cand["education"]
                if edu.startswith("Category:"):
                    edu = edu[len("Category:"):].strip()
                if not edu or edu in ("Not mentioned", "Not Given"):
                    edu = None

                occ = cand["profession"]
                if not occ or occ in ("Not mentioned", "Not Given"):
                    occ = None

                assets = parse_rupees(cand["total_assets"])
                liab = parse_rupees(cand["total_liabilities"])
                criminal = parse_criminal(cand["crime_status"])
                img = cand["image_url"]
                if not img or img == "None":
                    img = None

                session.add(Candidate(
                    election_id=election.id,
                    constituency_id=constituency.id,
                    party_id=party.id,
                    name=cand["name"],
                    age=cand["age"],
                    is_nota=False,
                    education=edu,
                    occupation=occ,
                    declared_assets=assets if assets > 0 else None,
                    liabilities=liab if liab > 0 else None,
                    criminal_cases=criminal,
                    image_url=img,
                ))
                imported += 1

        session.flush()
        print(f"Imported {imported} candidates")
        session.commit()

        # Summary
        print(f"\n=== Summary ===")
        total = session.query(Candidate).filter(Candidate.election_id == election.id).count()
        has_edu = session.query(Candidate).filter(Candidate.election_id == election.id, Candidate.education != None).count()
        has_assets = session.query(Candidate).filter(Candidate.election_id == election.id, Candidate.declared_assets != None).count()
        print(f"  Total candidates: {total}")
        print(f"  With education: {has_edu}, With assets: {has_assets}")

        # Party breakdown
        print(f"\n=== Top Parties by Candidates ===")
        from sqlalchemy import func as sqlfunc
        party_counts = session.query(
            Party.abbr, sqlfunc.count(Candidate.id)
        ).join(Candidate).filter(
            Candidate.election_id == election.id
        ).group_by(Party.abbr).order_by(sqlfunc.count(Candidate.id).desc()).limit(10).all()
        for abbr, count in party_counts:
            print(f"  {abbr}: {count} candidates")


if __name__ == "__main__":
    run()
