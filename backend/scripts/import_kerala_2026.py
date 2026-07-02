"""Import Kerala 2026 election data — all candidates from myneta CSV/JSON.

Sources (Kerala/Kerala_2026/):
  - All candidates:   Kerala_candidates_2026.csv  +  2026_candidate_profile.json
  - Constituency map: Kerala_Election_Data_Mapping.xlsx (AC sheet, 2026 cols)
  - Electors:        AC WISE ELECTORS.xlsx
  - Voter turnout:   AC WISE VOTERS INFORMATION.xlsx
  - Winner lookup:   LIST OF SUCCESSFUL CANDIDATES.xlsx  (to set position=1)
  - Party list:      LIST OF PARTICIPATING POLITICAL PARTIES.xlsx
"""
import sys
import os
import re
import csv
import json

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from app.database import engine, Base
from app.models.kerala import Election, District, Constituency, Party, Candidate

PROJECT_ROOT  = os.path.join(os.path.dirname(__file__), "..", "..")
DATA_DIR      = os.path.join(PROJECT_ROOT, "Kerala", "Kerala_2026")
CANDIDATES_CSV  = os.path.join(DATA_DIR, "Kerala_candidates_2026.csv")
PROFILES_JSON   = os.path.join(DATA_DIR, "2026_candidate_profile.json")
ELECTORS_FILE   = os.path.join(DATA_DIR, "AC WISE ELECTORS.xlsx")
VOTERS_FILE     = os.path.join(DATA_DIR, "AC WISE VOTERS INFORMATION.xlsx")
WINNERS_FILE    = os.path.join(DATA_DIR, "LIST OF SUCCESSFUL CANDIDATES.xlsx")
PARTY_LIST      = os.path.join(DATA_DIR, "LIST OF PARTICIPATING POLITICAL PARTIES.xlsx")
MAPPING_FILE    = os.path.join(PROJECT_ROOT, "Kerala", "Kerala_Election_Data_Mapping.xlsx")
GEOJSON_PATH    = os.path.join(PROJECT_ROOT, "frontend", "public", "kerala_AC.geojson")

PARTY_COLORS = {
    "BJP": "#FF9933", "INC": "#00BFFF", "CPI(M)": "#FF0000", "CPI": "#FF6600",
    "IUML": "#006400", "JD(S)": "#FFD700", "NCP": "#004080", "BSP": "#0000FF",
    "IND": "#808080", "NOTA": "#000000", "AAP": "#0066CC", "KC(M)": "#800080",
    "KEC(M)": "#8B008B", "RSP": "#FF4444", "AIFB": "#CC0000", "AAAP": "#0066CC",
    "INL": "#4B0082",
}

# Spelling differences between myneta CSV constituency names and 2026 mapping file
CONSTITUENCY_ALIASES = {
    "THIRUVAMBADY":   "THIRUVAMBADI",
    "VALLIKKUNNU":    "VALLIKUNNU",
    "OTTAPALAM":      "OTTAPPALAM",
    "MANNARKAD":      "MANNARKKAD",
    "IRINJALAKKUDA":  "IRINJALAKUDA",
    "AMBALAPUZHA":    "AMBALAPPUZHA",
    "MAVELIKARA":     "MAVELIKKARA",
    "CHATHANNUR":     "CHATHANNOOR",
    "KAZHAKKOOTTAM":  "KAZHAKOOTTAM",
    "NENMARA":        "NEMMARA",
    "ERANAD":         "ERANAD",        # fallback: use 2021 name if 2026 missing
}


def _strip_cat(name: str) -> str:
    return re.sub(r"\s*\((SC|ST)\)\s*$", "", name.strip(), flags=re.I).strip().upper()


def _norm_name(name: str) -> str:
    """Normalize a candidate name for fuzzy matching.
    Strips titles (ADV., PROF, DR.), S/O suffixes, punctuation, extra spaces."""
    n = name.upper().strip()
    n = re.sub(r"\bADV\.?\b|\bPROF\.?\b|\bDR\.?\b|\bMR\.?\b|\bSMT\.?\b", "", n)
    n = re.sub(r"\bS/O\b.*", "", n)
    n = re.sub(r"[^A-Z\s]", "", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n


def load_ac_name_to_no():
    """Build constituency name → AC number from 2026 mapping col (cols 9,10)."""
    wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True)
    ws = wb["AC"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    name_to_no = {}
    for row in rows[2:]:
        ac_no_val, ac_name_val = row[9], row[10]
        if not ac_no_val or not ac_name_val:
            continue
        ac_no = int(ac_no_val)
        raw = str(ac_name_val).strip().upper()
        clean = _strip_cat(raw)
        name_to_no[raw] = ac_no
        name_to_no[clean] = ac_no
    # Also index 2021 names (cols 6,7) as fallback for renamed constituencies
    for row in rows[2:]:
        ac_no_val, ac_name_val = row[6], row[7]
        if not ac_no_val or not ac_name_val:
            continue
        ac_no = int(ac_no_val)
        raw = str(ac_name_val).strip().upper()
        clean = _strip_cat(raw)
        if raw not in name_to_no:
            name_to_no[raw] = ac_no
        if clean not in name_to_no:
            name_to_no[clean] = ac_no
    return name_to_no


def resolve_ac_no(constituency_raw: str, name_to_no: dict) -> int | None:
    """Try multiple forms to match constituency name → AC number."""
    raw = constituency_raw.strip().upper()
    clean = _strip_cat(raw)
    alias_raw = CONSTITUENCY_ALIASES.get(raw, CONSTITUENCY_ALIASES.get(clean))
    alias_clean = _strip_cat(alias_raw) if alias_raw else None
    for key in [raw, clean, alias_raw, alias_clean]:
        if key and key in name_to_no:
            return name_to_no[key]
    return None


def load_ac_district_map():
    """GeoJSON → ac_no: district name."""
    with open(GEOJSON_PATH, encoding="utf-8") as f:
        data = json.load(f)
    ac_dist = {}
    for feat in data.get("features", []):
        props = feat.get("properties", {})
        try:
            ac_no = int(props.get("AC_NO", 0))
            ac_dist[ac_no] = props.get("DIST_NAME", "Unknown").title()
        except (ValueError, TypeError):
            pass
    return ac_dist


def load_ac_category_map():
    """2026 mapping → ac_no: category."""
    wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True)
    ws = wb["AC"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    cat_map = {}
    for row in rows[2:]:
        if not row[9] or not row[10]:
            continue
        ac_no = int(row[9])
        name = str(row[10]).strip()
        cat = "GEN"
        if "(SC)" in name:
            cat = "SC"
        elif "(ST)" in name:
            cat = "ST"
        cat_map[ac_no] = cat
    return cat_map


def load_ac_name_map():
    """2026 mapping → ac_no: clean AC name (without category suffix)."""
    wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True)
    ws = wb["AC"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    name_map = {}
    for row in rows[2:]:
        if not row[9] or not row[10]:
            continue
        ac_no = int(row[9])
        name_map[ac_no] = _strip_cat(str(row[10]).strip())
    return name_map


def load_electors_map():
    """AC WISE ELECTORS → ac_no: {male, female, third_gender, total}."""
    wb = openpyxl.load_workbook(ELECTORS_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    out = {}
    for row in rows[1:]:
        if not row[0]:
            continue
        try:
            ac_no = int(row[0])
        except (ValueError, TypeError):
            continue
        out[ac_no] = {
            "male":         int(row[9])  if row[9]  else 0,
            "female":       int(row[10]) if row[10] else 0,
            "third_gender": int(row[11]) if row[11] else 0,
            "total":        int(row[12]) if row[12] else 0,
        }
    return out


def load_voters_map():
    """AC WISE VOTERS INFO → ac_no: {total_votes, turnout_pct}."""
    wb = openpyxl.load_workbook(VOTERS_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    out = {}
    for row in rows[1:]:
        if not row[0]:
            continue
        try:
            ac_no = int(row[0])
        except (ValueError, TypeError):
            continue
        out[ac_no] = {
            "total_votes": int(row[13])   if row[13] else 0,
            "turnout_pct": float(row[15]) if row[15] else None,
        }
    return out


def load_party_names():
    wb = openpyxl.load_workbook(PARTY_LIST, read_only=True)
    ws = wb[wb.sheetnames[0]]
    mapping = {}
    for row in ws.iter_rows(values_only=True):
        if isinstance(row[1], (int, float)) and row[2] and row[3]:
            mapping[str(row[2]).strip()] = str(row[3]).strip()
    wb.close()
    return mapping


def load_winners():
    """LIST OF SUCCESSFUL CANDIDATES → set of (ac_no, winner_name_upper)."""
    wb = openpyxl.load_workbook(WINNERS_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    winners = {}  # ac_no → winner name upper
    for row in rows[1:]:
        if not row[0]:
            continue
        try:
            ac_no = int(row[0])
        except (ValueError, TypeError):
            continue
        winner_name = str(row[2]).strip().upper() if row[2] else ""
        winners[ac_no] = winner_name
    return winners


def _parse_rupees(val):
    if not val or "Nil" in str(val):
        return 0
    s = str(val).strip()
    if s.startswith("http"):
        return 0
    digits = re.sub(r"[^\d]", "", s.split("~")[0])
    return int(digits) if digits else 0


def _parse_criminal(status):
    if not status or "No criminal" in str(status):
        return 0
    m = re.search(r"(\d+)", str(status))
    return int(m.group(1)) if m else 0


def load_myneta_candidates():
    """Merge CSV and JSON → list of candidate dicts keyed by (constituency_upper, name_upper)."""
    # CSV: sno, candidate_id, candidate_name, candidate_link, constituency, party,
    #      criminal_cases, education, assets, liabilities
    csv_rows = {}
    with open(CANDIDATES_CSV, encoding="latin-1") as f:
        for row in csv.DictReader(f):
            cid = row.get("candidate_id", "").strip()
            if cid:
                csv_rows[cid] = row

    with open(PROFILES_JSON, encoding="utf-8") as f:
        profiles = json.load(f)

    candidates = []
    for p in profiles:
        cp = p.get("candidate_profile", {})
        cid = str(cp.get("candidate_id", ""))
        csv_r = csv_rows.get(cid, {})

        constituency_raw = (csv_r.get("constituency") or cp.get("ac_name") or "").strip()
        name = (csv_r.get("candidate_name") or cp.get("name") or "").strip()
        party_abbr = (csv_r.get("party") or cp.get("party") or "").strip()

        edu = (csv_r.get("education") or cp.get("education") or "").strip()
        if edu.startswith("Category:"):
            edu = edu[len("Category:"):].strip()

        assets_s = p.get("assets_summary", {})
        assets_raw = csv_r.get("assets", "")
        assets_val = _parse_rupees(assets_raw) if assets_raw and not str(assets_raw).startswith("http") else _parse_rupees(assets_s.get("total_assets"))

        liab_raw = csv_r.get("liabilities", "")
        liab_val = _parse_rupees(liab_raw) if liab_raw and not str(liab_raw).startswith("http") else _parse_rupees(assets_s.get("total_liabilities"))

        crime_val = int(csv_r.get("criminal_cases") or 0) or _parse_criminal(cp.get("crime_status"))

        occ = (p.get("profession", {}).get("self") or "").strip()
        img = (cp.get("image_url") or "").strip()

        age_raw = cp.get("age", "")
        try:
            age = int(str(age_raw).strip())
        except (ValueError, TypeError):
            age = None

        candidates.append({
            "constituency": constituency_raw,
            "name": name,
            "party_abbr": party_abbr,
            "education": edu if edu and edu not in ("Not mentioned", "Not Given", "") else None,
            "occupation": occ if occ and occ != "Not mentioned" else None,
            "declared_assets": assets_val if assets_val > 0 else None,
            "liabilities": liab_val if liab_val > 0 else None,
            "criminal_cases": crime_val,
            "image_url": img if img and img != "None" else None,
            "age": age,
            "gender": cp.get("gender", "").strip().upper() or None,
        })
    return candidates


def delete_existing():
    with Session(engine) as session:
        el = session.query(Election).filter_by(state="Kerala", year=2026).first()
        if not el:
            return
        eid = el.id
        session.query(Candidate).filter_by(election_id=eid).delete()
        session.query(Constituency).filter_by(election_id=eid).delete()
        session.query(District).filter_by(election_id=eid).delete()
        session.query(Election).filter_by(id=eid).delete()
        session.commit()
        print(f"Deleted existing Kerala 2026 (id={eid})")


def import_data():
    Base.metadata.create_all(bind=engine)
    delete_existing()

    name_to_no  = load_ac_name_to_no()
    cat_map     = load_ac_category_map()
    ac_name_map = load_ac_name_map()
    ac_dist_map = load_ac_district_map()
    electors    = load_electors_map()
    voters      = load_voters_map()
    party_names = load_party_names()
    winners     = load_winners()  # ac_no → winner name upper
    myneta_cands = load_myneta_candidates()

    with Session(engine) as session:
        election = Election(
            state="Kerala",
            year=2026,
            type="Assembly",
            name="Kerala Legislative Assembly Election 2026",
        )
        session.add(election)
        session.flush()
        print(f"Created election: {election.name} (id={election.id})")

        party_cache = {p.abbr.upper(): p for p in session.query(Party).all()}
        for abbr, full_name in party_names.items():
            abbr_stored = abbr[:50]
            key = abbr_stored.upper()
            if key not in party_cache:
                p = Party(name=full_name, abbr=abbr_stored, color=PARTY_COLORS.get(abbr_stored))
                session.add(p)
                session.flush()
                party_cache[key] = p

        district_cache = {}
        constituency_cache = {}  # ac_no → Constituency obj

        # Create all 140 constituencies
        for ac_no in sorted(ac_name_map.keys()):
            district_name = ac_dist_map.get(ac_no, "Unknown")
            if district_name not in district_cache:
                d = District(election_id=election.id, name=district_name)
                session.add(d)
                session.flush()
                district_cache[district_name] = d

            el_info = electors.get(ac_no, {})
            v_info  = voters.get(ac_no, {})

            con = Constituency(
                election_id=election.id,
                district_id=district_cache[district_name].id,
                ac_no=ac_no,
                name=ac_name_map[ac_no],
                category=cat_map.get(ac_no, "GEN"),
                total_electors=el_info.get("total"),
                male_electors=el_info.get("male"),
                female_electors=el_info.get("female"),
                third_gender_electors=el_info.get("third_gender"),
                total_votes_polled=v_info.get("total_votes"),
                turnout_pct=v_info.get("turnout_pct"),
            )
            session.add(con)
            session.flush()
            constituency_cache[ac_no] = con

        print(f"Created {len(constituency_cache)} constituencies, {len(district_cache)} districts")

        # Group myneta candidates by constituency → sort winners first
        ac_cands: dict[int, list] = {}
        skipped = 0
        for cand in myneta_cands:
            ac_no = resolve_ac_no(cand["constituency"], name_to_no)
            if ac_no is None:
                skipped += 1
                continue
            ac_cands.setdefault(ac_no, []).append(cand)

        if skipped:
            print(f"Warning: {skipped} candidates could not be matched to an AC")

        # Insert candidates — winner at position 1, rest at position 2
        total_inserted = 0
        for ac_no, cands in ac_cands.items():
            con = constituency_cache.get(ac_no)
            if not con:
                continue

            winner_name = winners.get(ac_no, "")
            winner_norm = _norm_name(winner_name)

            def is_winner(c, _wn=winner_name, _wno=winner_norm):
                n = c["name"].upper()
                return n == _wn or (_wno and _norm_name(n) == _wno)

            sorted_cands = sorted(cands, key=lambda c: (0 if is_winner(c) else 1))

            for pos_idx, cand in enumerate(sorted_cands):
                position = 1 if is_winner(cand) else 2

                party_abbr = cand["party_abbr"]
                party = None
                if party_abbr:
                    abbr_stored = party_abbr[:50]
                    key = abbr_stored.upper()
                    if key not in party_cache:
                        p = Party(name=party_abbr, abbr=abbr_stored, color=PARTY_COLORS.get(abbr_stored))
                        session.add(p)
                        session.flush()
                        party_cache[key] = p
                    party = party_cache[key]

                gender_raw = (cand.get("gender") or "").upper()
                gender = "MALE" if gender_raw in ("MALE", "M") else ("FEMALE" if gender_raw in ("FEMALE", "F") else (gender_raw or None))

                session.add(Candidate(
                    election_id=election.id,
                    constituency_id=con.id,
                    party_id=party.id if party else None,
                    name=cand["name"],
                    gender=gender,
                    age=cand.get("age"),
                    position=position,
                    votes_general=None,
                    votes_postal=None,
                    votes_total=None,
                    vote_pct=None,
                    is_nota=False,
                    education=cand.get("education"),
                    occupation=cand.get("occupation"),
                    declared_assets=cand.get("declared_assets"),
                    liabilities=cand.get("liabilities"),
                    criminal_cases=cand.get("criminal_cases", 0),
                    image_url=cand.get("image_url"),
                ))
                total_inserted += 1

        session.commit()
        print(f"\n=== Import Summary ===")
        print(f"Constituencies: {len(constituency_cache)}, Candidates inserted: {total_inserted}")
        winners_in_db = session.query(Candidate).filter_by(election_id=election.id, position=1).count()
        print(f"Winners (position=1): {winners_in_db}")


if __name__ == "__main__":
    import_data()
