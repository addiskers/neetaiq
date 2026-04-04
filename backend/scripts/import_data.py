"""
Master data import script for Assam 2026 election data.
Imports from 4 Excel files into PostgreSQL.
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from app.database import engine, Base, SessionLocal
from app.models import (
    Election, District, Constituency, PollingStation, Party, Candidate
)

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "data")

# Party name -> (short_name, hex_color)
PARTY_MAP = {
    "Bharatiya Janata Party": ("BJP", "#FF6B35"),
    "Indian National Congress": ("INC", "#00BFFF"),
    "All India Trinamool Congress": ("TMC", "#00A651"),
    "All India United Democratic Front": ("AIUDF", "#2E8B57"),
    "Asom Gana Parishad": ("AGP", "#FFD700"),
    "Bodoland Peoples Front": ("BPF", "#8B4513"),
    "Aam Aadmi Party": ("AAP", "#0066CC"),
    "Communist Party of India": ("CPI", "#FF0000"),
    "Communist Party of India (Marxist)": ("CPM", "#CC0000"),
    "Communist Party of India (Marxist-Leninist) (Liberation)": ("CPIML", "#B22222"),
    "Independent": ("IND", "#808080"),
    "Raijor Dal": ("RD", "#9932CC"),
    "Assam Jatiya Parishad": ("AJP", "#228B22"),
    "All India Forward Bloc": ("AIFB", "#DC143C"),
    "Nationalist Congress Party - Sharadchandra Pawar": ("NCP-SP", "#0000CD"),
    "Gana Suraksha Party": ("GSP", "#4682B4"),
    "Apni Janta Party": ("APJP", "#708090"),
    "Apni Jantantrik Party": ("AJTP", "#696969"),
    "Autonomous State Demand Committee": ("ASDC", "#556B2F"),
    "Bharatiya Gana Parishad": ("BGP", "#CD853F"),
    "Bhartiya Jan Samaj Party": ("BJSP", "#A0522D"),
    "Gondvana Gantantra Party": ("GGP", "#8FBC8F"),
    "Janata Dal (United)": ("JDU", "#006400"),
    "Jharkhand Mukti Morcha": ("JMM", "#B8860B"),
    "Peoples Party of Arunachal": ("PPA", "#DEB887"),
    "Rashtriya Janata Dal": ("RJD", "#32CD32"),
    "Social Democratic Party Of India": ("SDPI", "#191970"),
    "Socialist Unity Centre Of India (COMMUNIST)": ("SUCI", "#800000"),
    "Tipra Motha Party": ("TMP", "#FF69B4"),
    "United Peoples Party Liberal": ("UPPL", "#00CED1"),
    "Voice of the People Party": ("VOPP", "#9370DB"),
    "Zoram People's Movement": ("ZPM", "#FF8C00"),
    "National People's Party": ("NPP", "#FFA500"),
    "Rashtriya Ulama Council": ("RUC", "#4B0082"),
    "Republican Party of India (A)": ("RPIA", "#C71585"),
    "Republican Party of India (Athawale)": ("RPIA", "#C71585"),
    "Revolutionary Communist Party of India (Rasik Bhatt)": ("RCPI", "#8B0000"),
    "The National Road Map Party of India": ("NRMPI", "#5F9EA0"),
    "United People's Party, Liberal": ("UPPL2", "#00CED1"),
    "Vikas India Party": ("VIP", "#20B2AA"),
    "Voters Party International": ("VPI", "#778899"),
}


def import_all():
    # Create tables
    print("Creating tables...")
    Base.metadata.create_all(bind=engine)

    db = SessionLocal()
    try:
        # Check if data already exists
        existing = db.query(Election).first()
        if existing:
            print("Data already imported. Drop tables first to re-import.")
            print("Use: Base.metadata.drop_all(bind=engine)")
            return

        # Step 1: Create election
        print("\n[1/6] Creating election record...")
        election = Election(
            year=2026,
            election_name="Assembly GEN-BYE-Election-MAR-MAY-2026",
            election_type="AC - GENERAL",
            state="Assam",
        )
        db.add(election)
        db.flush()
        print(f"  Election ID: {election.id}")

        # Step 2: Import districts from mapping file
        print("\n[2/6] Importing districts...")
        district_map = _import_districts(db, election.id)
        print(f"  Imported {len(district_map)} districts")

        # Step 3: Import constituencies from electors file
        print("\n[3/6] Importing constituencies...")
        constituency_map = _import_constituencies(db, election.id, district_map)
        print(f"  Imported {len(constituency_map)} constituencies")

        # Step 4: Seed parties
        print("\n[4/6] Seeding parties...")
        party_map = _seed_parties(db)
        print(f"  Seeded {len(party_map)} parties")

        # Step 5: Import polling stations
        print("\n[5/6] Importing polling stations...")
        ps_count = _import_polling_stations(db, constituency_map)
        print(f"  Imported {ps_count} polling stations")

        # Step 6: Import candidates
        print("\n[6/6] Importing candidates...")
        cand_count = _import_candidates(db, election.id, constituency_map, party_map)
        print(f"  Imported {cand_count} candidates")

        db.commit()
        print("\n=== Import complete! ===")
        _print_summary(db, election.id)

    except Exception as e:
        db.rollback()
        print(f"\nERROR: {e}")
        raise
    finally:
        db.close()


def _import_districts(db: Session, election_id: int) -> dict:
    """Import districts from 'District and AC name of Assam.xlsx'. Returns {name: district_id}."""
    path = os.path.join(DATA_DIR, "District and AC name of Assam.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Aliases: electors file uses old names, mapping file has 2026 names
    DISTRICT_ALIASES = {
        "kamrup": "kamrup rural",
        "kamrup metro": "kamrup metropolitan",
        "sibsagar": "sivasagar",
        "south salmara": "south salmara-mankachar",
    }

    district_map = {}
    for row in rows[1:]:  # skip header
        name_2021 = str(row[0]).strip() if row[0] else None
        name_2026 = str(row[1]).strip() if row[1] else None
        status = str(row[2]).strip() if row[2] else None

        if not name_2026:
            continue

        district = District(
            election_id=election_id,
            name=name_2026,
            name_previous=name_2021 if name_2021 != name_2026 else None,
            rename_status=status,
        )
        db.add(district)
        db.flush()
        district_map[name_2026.lower()] = district.id

    # Register aliases so electors file old names resolve correctly
    for old, new in DISTRICT_ALIASES.items():
        if new in district_map:
            district_map[old] = district_map[new]

    wb.close()
    return district_map


def _import_constituencies(db: Session, election_id: int, district_map: dict) -> dict:
    """Import constituencies from electors Excel. Returns {ac_number: constituency_id}."""
    path = os.path.join(DATA_DIR, "Assam_Final_Electors_10-02-2026.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    constituency_map = {}
    for row in rows[4:]:  # data starts at row 5 (index 4)
        sl_no = row[0]
        if not sl_no or not isinstance(sl_no, int):
            continue  # skip district total rows

        district_name = str(row[1]).strip()
        ac_number = int(row[2])
        ac_name = str(row[3]).strip()
        ps_count = int(row[4]) if isinstance(row[4], int) else None
        men = int(row[5]) if isinstance(row[5], int) else None
        women = int(row[6]) if isinstance(row[6], int) else None
        third_gender = int(row[7]) if isinstance(row[7], int) else None
        total = int(row[8]) if isinstance(row[8], int) else None

        # Match district
        district_id = district_map.get(district_name.lower())
        if not district_id:
            print(f"  WARNING: District '{district_name}' not found in mapping, skipping AC {ac_number}")
            continue

        constituency = Constituency(
            election_id=election_id,
            district_id=district_id,
            ac_number=ac_number,
            name=ac_name,
            total_polling_stations=ps_count,
            male_electors=men,
            female_electors=women,
            third_gender_electors=third_gender,
            total_electors=total,
        )
        db.add(constituency)
        db.flush()
        constituency_map[ac_number] = constituency.id

    wb.close()
    return constituency_map


def _seed_parties(db: Session) -> dict:
    """Seed all parties. Also scans affidavit file for any missing parties. Returns {name: party_id}."""
    # First seed known parties
    party_id_map = {}
    for name, (short, color) in PARTY_MAP.items():
        party = Party(name=name, short_name=short, color=color)
        db.add(party)
        db.flush()
        party_id_map[name] = party.id

    # Scan affidavit file for any parties not in PARTY_MAP
    path = os.path.join(DATA_DIR, "affidavit_eci_nomination_final_2026 1.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    for row in rows[1:]:
        party_name = str(row[5]).strip() if row[5] else None
        if party_name and party_name not in party_id_map:
            party = Party(name=party_name, short_name="OTH", color="#A9A9A9")
            db.add(party)
            db.flush()
            party_id_map[party_name] = party.id
            print(f"  Added missing party: {party_name}")

    wb.close()
    return party_id_map


def _import_polling_stations(db: Session, constituency_map: dict) -> int:
    """Import polling stations from PS_LIST Excel."""
    path = os.path.join(DATA_DIR, "PS_LIST_Filtered 1.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    count = 0
    batch = []
    for row in rows[1:]:  # skip header
        ac_number = int(row[1]) if row[1] else None
        part_no = int(row[3]) if row[3] else None
        part_name = str(row[4]).strip() if row[4] else ""

        if not ac_number or not part_no:
            continue

        constituency_id = constituency_map.get(ac_number)
        if not constituency_id:
            continue

        batch.append(PollingStation(
            constituency_id=constituency_id,
            part_no=part_no,
            name=part_name,
        ))
        count += 1

        # Batch insert every 5000 rows
        if len(batch) >= 5000:
            db.bulk_save_objects(batch)
            db.flush()
            batch = []
            print(f"  ... {count} polling stations")

    if batch:
        db.bulk_save_objects(batch)
        db.flush()

    wb.close()
    return count


def _import_candidates(db: Session, election_id: int, constituency_map: dict, party_map: dict) -> int:
    """Import candidates from affidavit Excel."""
    path = os.path.join(DATA_DIR, "affidavit_eci_nomination_final_2026 1.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Build constituency name -> id lookup (need to query DB for names)
    from app.models import Constituency
    import re
    db_constituencies = db.query(Constituency).filter(Constituency.election_id == election_id).all()
    name_to_id = {}
    for c in db_constituencies:
        full_name = c.name.upper().strip()
        name_to_id[full_name] = c.id
        # Also register the name without (SC)/(ST) suffix for matching
        base_name = re.sub(r'\s*\((?:SC|ST)\)\s*$', '', full_name)
        if base_name != full_name:
            name_to_id[base_name] = c.id

    count = 0
    for row in rows[1:]:
        candidate_name = str(row[4]).strip() if row[4] else None
        party_name = str(row[5]).strip() if row[5] else None
        status = str(row[6]).strip() if row[6] else None
        contesting = str(row[7]).strip() if row[7] else "No"
        constituency_name = str(row[9]).strip().upper() if row[9] else None
        profile_url = str(row[10]).strip() if row[10] else None

        if not candidate_name or not constituency_name:
            continue

        # Try exact match, then without suffix, then fuzzy
        constituency_id = name_to_id.get(constituency_name)
        if not constituency_id:
            # Strip (SC)/(ST) suffix and normalize spacing
            normalized = re.sub(r'\s*\((?:SC|ST)\)\s*$', '', constituency_name)
            constituency_id = name_to_id.get(normalized)
        if not constituency_id:
            # Handle known spelling/spacing variants
            CONSTITUENCY_ALIASES = {
                "BHOWANIPUR-SORBHOG": "BHAWANIPUR-SORBHOG",
                "DOOMDOOMA": "DOOM DOOMA",
            }
            alias = CONSTITUENCY_ALIASES.get(constituency_name) or CONSTITUENCY_ALIASES.get(normalized)
            if alias:
                constituency_id = name_to_id.get(alias)
        if not constituency_id:
            # Try fuzzy: remove parenthetical suffixes
            for key, val in name_to_id.items():
                if constituency_name in key or key in constituency_name:
                    constituency_id = val
                    break
            if not constituency_id:
                print(f"  WARNING: Constituency '{constituency_name}' not matched, skipping {candidate_name}")
                continue

        party_id = party_map.get(party_name)
        if not party_id:
            print(f"  WARNING: Party '{party_name}' not found, skipping {candidate_name}")
            continue

        candidate = Candidate(
            election_id=election_id,
            constituency_id=constituency_id,
            party_id=party_id,
            name=candidate_name,
            status=status or "Accepted",
            is_contesting=(contesting == "Yes"),
            profile_url=profile_url,
        )
        db.add(candidate)
        count += 1

    db.flush()
    wb.close()
    return count


def _print_summary(db: Session, election_id: int):
    """Print import summary."""
    from sqlalchemy import func as sqlfunc
    print(f"\n--- Summary for Election ID {election_id} ---")
    print(f"  Districts:       {db.query(District).filter_by(election_id=election_id).count()}")
    print(f"  Constituencies:  {db.query(Constituency).filter_by(election_id=election_id).count()}")
    print(f"  Polling Stations:{db.query(PollingStation).count()}")
    print(f"  Parties:         {db.query(Party).count()}")
    print(f"  Candidates:      {db.query(Candidate).filter_by(election_id=election_id).count()}")
    contesting = db.query(Candidate).filter_by(election_id=election_id, is_contesting=True).count()
    print(f"  Contesting:      {contesting}")


if __name__ == "__main__":
    import_all()
