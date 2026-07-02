"""Import Tripura 2023 election data — full detailed results + myneta enrichment.

Sources (Tripura/Tripura_2023/):
  - Detailed Results:  10-Detailed Results.xlsx
  - Electors data:     Electors & Voters Data.xlsx
  - Party list:        3-List Of Political Parties Participated.xlsx
  - Candidate CSV:     Tripura_candidates_2023.csv
  - Profile JSON:      2023_candidate_profile.json
"""
import sys, os, re, csv, json
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from app.database import engine, Base
from app.models.tripura import Election, District, Constituency, Party, Candidate

PROJECT_ROOT     = os.path.join(os.path.dirname(__file__), "..", "..")
DATA_DIR         = os.path.join(PROJECT_ROOT, "Tripura", "Tripura_2023")
DETAILED_RESULTS = os.path.join(DATA_DIR, "10-Detailed Results.xlsx")
ELECTORS_FILE    = os.path.join(DATA_DIR, "Electors & Voters Data.xlsx")
PARTY_LIST       = os.path.join(DATA_DIR, "3-List Of Political Parties Participated.xlsx")
PROFILES_JSON    = os.path.join(DATA_DIR, "2023_candidate_profile.json")
CANDIDATES_CSV   = os.path.join(DATA_DIR, "Tripura_candidates_2023.csv")
MAPPING_FILE     = os.path.join(PROJECT_ROOT, "Tripura", "Tripura_Election_Data_Mapping.xlsx")

PARTY_COLORS = {
    "BJP": "#FF9933", "INC": "#00BFFF", "CPI(M)": "#FF0000", "CPM": "#FF0000",
    "CPI": "#FF6600", "AITC": "#00FF00", "TMC": "#00FF00", "IPFT": "#FF8C00",
    "TMP": "#8B4513", "INPT": "#556B2F", "IND": "#808080", "NOTA": "#000000",
    "AAP": "#0066CC", "BSP": "#0000FF", "RSP": "#FF4444",
}

ABBR_ALIASES = {"CPM": "CPI(M)", "TMC": "AITC"}


def load_ac_category_map():
    """2023 AC cols (6,7) from mapping file → ac_no: category."""
    wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True)
    ws = wb["AC"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    cat_map = {}
    for row in rows[2:]:
        if not row[6] or not row[7]:
            continue
        try:
            ac_no = int(row[6])
        except (ValueError, TypeError):
            continue
        name = str(row[7]).strip()
        cat = "GEN"
        if "(SC)" in name:
            cat = "SC"
        elif "(ST)" in name:
            cat = "ST"
        cat_map[ac_no] = cat
    return cat_map


def load_electors_map():
    """AC-wise electors → ac_no: {male, female, third_gender, total, total_votes}."""
    wb = openpyxl.load_workbook(ELECTORS_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    out = {}
    for row in rows[1:]:
        if row[2] is None:
            continue
        try:
            ac_no = int(row[2])
        except (ValueError, TypeError):
            continue
        out[ac_no] = {
            "male":         int(row[4])  if row[4]  else 0,
            "female":       int(row[5])  if row[5]  else 0,
            "third_gender": int(row[6])  if row[6]  else 0,
            "total":        int(row[7])  if row[7]  else 0,
            "total_votes":  int(row[13]) if row[13] else 0,
        }
    return out


def load_party_names():
    """Format: PARTY TYPE, SL No, ABBREVIATION, PARTY Name."""
    wb = openpyxl.load_workbook(PARTY_LIST, read_only=True)
    ws = wb[wb.sheetnames[0]]
    mapping = {}
    for row in ws.iter_rows(values_only=True):
        if isinstance(row[1], (int, float)) and row[2] and row[3]:
            abbr = str(row[2]).strip()
            canon = ABBR_ALIASES.get(abbr, abbr)
            mapping[canon] = str(row[3]).strip()
    wb.close()
    return mapping


def _parse_rupees(val):
    if not val or "Nil" in str(val):
        return 0
    s = str(val).strip()
    if s.startswith("http"):
        return 0
    digits = re.sub(r"[^\d]", "", s.split("~")[0])
    return int(digits) if digits else 0


def _parse_criminal(status):
    if not status or "No criminal" in str(status):
        return 0
    m = re.search(r"(\d+)", str(status))
    return int(m.group(1)) if m else 0


def load_profiles():
    """Merge CSV + JSON → dict keyed by (constituency_upper, name_upper)."""
    csv_by_id = {}
    with open(CANDIDATES_CSV, encoding="latin-1") as f:
        for row in csv.DictReader(f):
            cid = row.get("candidate_id", "").strip()
            if cid:
                csv_by_id[cid] = row

    with open(PROFILES_JSON, encoding="utf-8") as f:
        profiles = json.load(f)

    by_key = {}
    for p in profiles:
        cp = p.get("candidate_profile", {})
        cid = str(cp.get("candidate_id", ""))
        csv_r = csv_by_id.get(cid, {})
        con = (csv_r.get("constituency") or cp.get("ac_name") or "").strip().upper()
        name = (csv_r.get("candidate_name") or cp.get("name") or "").strip()
        if con and name:
            by_key[(con, name.upper())] = {**p, "_csv": csv_r}
    return by_key


def import_data():
    Base.metadata.create_all(bind=engine)

    cat_map    = load_ac_category_map()
    electors   = load_electors_map()
    party_names = load_party_names()

    # Cols: state(0), district(1), ac_no(2), ac_name(3), cand_no(4), name(5),
    # sex(6), age(7), category(8), party(9), symbol(10),
    # gen_votes(11), postal_votes(12), total_votes(13), vote_pct(14), total_electors(15)
    wb = openpyxl.load_workbook(DETAILED_RESULTS, read_only=True)
    ws = wb[wb.sheetnames[0]]
    data_rows = list(ws.iter_rows(values_only=True))[1:]
    wb.close()

    with Session(engine) as session:
        election = Election(
            state="Tripura", year=2023, type="Assembly",
            name="Tripura Legislative Assembly Election 2023",
        )
        session.add(election)
        session.flush()
        print(f"Created election: {election.name} (id={election.id})")

        party_cache = {p.abbr.upper(): p for p in session.query(Party).all()}
        for abbr, full_name in party_names.items():
            key = abbr.upper()
            if key not in party_cache:
                p = Party(name=full_name, abbr=abbr[:50], color=PARTY_COLORS.get(abbr))
                session.add(p)
                session.flush()
                party_cache[key] = p

        district_cache = {}
        ac_groups: dict[int, list] = {}
        ac_meta: dict[int, dict] = {}

        for row in data_rows:
            if row[2] is None:
                continue
            try:
                ac_no = int(row[2])
            except (ValueError, TypeError):
                continue
            dist = str(row[1]).strip().title() if row[1] else "Unknown"
            ac_name = str(row[3]).strip().upper() if row[3] else ""
            # Strip category suffix from AC name
            ac_name_clean = re.sub(r"\s*\((SC|ST)\)\s*$", "", ac_name, flags=re.I).strip()
            if ac_no not in ac_meta:
                ac_meta[ac_no] = {"district": dist, "ac_name": ac_name_clean}
            ac_groups.setdefault(ac_no, []).append(row)

        for ac_no in sorted(ac_groups.keys()):
            meta = ac_meta[ac_no]
            cand_rows = ac_groups[ac_no]
            dist_name = meta["district"]

            if dist_name not in district_cache:
                d = District(election_id=election.id, name=dist_name)
                session.add(d)
                session.flush()
                district_cache[dist_name] = d

            el = electors.get(ac_no, {})
            valid = [(r, int(r[13]) if r[13] else 0) for r in cand_rows if r[5]]
            valid.sort(key=lambda x: -x[1])

            winning_margin = (valid[0][1] - valid[1][1]) if len(valid) >= 2 else None
            total_votes  = el.get("total_votes") or (int(cand_rows[0][15]) if cand_rows and cand_rows[0][15] else None)
            total_electors = el.get("total") or (int(cand_rows[0][15]) if cand_rows and cand_rows[0][15] else None)
            turnout_pct  = round(total_votes / total_electors * 100, 2) if total_electors and total_votes else None

            con = Constituency(
                election_id=election.id,
                district_id=district_cache[dist_name].id,
                ac_no=ac_no, name=meta["ac_name"],
                category=cat_map.get(ac_no, "GEN"),
                total_electors=total_electors,
                male_electors=el.get("male"), female_electors=el.get("female"),
                third_gender_electors=el.get("third_gender"),
                total_votes_polled=total_votes, turnout_pct=turnout_pct,
                winning_margin=winning_margin,
            )
            session.add(con)
            session.flush()

            for pos, (row, _) in enumerate(valid, start=1):
                party_abbr = str(row[9]).strip() if row[9] else None
                party_abbr = ABBR_ALIASES.get(party_abbr, party_abbr)
                is_nota = bool(party_abbr and party_abbr.upper() == "NOTA")
                sex_raw = str(row[6]).strip().upper() if row[6] else None
                gender = "MALE" if sex_raw in ("M", "MALE") else ("FEMALE" if sex_raw in ("F", "FEMALE") else sex_raw)

                party = None
                if not is_nota and party_abbr:
                    key = party_abbr.upper()
                    if key not in party_cache:
                        p = Party(name=party_abbr, abbr=party_abbr[:50], color=PARTY_COLORS.get(party_abbr))
                        session.add(p)
                        session.flush()
                        party_cache[key] = p
                    party = party_cache[key]

                session.add(Candidate(
                    election_id=election.id, constituency_id=con.id,
                    party_id=party.id if party else None,
                    name=str(row[5]).strip(), gender=gender,
                    age=int(row[7]) if row[7] else None,
                    position=pos,
                    votes_general=int(row[11]) if row[11] else 0,
                    votes_postal=int(row[12]) if row[12] else 0,
                    votes_total=int(row[13]) if row[13] else 0,
                    vote_pct=float(row[14]) if row[14] else None,
                    is_nota=is_nota,
                ))

        session.commit()
        n_c = session.query(Constituency).filter_by(election_id=election.id).count()
        n_cand = session.query(Candidate).filter_by(election_id=election.id).count()
        print(f"Districts: {len(district_cache)}, Constituencies: {n_c}, Candidates: {n_cand}")


def enrich_profiles():
    profile_map = load_profiles()
    print(f"Loaded {len(profile_map)} profiles")

    with Session(engine) as session:
        el = session.query(Election).filter_by(state="Tripura", year=2023).first()
        if not el:
            print("Tripura 2023 not found — run import_data() first")
            return

        updated = missing = 0
        for con in session.query(Constituency).filter_by(election_id=el.id).all():
            for cand in session.query(Candidate).filter_by(constituency_id=con.id, is_nota=False).all():
                key = (con.name.upper(), cand.name.upper())
                prof = profile_map.get(key)
                if not prof:
                    missing += 1
                    continue

                cp = prof.get("candidate_profile", {})
                csv_r = prof.get("_csv", {})
                assets_s = prof.get("assets_summary", {})

                edu = (csv_r.get("education") or cp.get("education") or "").strip()
                if edu.startswith("Category:"):
                    edu = edu[len("Category:"):].strip()
                cand.education = edu if edu and edu not in ("Not mentioned", "Not Given", "") else None

                cand.criminal_cases = int(csv_r.get("criminal_cases") or 0) or _parse_criminal(cp.get("crime_status"))

                assets_raw = csv_r.get("assets", "")
                if not assets_raw or str(assets_raw).startswith("http"):
                    assets_val = _parse_rupees(assets_s.get("total_assets"))
                else:
                    assets_val = _parse_rupees(assets_raw)
                cand.declared_assets = assets_val if assets_val > 0 else None

                liab_raw = csv_r.get("liabilities", "")
                if not liab_raw or str(liab_raw).startswith("http"):
                    liab_val = _parse_rupees(assets_s.get("total_liabilities"))
                else:
                    liab_val = _parse_rupees(liab_raw)
                cand.liabilities = liab_val if liab_val > 0 else None

                occ = (prof.get("profession", {}).get("self") or "").strip()
                cand.occupation = occ if occ and occ != "Not mentioned" else None

                img = (cp.get("image_url") or "").strip()
                cand.image_url = img if img and img != "None" else None

                updated += 1

        session.commit()
        print(f"Enriched {updated} candidates, {missing} not matched")


if __name__ == "__main__":
    import_data()
    enrich_profiles()
