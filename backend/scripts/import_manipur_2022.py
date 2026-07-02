"""Import Manipur 2022 election data.

Data sources (all under Manipur/Manipur_2022/):
  - Results:  Detailed_Results_2022.xlsx
  - Electors: Electors & Voters Data_2022.xlsx
  - CSV:      Manipur_candidates_2022.csv
  - JSON:     2022_candidate_profile.json
  - Mapping:  Manipur/Manipur_Data_Mapping.xlsx

Results columns (0-indexed):
  STATE(0), District(1), AC_NO(2), AC_NAME(3), CandNo(4), CANDIDATE(5), SEX(6),
  AGE(7), CATEGORY(8), PARTY(9), SYMBOL(10), GENERAL(11), POSTAL(12),
  TOTAL(13), %(14), ELECTORS(15)

Electors columns:
  State(0), District(1), AC_NO(2), AC_Name(3), Male(4), Female(5), Third(6), Total(7)

Usage: cd backend && python -m scripts.import_manipur_2022
"""
import sys, os, re, json, csv
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from app.database import engine, Base
from app.models.manipur import Election, District, Constituency, Party, Candidate

PROJECT_ROOT   = os.path.join(os.path.dirname(__file__), "..", "..")
DATA_DIR       = os.path.join(PROJECT_ROOT, "Manipur", "Manipur_2022")
MAPPING_FILE   = os.path.join(PROJECT_ROOT, "Manipur", "Manipur_Data_Mapping.xlsx")
RESULTS_FILE   = os.path.join(DATA_DIR, "Detailed Results.xlsx")
ELECTORS_FILE  = os.path.join(DATA_DIR, "Electors & Voters Data_2022.xlsx")
PROFILES_JSON  = os.path.join(DATA_DIR, "2022_candidate_profile.json")
CANDIDATES_CSV = os.path.join(DATA_DIR, "Manipur_candidates_2022.csv")

PARTY_COLORS = {
    "BJP": "#FF9933", "INC": "#00BFFF", "IND": "#808080",
    "NPF": "#8B4513", "NPP": "#228B22", "AITC": "#00FF00",
    "CPI": "#FF4444", "CPI(M)": "#CC0000", "BSP": "#0000FF",
    "NCP": "#004080", "JDU": "#00CC66", "LJP": "#FFCC00",
    "NDA": "#FF6600", "KSD": "#DAA520",
}


def _normalize(name: str) -> str:
    return re.sub(r"[.\s@]+", "", str(name)).upper()


def load_ac_mapping():
    wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True)
    ws = wb["AC"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    name_to_ac = {}
    for row in rows[2:]:
        if not row[0]:
            continue
        ac_no = int(row[0])
        for col in [1, 4, 7]:
            if col < len(row) and row[col]:
                v = str(row[col]).strip()
                stripped = re.sub(r"\s*\(?(SC|ST)\)?\s*$", "", v, flags=re.I).strip()
                name_to_ac[_normalize(v)] = ac_no
                name_to_ac[_normalize(stripped)] = ac_no
        if len(row) > 5 and row[5]:
            comment = str(row[5])
            alt_m = re.search(r"showing as\s+(\w+)", comment, re.I)
            if alt_m:
                name_to_ac[_normalize(alt_m.group(1))] = ac_no
    name_to_ac.update({
        "NOURIYAPAKHANGLAKPA": 21,
        "NAORIAPAKHANGLAKPA":  21,
        "SAGOIBAND":           11,
        "THONJU":               5,
        "TIPAIMUK":            55,
    })
    return name_to_ac


def load_electors():
    """State(0), District(1), AC_NO(2), AC_Name(3), Male(4), Female(5), Third(6), Total(7)"""
    wb = openpyxl.load_workbook(ELECTORS_FILE, read_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()
    em = {}
    for row in rows[1:]:
        if not isinstance(row[2], (int, float)):
            continue
        ac_no = int(row[2])
        raw_name = re.sub(r"\s+", " ", str(row[3]).strip()) if row[3] else ""
        cat_m = re.search(r"\s*\((SC|ST)\)\s*$", raw_name, re.I)
        category = cat_m.group(1).upper() if cat_m else None
        name = re.sub(r"\s*\((SC|ST)\)\s*$", "", raw_name, flags=re.I).strip()
        em[ac_no] = {
            "name": name, "category": category,
            "male_electors":         int(row[4]) if row[4] else None,
            "female_electors":       int(row[5]) if row[5] else None,
            "third_gender_electors": int(row[6]) if row[6] else None,
            "total_electors":        int(row[7]) if row[7] else None,
        }
    return em


def load_profiles(name_to_ac):
    csv_by_id = {}
    with open(CANDIDATES_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            cid = row.get("candidate_id", "").strip()
            con_raw = re.sub(r"\s*\(?(SC|ST)\)?\s*$", "", row.get("constituency", "").strip(), flags=re.I).strip()
            if cid:
                csv_by_id[cid] = {
                    "name":           row.get("candidate_name", "").strip(),
                    "constituency":   con_raw,
                    "criminal_cases": int(row.get("criminal_cases") or 0),
                    "education":      row.get("education", "").strip(),
                    "assets":         row.get("assets", ""),
                    "liabilities":    row.get("liabilities", ""),
                }

    with open(PROFILES_JSON, encoding="utf-8") as f:
        profiles = json.load(f)

    by_acno = {}
    for p in profiles:
        cp   = p.get("candidate_profile", {})
        cid  = str(cp.get("candidate_id", ""))
        info = csv_by_id.pop(cid, None)

        ac_no = cand_name = None
        if info:
            ac_no     = name_to_ac.get(_normalize(info["constituency"]))
            cand_name = info["name"]
        if not ac_no:
            ac_raw = re.sub(r"\s*\(?(SC|ST)\)?\s*$", "", (cp.get("ac_name") or "").strip(), flags=re.I).strip()
            ac_no  = name_to_ac.get(_normalize(ac_raw))
            if not cand_name:
                cand_name = (cp.get("name") or "").strip()

        if ac_no and cand_name:
            by_acno.setdefault((ac_no, _normalize(cand_name)), {**p, "_csv": info})

    for cid, info in csv_by_id.items():
        ac_no = name_to_ac.get(_normalize(info["constituency"]))
        if ac_no and info["name"]:
            by_acno.setdefault((ac_no, _normalize(info["name"])),
                               {"candidate_profile": {}, "assets_summary": {}, "_csv": info})
    return by_acno


def parse_rupees(val):
    if not val or "Nil" in str(val):
        return 0
    s = str(val).strip()
    if s.startswith("http"):
        return 0
    digits = re.sub(r"[^\d]", "", s.split("~")[0])
    return int(digits) if digits else 0


def parse_criminal(status):
    if not status or "No criminal" in str(status):
        return 0
    m = re.search(r"(\d+)", str(status))
    return int(m.group(1)) if m else 0


def _name_words(name: str) -> frozenset:
    """Significant words (3+ chars) from name, stripping @alias prefixes."""
    base = re.split(r"\s*@", name)[0].strip()
    return frozenset(w for w in re.findall(r"[A-Za-z]{3,}", base) if w.upper() not in ("MRS", "MR", "DR", "SH", "SHRI"))


def build_word_map(profile_map: dict) -> dict:
    """Secondary index: (ac_no, frozenset_of_profile_words) → profile, for fuzzy fallback.
    Uses original name (with spaces) so word boundaries are preserved."""
    wm = {}
    for (ac_no, _), profile in profile_map.items():
        csv_info = profile.get("_csv")
        cp = profile.get("candidate_profile", {})
        orig = (csv_info.get("name") if csv_info else None) or (cp.get("name") or "")
        words = frozenset(w.upper() for w in re.findall(r"[A-Za-z]{3,}", orig)
                         if w.upper() not in ("MRS", "MR", "DR", "SH", "SHRI"))
        if words:
            wm.setdefault((ac_no, words), profile)
    return wm


def lookup_profile(ac_no, name, profile_map, word_map):
    """Exact match first; then @-stripped; then word-subset fallback."""
    prof = profile_map.get((ac_no, _normalize(name or "")))
    if prof:
        return prof
    # Strip @alias suffix and retry exact
    base = re.split(r"\s*@", name)[0].strip() if name else ""
    if base != name:
        prof = profile_map.get((ac_no, _normalize(base)))
        if prof:
            return prof
    # Word-subset: all significant words in results name appear in profile name
    rwords = frozenset(w.upper() for w in _name_words(base))
    if rwords:
        for (k_ac, k_words), profile in word_map.items():
            if k_ac == ac_no and rwords <= k_words:
                return profile
    return None


def run():
    Base.metadata.create_all(bind=engine)

    name_to_ac  = load_ac_mapping()
    em          = load_electors()
    profile_map = load_profiles(name_to_ac)
    word_map    = build_word_map(profile_map)
    print(f"Loaded {len(em)} ACs, {len(profile_map)} profiles")

    # Results: STATE(0), District(1), AC_NO(2), AC_NAME(3), CandNo(4), CANDIDATE(5),
    #          SEX(6), AGE(7), CATEGORY(8), PARTY(9), SYMBOL(10),
    #          GENERAL(11), POSTAL(12), TOTAL(13), %(14), ELECTORS(15)
    wb = openpyxl.load_workbook(RESULTS_FILE, read_only=True)
    result_rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()

    ac_data = {}
    for row in result_rows[1:]:
        if not isinstance(row[2], (int, float)):
            continue
        ac_no    = int(row[2])
        district = str(row[1]).strip() if row[1] else "Unknown"
        cat_raw  = str(row[8]).strip().upper() if row[8] else "GEN"
        category = "GEN" if cat_raw in ("GEN", "GENERAL") else cat_raw

        if ac_no not in ac_data:
            ac_data[ac_no] = {
                "district":              district,
                "name":                  str(row[3]).strip() if row[3] else "",
                "category":              category,
                "electors_from_results": int(row[15]) if row[15] else None,
                "candidates":            [],
            }
        sex_raw = str(row[6]).strip().upper() if row[6] else ""
        ac_data[ac_no]["candidates"].append({
            "name":         str(row[5]).strip() if row[5] else None,
            "sex":          "M" if sex_raw in ("M", "MALE") else ("F" if sex_raw in ("F", "FEMALE") else None),
            "age":          int(row[7]) if row[7] else None,
            "party":        str(row[9]).strip() if row[9] else None,
            "gen_votes":    int(row[11]) if row[11] else 0,
            "postal_votes": int(row[12]) if row[12] else 0,
            "total_votes":  int(row[13]) if row[13] else 0,
        })

    with Session(engine) as session:
        if session.query(Election).filter_by(state="Manipur", year=2022).first():
            print("Manipur 2022 already exists — skipping.")
            return

        election = Election(state="Manipur", year=2022, type="Assembly",
                            name="Manipur Legislative Assembly Election 2022")
        session.add(election)
        session.flush()

        all_districts = sorted(set(info["district"] for info in ac_data.values()))
        district_cache = {dname: District(election_id=election.id, name=dname)
                          for dname in all_districts}
        for d in district_cache.values():
            session.add(d)
        session.flush()

        constituency_cache = {}
        for ac_no, info in ac_data.items():
            elec = em.get(ac_no, {})
            dist = district_cache.get(info["district"])
            total_electors = elec.get("total_electors") or info["electors_from_results"]
            real = [c for c in info["candidates"] if c["party"] not in ("NOTA", None)
                    and c["name"] != "None of the Above"]
            real_sorted = sorted(real, key=lambda x: x["total_votes"], reverse=True)
            total_votes = sum(c["total_votes"] for c in info["candidates"])
            turnout = round(total_votes / total_electors * 100, 2) if total_electors and total_votes else None
            margin  = (real_sorted[0]["total_votes"] - real_sorted[1]["total_votes"]
                       if len(real_sorted) >= 2 else None)
            category = elec.get("category") or info.get("category") or "GEN"
            con = Constituency(
                election_id=election.id, district_id=dist.id if dist else None,
                ac_no=ac_no, name=elec.get("name") or info["name"],
                category=category,
                total_electors=total_electors,
                male_electors=elec.get("male_electors"),
                female_electors=elec.get("female_electors"),
                third_gender_electors=elec.get("third_gender_electors"),
                total_votes_polled=total_votes, turnout_pct=turnout, winning_margin=margin,
            )
            session.add(con)
            constituency_cache[ac_no] = con
        session.flush()
        print(f"Created {len(constituency_cache)} constituencies")

        party_cache = {p.abbr.upper(): p for p in session.query(Party).all()}
        imported = 0

        for ac_no, info in ac_data.items():
            con  = constituency_cache[ac_no]
            real = [c for c in info["candidates"] if c["party"] not in ("NOTA", None)
                    and c["name"] != "None of the Above"]
            nota = [c for c in info["candidates"] if c["party"] == "NOTA"
                    or c["name"] == "None of the Above"]
            sorted_cands = sorted(real, key=lambda x: x["total_votes"], reverse=True) + nota
            total_valid  = sum(c["total_votes"] for c in sorted_cands)

            for pos, cand in enumerate(sorted_cands, start=1):
                is_nota = cand["party"] == "NOTA" or cand["name"] == "None of the Above"
                name    = "NOTA" if is_nota else cand["name"]

                party = None
                if not is_nota and cand["party"]:
                    abbr = cand["party"].strip()
                    if abbr.upper() not in party_cache:
                        p = Party(name=abbr, abbr=abbr, color=PARTY_COLORS.get(abbr))
                        session.add(p); session.flush()
                        party_cache[abbr.upper()] = p
                    party = party_cache[abbr.upper()]

                vote_pct  = round(cand["total_votes"] / total_valid * 100, 2) if total_valid else 0
                prof      = lookup_profile(ac_no, name or "", profile_map, word_map)
                csv_info  = prof.get("_csv") if prof else None
                cp        = (prof or {}).get("candidate_profile", {})
                assets_s  = (prof or {}).get("assets_summary", {})

                # Use myneta canonical name (title-cased CSV name preferred)
                if not is_nota and prof:
                    if csv_info and csv_info.get("name", "").strip():
                        name = csv_info["name"].strip()
                    elif cp and cp.get("name", "").strip():
                        name = cp["name"].strip().title()

                education = declared_assets = liabilities_val = image_url = None
                criminal_cases = 0

                if csv_info:
                    edu = csv_info.get("education", "").strip()
                    if edu and edu not in ("Not mentioned", "Not Given", ""):
                        education = edu
                    criminal_cases = int(csv_info.get("criminal_cases") or 0)
                elif cp:
                    edu = (cp.get("education") or "").strip()
                    if edu.startswith("Category:"):
                        edu = edu[len("Category:"):].strip()
                    if edu and edu not in ("Not mentioned", "Not Given"):
                        education = edu
                    criminal_cases = parse_criminal(cp.get("crime_status"))

                # Assets/liabilities: CSV first, fall through to JSON when CSV has URL or returns 0
                a = parse_rupees(csv_info.get("assets")) if csv_info else 0
                if a > 0:
                    declared_assets = a
                elif assets_s:
                    a = parse_rupees(assets_s.get("total_assets"))
                    if a > 0: declared_assets = a

                l = parse_rupees(csv_info.get("liabilities")) if csv_info else 0
                if l > 0:
                    liabilities_val = l
                elif assets_s:
                    l = parse_rupees(assets_s.get("total_liabilities"))
                    if l > 0: liabilities_val = l

                if cp:
                    img = (cp.get("image_url") or "").strip()
                    # 2022 JSON has incomplete base paths ending with "/"
                    if img and img != "None" and not img.endswith("/"):
                        image_url = img

                session.add(Candidate(
                    election_id=election.id, constituency_id=con.id,
                    party_id=party.id if party else None, name=name,
                    gender=cand["sex"], age=cand["age"], position=pos,
                    votes_general=cand["gen_votes"], votes_postal=cand["postal_votes"],
                    votes_total=cand["total_votes"], vote_pct=vote_pct, is_nota=is_nota,
                    education=education, declared_assets=declared_assets,
                    liabilities=liabilities_val, criminal_cases=criminal_cases,
                    image_url=image_url,
                ))
                imported += 1

        session.commit()
        election_id = election.id

    print(f"\n=== Manipur 2022 ===")
    print(f"Constituencies: {len(constituency_cache)}, Candidates: {imported}")
    with Session(engine) as session:
        winners = session.query(Candidate).filter_by(
            election_id=election_id, position=1, is_nota=False
        ).all()
        seats = {}
        for w in winners:
            p = w.party.abbr if w.party else "IND"
            seats[p] = seats.get(p, 0) + 1
        for p, s in sorted(seats.items(), key=lambda x: -x[1]):
            print(f"  {p}: {s} seats")


if __name__ == "__main__":
    run()
