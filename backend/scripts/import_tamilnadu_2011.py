"""Import Tamil Nadu 2011 election data.

Primary source:  my_neta CSV/JSON (~1046 candidates with affidavit profile data)
Vote enrichment: Constituency_Wise_Voters&Electors.xlsx (all 2744, provides vote totals)
Constituency stats (turnout, winning_margin, total_votes_polled) computed from full 2744.

Usage: cd backend && python -m scripts.import_tamilnadu_2011
"""
import sys, os, re, json, csv
from collections import defaultdict
from difflib import SequenceMatcher
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from app.database import engine, Base
from app.models.tamilnadu import Election, District, Constituency, Party, Candidate

PROJECT_ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
DATA_DIR = os.path.join(PROJECT_ROOT, "tamil_nadu", "Tamil Nadu_2011")

ELECTORS_FILE  = os.path.join(DATA_DIR, "Dist_and_AC_Electors & Voters Data_2011.xlsx")
VOTERS_FILE    = os.path.join(DATA_DIR, "Constituency_Wise_Voters&Electors.xlsx")
PARTIES_FILE   = os.path.join(DATA_DIR, "List of Political Parties Participated_2011.xlsx")
PROFILES_JSON  = os.path.join(DATA_DIR, "my_neta_2011_candidate_profile.json")
CANDIDATES_CSV = os.path.join(DATA_DIR, "my_neta_tamilnadu_candidates_2011.csv")

PARTY_COLORS = {
    "DMK": "#FF0000", "AIADMK": "#006400", "INC": "#00BFFF",
    "BJP": "#FF9933", "CPI(M)": "#FF0000", "CPI": "#FF4444",
    "PMK": "#FFFF00", "AITC": "#00FF00", "MDMK": "#FF6600",
    "VCK": "#0000FF", "DMDK": "#800080", "IND": "#808080",
}

ABBR_ALIASES = {"CPM": "CPI(M)", "ADMK": "AIADMK"}


def _normalize(name):
    return re.sub(r"[.\s@\-,]+", "", name).upper()


def _name_score(a: str, b: str) -> float:
    """Return match score 0-1. Handles initial-position swaps."""
    na, nb = _normalize(a), _normalize(b)
    if na == nb:
        return 1.0
    if sorted(na) == sorted(nb):
        return 0.99  # same chars, different order (e.g. 'SEKARCH' vs 'CHSEKAR')
    if (na in nb or nb in na) and min(len(na), len(nb)) >= 5:
        return 0.95
    return SequenceMatcher(None, na, nb).ratio()


def parse_rupees(val):
    if not val or str(val).strip() in ("", "Nil", "Rs\xa00 ~", "Rs 0 ~"):
        return None
    cleaned = re.sub(r"[Rs\s\xa0,~]", "", str(val))
    cleaned = cleaned.split("Lac")[0].split("Cr")[0].strip()
    digits = re.match(r"^([\d.]+)", cleaned)
    if not digits:
        return None
    v = int(float(digits.group(1)))
    return v if v > 0 else None


def parse_criminal(status):
    if not status or "No criminal" in status:
        return 0
    m = re.search(r"(\d+)", status)
    return int(m.group(1)) if m else 0


def load_electors():
    # Cols: District(0), AC No(1), AC Name(2), Type(3), Male(4), Female(5), Third(6), Total(7)
    wb = openpyxl.load_workbook(ELECTORS_FILE, read_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()
    em = {}
    for row in rows[1:]:
        if not row[1]:
            continue
        try:
            ac_no = int(row[1])
        except (TypeError, ValueError):
            continue
        em[ac_no] = {
            "district": str(row[0]).strip().upper() if row[0] else "UNKNOWN",
            "name": str(row[2]).strip().upper() if row[2] else "",
            "category": str(row[3]).strip().upper() if row[3] else "GEN",
            "male_electors": int(row[4]) if row[4] else None,
            "female_electors": int(row[5]) if row[5] else None,
            "third_gender_electors": int(row[6]) if row[6] else None,
            "total_electors": int(row[7]) if row[7] else None,
        }
    return em


def load_voters_index(em):
    """Build per-AC vote lookup and constituency stats from the full voters file.

    Returns:
        ac_stats: {ac_no: {total_votes_polled, winning_margin, turnout_pct}}
        ac_vote_lookup: {ac_no: {norm_name: {votes_general, votes_postal, votes_total, position, gender, age}}}
    """
    # Cols: STATE(0), AC NO(1), AC NAME(2), SL NO(3), NAME(4), SEX(5), AGE(6),
    #       CAT(7), PARTY(8), GENERAL(9), POSTAL(10), TOTAL(11), %(12), ELECTORS(13)
    wb = openpyxl.load_workbook(VOTERS_FILE, read_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()

    ac_raw: dict[int, list] = defaultdict(list)
    for row in rows[1:]:
        if not row[1]:
            continue
        try:
            ac_no = int(row[1])
        except (TypeError, ValueError):
            continue
        sex = str(row[5]).strip().upper() if row[5] else ""
        gender = "M" if sex in ("M", "MALE") else ("F" if sex in ("F", "FEMALE") else None)
        ac_raw[ac_no].append({
            "name": str(row[4]).strip() if row[4] else "",
            "gender": gender,
            "age": int(row[6]) if row[6] else None,
            "votes_general": int(row[9]) if row[9] else 0,
            "votes_postal": int(row[10]) if row[10] else 0,
            "votes_total": int(row[11]) if row[11] else 0,
        })

    ac_stats = {}
    ac_vote_lookup = {}

    for ac_no, cands in ac_raw.items():
        sorted_cands = sorted(cands, key=lambda x: x["votes_total"], reverse=True)
        total_votes = sum(c["votes_total"] for c in cands)
        winning_margin = (
            sorted_cands[0]["votes_total"] - sorted_cands[1]["votes_total"]
            if len(sorted_cands) >= 2 else None
        )
        elec_info = em.get(ac_no, {})
        turnout_pct = (
            round(total_votes / elec_info["total_electors"] * 100, 2)
            if elec_info.get("total_electors") else None
        )
        ac_stats[ac_no] = {
            "total_votes_polled": total_votes,
            "winning_margin": winning_margin,
            "turnout_pct": turnout_pct,
        }
        # Build lookup: norm_name → vote info + rank
        lookup = {}
        for pos, c in enumerate(sorted_cands, start=1):
            lookup[_normalize(c["name"])] = {
                "votes_general": c["votes_general"],
                "votes_postal": c["votes_postal"],
                "votes_total": c["votes_total"],
                "vote_pct": round(c["votes_total"] / total_votes * 100, 2) if total_votes else None,
                "position": pos,
                "gender": c["gender"],
                "age": c["age"],
            }
        ac_vote_lookup[ac_no] = lookup

    return ac_stats, ac_vote_lookup


def find_voter_data(ac_no: int, csv_name: str, ac_vote_lookup: dict) -> dict:
    """Find vote data for a CSV candidate by fuzzy name match within their AC."""
    lookup = ac_vote_lookup.get(ac_no, {})
    if not lookup:
        return {}
    norm = _normalize(csv_name)
    # Exact match
    if norm in lookup:
        return lookup[norm]
    # Scored match: sorted-char anagram + fuzzy fallback
    best_score, best_data = 0.0, {}
    for vname, vdata in lookup.items():
        score = _name_score(norm, vname)
        if score > best_score:
            best_score, best_data = score, vdata
    return best_data if best_score >= 0.85 else {}


def load_party_names():
    wb = openpyxl.load_workbook(PARTIES_FILE, read_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()
    return {str(r[2]).strip(): str(r[3]).strip() for r in rows[1:] if r[2] and r[3]}


def load_profiles_and_csv():
    profiles_by_id = {}
    if os.path.exists(PROFILES_JSON):
        with open(PROFILES_JSON, encoding="utf-8") as f:
            for item in json.load(f):
                cp = item.get("candidate_profile", {})
                cid = str(cp.get("candidate_id", "")).strip()
                if cid:
                    profiles_by_id[cid] = item
    csv_rows = []
    if os.path.exists(CANDIDATES_CSV):
        with open(CANDIDATES_CSV, encoding="utf-8") as f:
            csv_rows = list(csv.DictReader(f))
    return profiles_by_id, csv_rows


def run():
    Base.metadata.create_all(bind=engine)

    em = load_electors()
    ac_stats, ac_vote_lookup = load_voters_index(em)
    party_names = load_party_names()
    profiles_by_id, csv_rows = load_profiles_and_csv()

    total_voter_cands = sum(len(v) for v in ac_vote_lookup.values())
    print(f"Loaded {len(em)} ACs from electors, {total_voter_cands} candidates from voters file")
    print(f"Primary source: {len(csv_rows)} CSV candidates, {len(profiles_by_id)} JSON profiles")

    # Build constituency name → ac_no resolver (from electors file)
    con_name_to_ac = {_normalize(info["name"]): ac_no for ac_no, info in em.items() if info.get("name")}

    def resolve_ac_no(csv_con_name: str) -> int | None:
        norm = _normalize(csv_con_name)
        if norm in con_name_to_ac:
            return con_name_to_ac[norm]
        best_ratio, best_ac = 0.0, None
        for key, ac_no in con_name_to_ac.items():
            r = SequenceMatcher(None, norm, key).ratio()
            if r > best_ratio:
                best_ratio, best_ac = r, ac_no
        return best_ac if best_ratio >= 0.85 else None

    district_set = {info["district"] for info in em.values()
                    if info.get("district") and info["district"] != "UNKNOWN"}

    with Session(engine) as session:
        if session.query(Election).filter_by(state="Tamil Nadu", year=2011).first():
            print("Tamil Nadu 2011 already exists — skipping.")
            return

        election = Election(
            state="Tamil Nadu", year=2011, type="Assembly",
            name="Tamil Nadu Legislative Assembly Election 2011",
        )
        session.add(election)
        session.flush()

        if not district_set:
            district_set.add("Tamil Nadu")
        district_cache = {dname: District(election_id=election.id, name=dname)
                          for dname in sorted(district_set)}
        for d in district_cache.values():
            session.add(d)
        session.flush()
        default_district = list(district_cache.values())[0]

        # Create all 234 constituencies with stats from full voters file
        constituency_cache = {}
        for ac_no in sorted(em.keys()):
            elec_info = em[ac_no]
            dist_name = elec_info.get("district", "UNKNOWN")
            dist = district_cache.get(dist_name, default_district)
            con_name = re.sub(r"\s+", " ", elec_info.get("name", f"CONSTITUENCY_{ac_no}")).strip()
            category = elec_info.get("category", "GEN")
            if category not in ("GEN", "SC", "ST"):
                category = "GEN"
            stats = ac_stats.get(ac_no, {})
            con = Constituency(
                election_id=election.id,
                district_id=dist.id,
                ac_no=ac_no,
                name=con_name,
                category=category,
                total_electors=elec_info.get("total_electors"),
                male_electors=elec_info.get("male_electors"),
                female_electors=elec_info.get("female_electors"),
                third_gender_electors=elec_info.get("third_gender_electors"),
                total_votes_polled=stats.get("total_votes_polled"),
                winning_margin=stats.get("winning_margin"),
                turnout_pct=stats.get("turnout_pct"),
            )
            session.add(con)
            constituency_cache[ac_no] = con
        session.flush()
        print(f"Created {len(constituency_cache)} constituencies")

        party_cache = {p.abbr.upper(): p for p in session.query(Party).all()}

        def get_or_create_party(abbr):
            abbr = ABBR_ALIASES.get(abbr, abbr) if abbr else "IND"
            if not abbr or abbr.upper() in ("", "NOTA"):
                return None
            key = abbr.upper()
            if key not in party_cache:
                full_name = party_names.get(abbr, abbr)
                p = Party(name=full_name, abbr=abbr, color=PARTY_COLORS.get(abbr))
                session.add(p)
                session.flush()
                party_cache[key] = p
            return party_cache[key]

        imported = skipped = vote_matched = 0

        for row in csv_rows:
            cid = str(row.get("candidate_id", "")).strip()
            cname = row.get("candidate_name", "").strip()
            con_raw = re.sub(r"\s*\((SC|ST)\)\s*$", "", row.get("constituency", "").strip(), flags=re.I).upper()

            ac_no = resolve_ac_no(con_raw)
            if not ac_no:
                skipped += 1
                continue

            con = constituency_cache.get(ac_no)
            if not con:
                skipped += 1
                continue

            # Look up vote data from voters file
            voter = find_voter_data(ac_no, cname, ac_vote_lookup)
            if voter:
                vote_matched += 1

            # Profile from JSON
            prof = profiles_by_id.get(cid, {})
            cp = prof.get("candidate_profile", {})

            # Gender: voters file > JSON
            gender = voter.get("gender") if voter else None
            if not gender:
                g_raw = (cp.get("gender") or "").strip().upper()
                gender = "M" if g_raw in ("M", "MALE") else ("F" if g_raw in ("F", "FEMALE") else None)

            # Age: voters file > JSON
            age = voter.get("age") if voter else None
            if not age:
                age_str = (cp.get("age") or "").strip()
                age = int(age_str) if age_str.lstrip(" ").isdigit() else None

            # Education
            edu_raw = row.get("education", "").strip() or (cp.get("education") or "").strip()
            if edu_raw.startswith("Category:"):
                edu_raw = edu_raw[len("Category:"):].strip()
            education = edu_raw if edu_raw not in ("Not mentioned", "Not Given", "Unknown", "", None) else None

            # Occupation
            occupation = (prof.get("profession", {}).get("self") or "").strip() or None

            # Criminal cases
            criminal_raw = row.get("criminal_cases", "0").strip()
            criminal_cases = int(criminal_raw) if criminal_raw.isdigit() else parse_criminal(cp.get("crime_status", ""))

            # Party
            party_raw = ABBR_ALIASES.get(row.get("party", "").strip(), row.get("party", "").strip())
            party = get_or_create_party(party_raw)

            # Assets/liabilities: JSON assets_summary is more reliable (CSV column often has URLs)
            summary = prof.get("assets_summary", {})
            declared_assets = (
                parse_rupees(summary.get("total_assets"))
                or parse_rupees(row.get("assets", ""))
            )
            liabilities = (
                parse_rupees(summary.get("total_liabilities"))
                or parse_rupees(row.get("liabilities", ""))
            )

            session.add(Candidate(
                election_id=election.id,
                constituency_id=con.id,
                party_id=party.id if party else None,
                name=cname,
                gender=gender,
                age=age,
                position=voter.get("position") if voter else None,
                votes_general=voter.get("votes_general") if voter else None,
                votes_postal=voter.get("votes_postal") if voter else None,
                votes_total=voter.get("votes_total") if voter else None,
                vote_pct=voter.get("vote_pct") if voter else None,
                is_nota=False,
                education=education,
                occupation=occupation,
                declared_assets=declared_assets,
                liabilities=liabilities,
                criminal_cases=criminal_cases,
                image_url=(cp.get("image_url") or "").strip() or None,
            ))
            imported += 1

        session.commit()
        election_id = election.id

    print(f"\n=== Tamil Nadu 2011 ===")
    print(f"Candidates: {imported} imported, {skipped} skipped, {vote_matched} with vote data")

    with Session(engine) as session:
        from sqlalchemy import func
        winners_db = session.query(Candidate).filter_by(election_id=election_id, position=1).all()
        seats = {}
        for w in winners_db:
            p = w.party.abbr if w.party else "IND"
            seats[p] = seats.get(p, 0) + 1
        for p, s in sorted(seats.items(), key=lambda x: -x[1]):
            print(f"  {p}: {s} seats")
        total_votes = session.query(func.sum(Constituency.total_votes_polled)).filter_by(election_id=election_id).scalar()
        avg_turnout = session.query(func.avg(Constituency.turnout_pct)).filter_by(election_id=election_id).scalar()
        print(f"Total votes polled: {total_votes:,}")
        print(f"Average turnout: {float(avg_turnout):.1f}%")


if __name__ == "__main__":
    run()
