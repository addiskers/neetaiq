"""Import Goa 2017 election data.

Data sources (all under Goa/Goa_2017/):
  - Results:  Detailed Results_2017.xlsx
  - Electors: Electors & Voters Data_2017.xlsx
  - CSV bridge: Goa_candidates_2017.csv
  - JSON profiles: 2017_candidate_profile.json
  - Mapping: Goa/Goa_Data_Mapping.xlsx

Usage: cd backend && python -m scripts.import_goa_2017
"""
import sys, os, re, json, csv
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session
from app.database import engine, Base
from app.models.goa import Election, District, Constituency, Party, Candidate

PROJECT_ROOT = os.path.join(os.path.dirname(__file__), "..", "..")
DATA_DIR     = os.path.join(PROJECT_ROOT, "Goa", "Goa_2017")
MAPPING_FILE = os.path.join(PROJECT_ROOT, "Goa", "Goa_Data_Mapping.xlsx")

RESULTS_FILE   = os.path.join(DATA_DIR, "Detailed Results_2017.xlsx")
ELECTORS_FILE  = os.path.join(DATA_DIR, "Electors & Voters Data_2017.xlsx")
PROFILES_JSON  = os.path.join(DATA_DIR, "2017_candidate_profile.json")
CANDIDATES_CSV = os.path.join(DATA_DIR, "Goa_candidates_2017.csv")

PARTY_COLORS = {
    "BJP": "#FF9933", "INC": "#00BFFF", "MGP": "#006400", "AAP": "#0066CC",
    "GFP": "#FF4500", "NCP": "#004080", "IND": "#808080", "AITC": "#00FF00",
    "BSP": "#0000FF", "CPI": "#FF4444", "CPI(M)": "#CC0000",
}

AC_DISTRICT = {**{n: "North Goa" for n in range(1, 24)},
               **{n: "South Goa" for n in range(24, 41)}}


def _normalize(name: str) -> str:
    return re.sub(r"[.\s@]+", "", str(name)).upper()


def load_ac_mapping():
    wb = openpyxl.load_workbook(MAPPING_FILE, read_only=True)
    ws = wb["AC"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    name_to_ac = {}
    for row in rows[2:]:
        if not row[0]:
            continue
        ac_no = int(row[0])
        for col in [1, 4, 7]:
            if col < len(row) and row[col]:
                v = str(row[col]).strip()
                stripped = re.sub(r"\s*\(?(SC|ST)\)?\s*$", "", v, flags=re.I).strip()
                name_to_ac[_normalize(v)] = ac_no
                name_to_ac[_normalize(stripped)] = ac_no
    return name_to_ac


def load_electors():
    # Cols: District(0), AC NO(1), AC Name(2), MALE(3), FEMALE(4), Third(5), TOTAL(6)
    wb = openpyxl.load_workbook(ELECTORS_FILE, read_only=True)
    rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()
    em = {}
    for row in rows[1:]:
        if not isinstance(row[1], (int, float)):
            continue
        ac_no = int(row[1])
        raw_name = re.sub(r"\s+", " ", str(row[2]).strip()) if row[2] else ""
        cat_m = re.search(r"\s*\((SC|ST)\)\s*$", raw_name, re.I)
        category = cat_m.group(1).upper() if cat_m else "GEN"
        name = re.sub(r"\s*\((SC|ST)\)\s*$", "", raw_name, flags=re.I).strip()
        em[ac_no] = {
            "name": name, "category": category,
            "male_electors": int(row[3]) if row[3] else None,
            "female_electors": int(row[4]) if row[4] else None,
            "third_gender_electors": int(row[5]) if row[5] else None,
            "total_electors": int(row[6]) if row[6] else None,
        }
    return em


def load_profiles(name_to_ac):
    csv_by_id = {}
    with open(CANDIDATES_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            cid = row.get("candidate_id", "").strip()
            con_raw = re.sub(r"\s*\(?(SC|ST)\)?\s*$", "", row.get("constituency", "").strip(), flags=re.I).strip()
            if cid:
                csv_by_id[cid] = {
                    "name": row.get("candidate_name", "").strip(),
                    "constituency": con_raw,
                    "criminal_cases": int(row.get("criminal_cases") or 0),
                    "education": row.get("education", "").strip(),
                    "assets": row.get("assets", ""),
                    "liabilities": row.get("liabilities", ""),
                }

    with open(PROFILES_JSON, encoding="utf-8") as f:
        profiles = json.load(f)

    by_acno = {}
    for p in profiles:
        cp = p.get("candidate_profile", {})
        cid = str(cp.get("candidate_id", ""))
        info = csv_by_id.pop(cid, None)
        ac_no = None
        cand_name = None
        if info:
            ac_no = name_to_ac.get(_normalize(info["constituency"]))
            cand_name = info["name"]
        if not ac_no:
            ac_raw = re.sub(r"\s*\(?(SC|ST)\)?\s*$", "", (cp.get("ac_name") or "").strip(), flags=re.I).strip()
            ac_no = name_to_ac.get(_normalize(ac_raw))
            if not cand_name:
                cand_name = (cp.get("name") or "").strip()
        if ac_no and cand_name:
            by_acno.setdefault((ac_no, _normalize(cand_name)), {**p, "_csv": info})

    for cid, info in csv_by_id.items():
        ac_no = name_to_ac.get(_normalize(info["constituency"]))
        if ac_no and info["name"]:
            by_acno.setdefault((ac_no, _normalize(info["name"])),
                               {"candidate_profile": {}, "assets_summary": {}, "_csv": info})
    return by_acno


def parse_rupees(val):
    if not val or "Nil" in str(val):
        return 0
    cleaned = re.sub(r"[Rs\s\xa0,~]", "", str(val))
    m = re.match(r"^([\d.]+)", cleaned)
    return int(float(m.group(1))) if m else 0


def parse_criminal(status):
    if not status or "No criminal" in str(status):
        return 0
    m = re.search(r"(\d+)", str(status))
    return int(m.group(1)) if m else 0


def run():
    Base.metadata.create_all(bind=engine)

    name_to_ac = load_ac_mapping()
    em = load_electors()
    profile_map = load_profiles(name_to_ac)
    print(f"Loaded {len(em)} ACs, {len(profile_map)} profiles")

    # Results: District(0), AC NO(1), AC Name(2), Candidate Name(3), Sex(4), Age(5),
    #          Category(6), Party(7), General(8), Postal(9), Total(10),
    #          Total Electors(11), Total+NOTA(12), %(13)
    wb = openpyxl.load_workbook(RESULTS_FILE, read_only=True)
    result_rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    wb.close()

    ac_data = {}
    for row in result_rows[1:]:
        if not isinstance(row[1], (int, float)):
            continue
        ac_no = int(row[1])
        if ac_no not in ac_data:
            ac_data[ac_no] = {
                "district": str(row[0]).strip() if row[0] else AC_DISTRICT.get(ac_no, "North Goa"),
                "name": str(row[2]).strip() if row[2] else "",
                "electors_from_results": int(row[11]) if row[11] else None,
                "candidates": [],
            }
        sex_raw = str(row[4]).strip().upper() if row[4] else ""
        party_raw = str(row[7]).strip() if row[7] else None
        name_raw = str(row[3]).strip() if row[3] else None
        ac_data[ac_no]["candidates"].append({
            "name": name_raw,
            "sex": "M" if sex_raw in ("M", "MALE") else ("F" if sex_raw in ("F", "FEMALE") else None),
            "age": int(row[5]) if row[5] else None,
            "party": party_raw,
            "gen_votes": int(row[8]) if row[8] else 0,
            "postal_votes": int(row[9]) if row[9] else 0,
            "total_votes": int(row[10]) if row[10] else 0,
        })

    with Session(engine) as session:
        if session.query(Election).filter_by(state="Goa", year=2017).first():
            print("Goa 2017 already exists — skipping.")
            return

        election = Election(state="Goa", year=2017, type="Assembly",
                            name="Goa Legislative Assembly Election 2017")
        session.add(election)
        session.flush()

        district_cache = {dname: District(election_id=election.id, name=dname)
                          for dname in sorted(set(AC_DISTRICT.values()))}
        for d in district_cache.values():
            session.add(d)
        session.flush()

        constituency_cache = {}
        for ac_no, info in ac_data.items():
            elec = em.get(ac_no, {})
            dist = district_cache.get(info["district"])
            total_electors = elec.get("total_electors") or info["electors_from_results"]
            real = [c for c in info["candidates"] if c["party"] not in ("NOTA", None) and c["name"] != "None of the Above"]
            real_sorted = sorted(real, key=lambda x: x["total_votes"], reverse=True)
            total_votes = sum(c["total_votes"] for c in info["candidates"])
            turnout = round(total_votes / total_electors * 100, 2) if total_electors and total_votes else None
            margin = real_sorted[0]["total_votes"] - real_sorted[1]["total_votes"] if len(real_sorted) >= 2 else None
            con = Constituency(
                election_id=election.id, district_id=dist.id if dist else None,
                ac_no=ac_no,
                name=elec.get("name") or info["name"],
                category=elec.get("category") or "GEN",
                total_electors=total_electors,
                male_electors=elec.get("male_electors"),
                female_electors=elec.get("female_electors"),
                third_gender_electors=elec.get("third_gender_electors"),
                total_votes_polled=total_votes, turnout_pct=turnout, winning_margin=margin,
            )
            session.add(con)
            constituency_cache[ac_no] = con
        session.flush()
        print(f"Created {len(constituency_cache)} constituencies")

        party_cache = {p.abbr.upper(): p for p in session.query(Party).all()}
        imported = 0

        for ac_no, info in ac_data.items():
            con = constituency_cache[ac_no]
            real = [c for c in info["candidates"] if c["party"] not in ("NOTA", None) and c["name"] != "None of the Above"]
            nota = [c for c in info["candidates"] if c["party"] == "NOTA" or c["name"] == "None of the Above"]
            sorted_cands = sorted(real, key=lambda x: x["total_votes"], reverse=True) + nota
            total_valid = sum(c["total_votes"] for c in sorted_cands)

            for pos, cand in enumerate(sorted_cands, start=1):
                is_nota = cand["party"] == "NOTA" or cand["name"] == "None of the Above"
                name = "NOTA" if is_nota else cand["name"]

                party = None
                if not is_nota and cand["party"]:
                    abbr = cand["party"].strip()
                    if abbr.upper() not in party_cache:
                        p = Party(name=abbr, abbr=abbr, color=PARTY_COLORS.get(abbr))
                        session.add(p); session.flush()
                        party_cache[abbr.upper()] = p
                    party = party_cache[abbr.upper()]

                vote_pct = round(cand["total_votes"] / total_valid * 100, 2) if total_valid else 0
                prof = profile_map.get((ac_no, _normalize(name or "")))
                csv_info = prof.get("_csv") if prof else None
                cp = (prof or {}).get("candidate_profile", {})
                assets_s = (prof or {}).get("assets_summary", {})

                education = declared_assets = liabilities_val = image_url = None
                criminal_cases = 0

                if csv_info:
                    edu = csv_info.get("education", "").strip()
                    if edu and edu not in ("Not mentioned", "Not Given", ""):
                        education = edu
                    criminal_cases = csv_info.get("criminal_cases", 0)
                    a = parse_rupees(csv_info.get("assets"))
                    if a > 0:
                        declared_assets = a
                    l = parse_rupees(csv_info.get("liabilities"))
                    if l > 0:
                        liabilities_val = l
                elif cp:
                    edu = (cp.get("education") or "").strip()
                    if edu.startswith("Category:"):
                        edu = edu[len("Category:"):].strip()
                    if edu and edu not in ("Not mentioned", "Not Given"):
                        education = edu
                    criminal_cases = parse_criminal(cp.get("crime_status"))
                    a = parse_rupees(assets_s.get("total_assets"))
                    if a > 0:
                        declared_assets = a
                    l = parse_rupees(assets_s.get("total_liabilities"))
                    if l > 0:
                        liabilities_val = l

                if cp:
                    img = (cp.get("image_url") or "").strip()
                    if img and img != "None":
                        image_url = img

                session.add(Candidate(
                    election_id=election.id, constituency_id=con.id,
                    party_id=party.id if party else None, name=name,
                    gender=cand["sex"], age=cand["age"], position=pos,
                    votes_general=cand["gen_votes"], votes_postal=cand["postal_votes"],
                    votes_total=cand["total_votes"], vote_pct=vote_pct, is_nota=is_nota,
                    education=education, declared_assets=declared_assets,
                    liabilities=liabilities_val, criminal_cases=criminal_cases,
                    image_url=image_url,
                ))
                imported += 1

        session.commit()
        election_id = election.id

    print(f"\n=== Goa 2017 ===")
    print(f"Constituencies: {len(constituency_cache)}, Candidates: {imported}")
    with Session(engine) as session:
        winners = session.query(Candidate).filter_by(election_id=election_id, position=1, is_nota=False).all()
        seats = {}
        for w in winners:
            p = w.party.abbr if w.party else "IND"
            seats[p] = seats.get(p, 0) + 1
        for p, s in sorted(seats.items(), key=lambda x: -x[1]):
            print(f"  {p}: {s} seats")


if __name__ == "__main__":
    run()
